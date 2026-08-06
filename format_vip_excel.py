"""
Format VIP Phone Numbers Excel with professional styling
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

print("=" * 100)
print("FORMATTING VIP PHONE NUMBERS EXCEL")
print("=" * 100)

# Load the Excel file
input_file = 'VIP_Phone_Numbers_Complete_20260805_1003.xlsx'
output_file = f'VIP_Phone_Numbers_FORMATTED_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'

print(f"\n📂 Loading: {input_file}")

# Load data
df_all = pd.read_excel(input_file, sheet_name='All VIPs')
df_found = pd.read_excel(input_file, sheet_name='Found')
df_not_found = pd.read_excel(input_file, sheet_name='Not Found')

print(f"   All VIPs: {len(df_all)}")
print(f"   Found: {len(df_found)}")
print(f"   Not Found: {len(df_not_found)}")

# Create a new Excel with better formatting
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # Write all sheets
    df_all.to_excel(writer, sheet_name='All VIPs', index=False)
    df_found.to_excel(writer, sheet_name='Found (Phone Numbers)', index=False)
    df_not_found.to_excel(writer, sheet_name='Not Found', index=False)
    
    # Get workbook
    wb = writer.book
    
    # Define colors
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    found_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green
    not_found_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Light red
    
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    normal_font = Font(name='Arial', size=10)
    
    border_thin = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Format each sheet
    for sheet_name in ['All VIPs', 'Found (Phone Numbers)', 'Not Found']:
        ws = wb[sheet_name]
        
        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border_thin
        
        # Format data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border_thin
                
                # Color code based on status
                if cell.column == 1:  # Status column
                    if cell.value == '✅':
                        for c in row:
                            c.fill = found_fill
                    elif cell.value == '❌':
                        for c in row:
                            c.fill = not_found_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set header row height
        ws.row_dimensions[1].height = 30
        
        # Enable filters
        ws.auto_filter.ref = ws.dimensions
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    # Add summary sheet
    summary_data = {
        'Metric': [
            'Total VIP Customers',
            'Found with Phone Numbers',
            'Not Found',
            'Success Rate',
            '',
            'Generated Date',
            'Data Source',
            'Search Period'
        ],
        'Value': [
            len(df_all),
            len(df_found),
            len(df_not_found),
            f"{len(df_found)/len(df_all)*100:.1f}%",
            '',
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            'Latest orders from API/Cache',
            'Last 14 days'
        ]
    }
    
    df_summary = pd.DataFrame(summary_data)
    df_summary.to_excel(writer, sheet_name='Summary', index=False)
    
    # Format summary sheet
    ws_summary = wb['Summary']
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 40
    
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
    
    for row in ws_summary.iter_rows(min_row=2, max_row=ws_summary.max_row):
        for cell in row:
            cell.font = Font(name='Arial', size=11)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border_thin
    
    ws_summary.freeze_panes = 'A2'
    
    # Make summary sheet the first sheet
    wb._sheets.insert(0, wb._sheets.pop(wb.sheetnames.index('Summary')))

print(f"\n✅ FORMATTED EXCEL CREATED: {output_file}")
print("\n📊 Excel Structure:")
print("   Sheet 1: Summary - Overview of search results")
print("   Sheet 2: All VIPs - Complete list with status")
print("   Sheet 3: Found (Phone Numbers) - 72 VIPs with phones")
print("   Sheet 4: Not Found - 17 VIPs without phones")
print("\n✨ Features:")
print("   ✓ Color-coded rows (Green=Found, Red=Not Found)")
print("   ✓ Auto-filters on all sheets")
print("   ✓ Frozen header rows")
print("   ✓ Auto-adjusted column widths")
print("   ✓ Professional formatting")
print("=" * 100)
