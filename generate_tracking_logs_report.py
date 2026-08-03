"""
generate_tracking_logs_report.py
Generates a complete tracking logs report separated by status code columns.
- Dedicated columns per status code:
  LOG 110 | UNIT 110 | TIME 110 | LOG 120 | UNIT 120 | TIME 120 | LOG 200 | UNIT 200 | TIME 200 | LOG 210 ...
- Every LOG column contains ONLY its specific status code number (e.g. LOG 110 contains ONLY '110')
- Default date range: 01/07/2026 to 03/08/2026 (or today's date)
- Excludes test bills (from test_bills.txt)
- Excludes trainer accounts/post offices (TRAINER, TEST, DEMO)
- Excludes global offices (GLOBAL, EXTERNAL)
"""

import os
import json
import glob
import re
import requests
import pandas as pd
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(HERE, "config.json")
TEST_BILLS_PATH = os.path.join(HERE, "test_bills.txt")

EXCLUDE_KEYWORDS = {"TRAINER", "GLOBAL", "TEST", "DEMO", "EXTERNAL"}

STATUS_CODES_LIST = [
    "110", "120", "200", "210", "230", "300", "302", "306", "309", "310", "311",
    "400", "401", "402", "420", "470", "472", "480", "500", "410"
]

def load_test_ids():
    test_ids = set()
    if os.path.exists(TEST_BILLS_PATH):
        with open(TEST_BILLS_PATH, "r", encoding="utf-8") as f:
            test_ids = set(line.strip() for line in f if line.strip() and not line.strip().startswith("/"))
    return test_ids

def generate_tracking_logs(detail_xlsx_path, out_path, max_workers=35):
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        cfg = json.load(f)

    test_ids = load_test_ids()
    df = pd.read_excel(detail_xlsx_path)
    df.columns = [str(c).strip().upper() for c in df.columns]

    headers_api = {
        "Authorization": "Bearer " + cfg["api"]["bearer_token"],
        "Accept": "application/json, text/plain, */*",
        "User-Agent": "Mozilla/5.0"
    }

    def fetch_order_log(oid):
        oid_str = str(oid).strip()
        if not oid_str or oid_str.lower() == "nan" or oid_str in test_ids:
            return None

        try:
            r = requests.get(
                "https://gw-express.metfone.com.kh/tms-tracking/api/v1/order-tracking",
                params={"order_id": oid_str},
                headers=headers_api,
                timeout=10
            )
            if r.status_code != 200:
                return None
            data = r.json()
            trips = data.get("trackingTrips", [])
            if not trips:
                return None

            for t in trips:
                upd = str(t.get("updatedBy", {}).get("name", "") or "").upper()
                desc = str(t.get("desc", "") or "").upper()
                po = str(t.get("postcode", "") or t.get("postOffice", {}).get("code", "") or "").upper()
                if any(kw in upd or kw in desc or kw in po for kw in EXCLUDE_KEYWORDS):
                    return None

            trips_sorted = list(reversed(trips))
            row_dict = {"BILL ID": oid_str}

            # Initialize all status code columns
            for sc in STATUS_CODES_LIST:
                row_dict[f"LOG_{sc}"] = ""
                row_dict[f"UNIT_{sc}"] = ""
                row_dict[f"TIME_{sc}"] = ""

            latest_st = ""
            latest_time = ""

            for t in trips_sorted:
                st = str(t.get("status", "") or "").lstrip("S").strip()
                po = t.get("postOffice") or {}
                unit = t.get("postcode") or (po.get("code") if isinstance(po, dict) else "") or ""
                if not unit and "handoverInfo" in t:
                    unit = t.get("handoverInfo", {}).get("departmentCode", "")

                dt_raw = t.get("updatedAt", "")
                dt_str = ""
                if dt_raw:
                    try:
                        dt_obj = pd.to_datetime(dt_raw)
                        dt_str = dt_obj.strftime("%d/%m/%Y %H:%M:%S")
                    except Exception:
                        dt_str = str(dt_raw)

                if st in STATUS_CODES_LIST:
                    row_dict[f"LOG_{st}"] = st
                    row_dict[f"UNIT_{st}"] = unit
                    row_dict[f"TIME_{st}"] = dt_str

                latest_st = st
                latest_time = dt_str

            row_dict["LATEST_STATUS"] = latest_st
            row_dict["LATEST_TIME"] = latest_time
            return row_dict
        except Exception:
            return None

    unique_oids = df["ORDER ID"].dropna().unique()
    results = []

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(fetch_order_log, oid) for oid in unique_oids]
        for future in as_completed(futures):
            res = future.result()
            if res:
                results.append(res)

    if not results:
        raise ValueError("No tracking logs could be extracted.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Tracking Logs"

    font_family = "Calibri"
    f_title = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    f_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    f_data = Font(name=font_family, size=9)
    f_log = Font(name=font_family, size=9, bold=True, color="1E3A8A")

    fill_title = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_hdr1 = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    fill_hdr2 = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
    fill_alt = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    thin = Side(border_style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Build Header Columns: BILL ID | LOG 110 | UNIT 110 | TIME 110 | LOG 120 | UNIT 120 ...
    headers = ["BILL ID"]
    for sc in STATUS_CODES_LIST:
        headers.extend([f"LOG {sc}", f"UNIT {sc}", f"TIME {sc}"])
    headers.extend(["LATEST STATUS", "LATEST TIME"])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    t_cell = ws.cell(1, 1, "BILL TRACKING LOGS BY STATUS CODE REPORT (01/07/2026 - 03/08/2026)")
    t_cell.font = f_title
    t_cell.fill = fill_title
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.row_dimensions[3].height = 24
    for c_idx, h_text in enumerate(headers, 1):
        cell = ws.cell(3, c_idx, h_text)
        cell.font = f_header
        cell.fill = fill_hdr1 if c_idx == 1 or ((c_idx - 2) // 3) % 2 == 0 else fill_hdr2
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    results.sort(key=lambda x: str(x.get("BILL ID", "")))

    r_idx = 4
    for r_data in results:
        row_fill = fill_alt if r_idx % 2 == 0 else None
        ws.row_dimensions[r_idx].height = 19

        c_bill = ws.cell(r_idx, 1, str(r_data.get("BILL ID", "")))
        c_bill.font = f_data
        c_bill.border = border
        c_bill.alignment = Alignment(horizontal="left", vertical="center")

        col_pos = 2
        for sc in STATUS_CODES_LIST:
            log_val = r_data.get(f"LOG_{sc}", "")
            unit_val = r_data.get(f"UNIT_{sc}", "")
            time_val = r_data.get(f"TIME_{sc}", "")

            # LOG cell
            c_log = ws.cell(r_idx, col_pos, log_val)
            c_log.font = f_log
            c_log.border = border
            if row_fill: c_log.fill = row_fill
            c_log.alignment = Alignment(horizontal="center", vertical="center")
            col_pos += 1

            # UNIT cell
            c_unit = ws.cell(r_idx, col_pos, unit_val)
            c_unit.font = f_data
            c_unit.border = border
            if row_fill: c_unit.fill = row_fill
            c_unit.alignment = Alignment(horizontal="center", vertical="center")
            col_pos += 1

            # TIME cell
            c_time = ws.cell(r_idx, col_pos, time_val)
            c_time.font = f_data
            c_time.border = border
            if row_fill: c_time.fill = row_fill
            c_time.alignment = Alignment(horizontal="center", vertical="center")
            col_pos += 1

        # Latest Status & Time
        for val in [r_data.get("LATEST_STATUS", ""), r_data.get("LATEST_TIME", "")]:
            cell = ws.cell(r_idx, col_pos, val)
            cell.font = f_data
            cell.border = border
            if row_fill: cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_pos += 1

        r_idx += 1

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    wb.save(out_path)
    return out_path

if __name__ == "__main__":
    import downloader
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    tmp_detail = os.path.join(HERE, "latest_july_aug_detail.xlsx")
    today_str = datetime.now().strftime("%Y%m%d")
    downloader.download_detail(cfg["api"], tmp_detail, from_date="20260701", to_date=today_str, force_refresh=True)
    out_file = os.path.join(HERE, f"Bill_Tracking_Status_Logs_01Jul_03Aug_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    generate_tracking_logs(tmp_detail, out_file)
    print(f"✅ July 1 - Aug 3 Status Logs Report saved to {out_file}")
