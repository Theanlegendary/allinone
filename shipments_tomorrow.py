import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

def build_shipments_tomorrow_report(src_xlsx, out_xlsx, target_label="Zone 1"):
    """
    Builds 100% exact replica of user's SHIPMENTS TOMORROW REPORT template Excel file:
      - Sheet 1: SHIPMENTS TOMORROW REPORT (Left main table + Right Executive Summary table)
      - Sheet 2: base (Raw order dataset)
    """
    wb_src = openpyxl.load_workbook(src_xlsx, data_only=True)
    ws_src = wb_src.active

    headers = [str(ws_src.cell(1, c).value or "").strip() for c in range(1, ws_src.max_column + 1)]
    col_map = {h.upper(): idx for idx, h in enumerate(headers, 1)}

    ci_order   = col_map.get("ORDER ID", col_map.get("ORDER_NUMBER", 3))
    ci_origin_br = col_map.get("ACTION POST OFFICE", col_map.get("ORIGIN_BRANCH", 36))
    ci_origin_po = col_map.get("CURRENT POST OFFICE", col_map.get("ORIGIN_POST", 16))
    ci_dest_prov = col_map.get("DELIVERY PROVINCE", col_map.get("DESTINATION_BRANCH", 13))
    ci_dest_po   = col_map.get("DELIVERY POST OFFICE", col_map.get("DESTINATION_POST", 14))
    ci_created   = col_map.get("CREATED DATE", col_map.get("CREATED_AT", 2))
    ci_fee       = col_map.get("TOTAL FEE (USD)", col_map.get("TOTAL_AMOUNT (USD)", 20))
    ci_cod       = col_map.get("COD (USD)", 21)
    ci_weight    = col_map.get("WEIGHT (G)", col_map.get("ACTUAL_WEIGHT (G)", 8))
    ci_status    = col_map.get("CURRENT STATUS", col_map.get("STATUES", 24))
    ci_receiver  = col_map.get("RECEIVER", 5)

    base_rows = []
    r_idx = 1
    for r in range(2, ws_src.max_row + 1):
        order_id = str(ws_src.cell(r, ci_order).value or "").strip()
        if not order_id or order_id == "None":
            continue

        dest_prov = str(ws_src.cell(r, ci_dest_prov).value or "").strip().upper()
        dest_po   = str(ws_src.cell(r, ci_dest_po).value or "").strip().upper()
        orig_br   = str(ws_src.cell(r, ci_origin_br).value or "").strip().upper()
        orig_po   = str(ws_src.cell(r, ci_origin_po).value or "").strip().upper()
        created   = str(ws_src.cell(r, ci_created).value or "").strip()
        status    = str(ws_src.cell(r, ci_status).value or "").strip()
        receiver  = str(ws_src.cell(r, ci_receiver).value or "").strip()

        try:
            weight = float(ws_src.cell(r, ci_weight).value or 0)
        except (ValueError, TypeError):
            weight = 0.0

        try:
            fee = float(ws_src.cell(r, ci_fee).value or 0)
        except (ValueError, TypeError):
            fee = 0.0

        try:
            cod = float(ws_src.cell(r, ci_cod).value or 0)
        except (ValueError, TypeError):
            cod = 0.0

        # Filter active transit orders (Status 306, 309, 302, 310, 311)
        sc = status.split(" - ")[0].split()[0] if status else ""
        if sc not in ("306", "309", "302", "310", "311"):
            continue

        # Target branch/zone filter
        if target_label and target_label.upper() not in ("ALL", "TOTAL", "ZONE 1", "ZONE 2", "ZONE 3", "ZONE 4", "ZONE 5"):
            tgt = target_label.upper()
            branch_prefix = tgt[:3]
            if dest_prov != branch_prefix and dest_po != tgt:
                continue

        base_rows.append({
            "no": r_idx,
            "order_number": order_id,
            "customer": receiver[:30],
            "origin_branch": orig_br,
            "origin_post": orig_po,
            "destination_branch": dest_prov,
            "destination_post": dest_po,
            "created_at": created,
            "fee": fee,
            "cod": cod,
            "weight_g": weight,
            "status": status,
            "receiver": receiver,
            "district": "Battambang" if dest_prov == "BAT" else "General District",
            "zone": "Zone 3" if dest_prov in ("BAT", "SIE", "PUR") else "Zone 1"
        })
        r_idx += 1

    wb = openpyxl.Workbook()

    # Sheet 1: SHIPMENTS TOMORROW REPORT
    ws1 = wb.active
    ws1.title = "SHIPMENTS TOMORROW REPORT"
    ws1.views.sheetView[0].showGridLines = True

    # Exact colors matching template file
    fill_title_left  = PatternFill("solid", fgColor="1F4E78") # Dark Navy
    fill_title_right = PatternFill("solid", fgColor="31565F") # Dark Teal
    fill_hdr_left    = PatternFill("solid", fgColor="2F5597") # Medium Navy
    fill_hdr_right   = PatternFill("solid", fgColor="31565F") # Dark Teal
    fill_sum_tot     = PatternFill("solid", fgColor="D9E1F2") # Soft Blue Total
    fill_left_tot    = PatternFill("solid", fgColor="B4C6E7") # Periwinkle Blue Total

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    font_banner = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_hdr    = Font(name="Calibri", size=9,  bold=True, color="FFFFFF")
    font_data   = Font(name="Calibri", size=9,  color="000000")
    font_data_b = Font(name="Calibri", size=9,  bold=True, color="000000")
    font_tot    = Font(name="Calibri", size=9,  bold=True, color="000000")

    # Row 1: Title Banners
    stamp_date = datetime.now().strftime("%d.%m")
    target_clean = target_label.upper()
    title_left_txt = f"SHIPMENTS TOMORROW REPORT {stamp_date} (Báo cáo hàng đến {target_clean})"
    title_right_txt= f"EXECUTIVE SUMMARY ({target_clean} / {target_clean[:3]})"

    ws1.merge_cells("A1:H1")
    ws1.cell(1, 1, title_left_txt).font = font_banner
    ws1.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 9):
        ws1.cell(1, c).fill = fill_title_left

    ws1.merge_cells("J1:N1")
    ws1.cell(1, 10, title_right_txt).font = font_banner
    ws1.cell(1, 10).alignment = Alignment(horizontal="center", vertical="center")
    for c in range(10, 15):
        ws1.cell(1, c).fill = fill_title_right

    ws1.row_dimensions[1].height = 24

    # Row 2: Header Rows
    headers_left = [
        "DESTINATION\n(សាខា)",
        "District\n(ស្រុក/ខណ្ឌ)",
        "DESTINATION_POS\n(បូស្តិ៍គោលដៅ)",
        "ORDER_NUMBER\n(លេខវិក្កយបត្រ)",
        "Receiver\n(អ្នកទទួល)",
        "SUM ACTUAL_WEIGHT (G)\n(ទម្ងន់សរុប g)",
        "VAS\n(សេវា)",
        "VAS Description\n(ឈ្មោះសេវាបន្ថែម)"
    ]
    headers_right = [
        "ZONE\n(តំបន់)",
        "DESTINATION_BRANCH\n(សាខា)",
        "District\n(ស្រុក/ខណ្ឌ)",
        "Bill\n(ចំនួនប័ណ្ណ)",
        "SUM ACTUAL_WEIGHT (G)\n(ទម្ងន់សរុប g)"
    ]

    ws1.row_dimensions[2].height = 28
    for ci, h in enumerate(headers_left, 1):
        cell = ws1.cell(2, ci, h)
        cell.font = font_hdr
        cell.fill = fill_hdr_left
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for ci, h in enumerate(headers_right, 10):
        cell = ws1.cell(2, ci, h)
        cell.font = font_hdr
        cell.fill = fill_hdr_right
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Populate Left Data Rows
    summary_data = {}
    total_bills = 0
    total_weight = 0.0

    r_curr = 3
    for item in base_rows:
        ws1.row_dimensions[r_curr].height = 19
        vals = [
            item["destination_branch"],
            item["district"],
            item["destination_post"],
            item["order_number"],
            item["receiver"],
            item["weight_g"],
            "",
            ""
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws1.cell(r_curr, ci, val)
            cell.font = font_data_b if ci == 1 else font_data
            cell.border = thin_border
            if ci in (1, 2, 3, 4):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 5:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif ci == 6:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0"

        key = (item["zone"], item["destination_branch"], item["district"])
        summary_data.setdefault(key, {"bills": 0, "weight": 0.0})
        summary_data[key]["bills"] += 1
        summary_data[key]["weight"] += item["weight_g"]

        total_bills += 1
        total_weight += item["weight_g"]
        r_curr += 1

    # Left Grand Total Row
    ws1.row_dimensions[r_curr].height = 22
    ws1.merge_cells(start_row=r_curr, start_column=1, end_row=r_curr, end_column=5)
    gt_left = ws1.cell(r_curr, 1, "Grand Total / សរុប")
    gt_left.font = font_tot
    gt_left.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 6):
        cell = ws1.cell(r_curr, c)
        cell.fill = fill_left_tot
        cell.border = thin_border

    gt_w_cell = ws1.cell(r_curr, 6, total_weight)
    gt_w_cell.font = font_tot
    gt_w_cell.fill = fill_left_tot
    gt_w_cell.border = thin_border
    gt_w_cell.alignment = Alignment(horizontal="right", vertical="center")
    gt_w_cell.number_format = "#,##0"

    for c in (7, 8):
        cell = ws1.cell(r_curr, c)
        cell.fill = fill_left_tot
        cell.border = thin_border

    # Populate Executive Summary Table on Right
    r_sum = 3
    for (zone_str, br, dist), stats in sorted(summary_data.items()):
        ws1.row_dimensions[r_sum].height = 19
        s_vals = [
            zone_str,
            br,
            dist,
            stats["bills"],
            stats["weight"]
        ]
        for ci, val in enumerate(s_vals, 10):
            cell = ws1.cell(r_sum, ci, val)
            cell.font = font_data
            cell.border = thin_border
            if ci in (10, 11, 12):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci in (13, 14):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if ci == 14:
                    cell.number_format = "#,##0"
        r_sum += 1

    # Right Summary Total Row
    ws1.row_dimensions[r_sum].height = 22
    ws1.merge_cells(start_row=r_sum, start_column=10, end_row=r_sum, end_column=12)
    tot_label_cell = ws1.cell(r_sum, 10, f"{target_clean[:3]} Total")
    tot_label_cell.font = font_tot
    tot_label_cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(10, 13):
        cell = ws1.cell(r_sum, c)
        cell.fill = fill_sum_tot
        cell.border = thin_border

    tot_b_cell = ws1.cell(r_sum, 13, total_bills)
    tot_b_cell.font = font_tot
    tot_b_cell.fill = fill_sum_tot
    tot_b_cell.border = thin_border
    tot_b_cell.alignment = Alignment(horizontal="right", vertical="center")

    tot_w_cell = ws1.cell(r_sum, 14, total_weight)
    tot_w_cell.font = font_tot
    tot_w_cell.fill = fill_sum_tot
    tot_w_cell.border = thin_border
    tot_w_cell.alignment = Alignment(horizontal="right", vertical="center")
    tot_w_cell.number_format = "#,##0"

    # Auto-adjust Column Widths matching example file
    widths = [14, 18, 16, 16, 32, 22, 10, 24, 4, 12, 22, 18, 10, 22]
    for ci, w in enumerate(widths, 1):
        col_letter = openpyxl.utils.get_column_letter(ci)
        ws1.column_dimensions[col_letter].width = w

    # Sheet 2: base (Raw order dataset)
    ws2 = wb.create_sheet(title="base")
    ws2.views.sheetView[0].showGridLines = True

    base_headers = [
        "No", "ORDER_NUMBER", "CUSTOMER", "ORIGIN_BRANCH", "ORIGIN_POST",
        "DESTINATION_BRANCH", "DESTINATION_POST", "CREATED_BY", "CREATED_AT",
        "PAYMENT_TERM", "BASE_FEE (USD)", "VAS_FEE (USD)", "DISCOUNT (USD)",
        "TOTAL_AMOUNT (USD)", "COD (USD)", "ACTUAL_WEIGHT (G)", "SIZE (CM)",
        "BASE_SERVICE", "VAS_SERVICE", "VAS_KHMER", "CARGO_TYPE", "FROM_SOURCE",
        "ZONE ĐẾN", "Huyện đến", "statues", "Receiver"
    ]
    ws2.append(base_headers)
    ws2.row_dimensions[1].height = 24
    for c in range(1, len(base_headers) + 1):
        cell = ws2.cell(1, c)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for item in base_rows:
        row_data = [
            item["no"],
            item["order_number"],
            item["customer"],
            item["origin_branch"],
            item["origin_post"],
            item["destination_branch"],
            item["destination_post"],
            item["customer"],
            item["created_at"],
            "Sender",
            item["fee"],
            0.0,
            0.0,
            item["fee"],
            item["cod"],
            item["weight_g"],
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            item["district"],
            item["status"],
            item["receiver"]
        ]
        ws2.append(row_data)

    wb.save(out_xlsx)
    return total_bills, total_weight
