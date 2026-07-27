"""
pivot.py
Doc file Excel chi tiet don (export-detail) va dung bang pivot
"PENDING BILL CHECK" giong mau:
  - Loc: CURRENT STATUS thuoc danh sach pending_status_codes
         (+ loai test neu exclude_test)
  - Hang (rows): ZONE > CURRENT POST OFFICE > ORDER ID
  - Cot (columns): MONTH / DAY theo CREATED DATE
  - Gia tri: Count of ORDER ID
  - Co dong/cot Grand Total
"""

from datetime import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---- vi tri cot trong file nguon (0-based) ----
COL_CREATED_DATE    = 1   # CREATED DATE  (dd/mm/yyyy HH:MM:SS)
COL_ORDER_ID        = 2   # ORDER ID
COL_DELIVERY_PROV   = 12  # DELIVERY PROVINCE (e.g. MON, BAT, PNP, SIE)
COL_CURRENT_PO      = 15  # CURRENT POST OFFICE
COL_CURRENT_STATUS  = 23  # CURRENT STATUS  ("110 - Chua tiep nhan")
COL_SENDER          = 3
COL_RECEIVER        = 4
COL_ACTION_PO_HUB   = 35  # ACTION POST OFFICE at ORIGIN HUB (col 35)
COL_TOTAL_FEE       = 19  # TOTAL FEE (USD)
COL_COD             = 20  # COD (USD)


def _status_code(value):
    """Lay ma so dau chuoi trang thai: '110 - Chua tiep nhan' -> '110'."""
    if value is None:
        return ""
    s = str(value).strip()
    if " - " in s:
        s = s.split(" - ", 1)[0]
    return s.split()[0].strip() if s else ""


def _parse_day(value):
    """Tra ve (month, day) tu CREATED DATE. Ho tro dd/mm/yyyy ..."""
    if value is None:
        return None, None
    if isinstance(value, datetime):
        return value.month, value.day
    s = str(value).strip()
    date_part = s.split(" ")[0]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            dt = datetime.strptime(date_part, fmt)
            return dt.month, dt.day
        except ValueError:
            continue
    return None, None


def _zone_for(po_code, zone_cfg):
    if not po_code:
        return zone_cfg.get("default_zone", "Khac")
    po = str(po_code).strip()
    by_po = zone_cfg.get("by_post_office", {})
    if po in by_po:
        return by_po[po]
    by_prefix = zone_cfg.get("by_prefix", {})
    for prefix, zone in by_prefix.items():
        if po.upper().startswith(prefix.upper()):
            return zone
    return zone_cfg.get("default_zone", "Khac")


def _is_test_row(row, test_keywords):
    blob = " ".join(str(row[c] or "") for c in (COL_SENDER, COL_RECEIVER)).lower()
    return any(k.lower() in blob for k in test_keywords)


def read_source(path):
    """Doc file Excel nguon -> list rows (tuple), bo qua header."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(r)
    return rows[1:] if rows else []


def build_pivot(rows, pivot_cfg, zone_cfg):
    """
    Builds a standard pivot tree:
      zone -> po -> {order_id: (month, day)}
    """
    pending = set(str(c).strip() for c in pivot_cfg.get("pending_status_codes", []))
    exclude_test = pivot_cfg.get("exclude_test", False)
    test_keywords = pivot_cfg.get("test_keywords", ["test"])

    tree = defaultdict(lambda: defaultdict(dict))
    today = datetime.now().date()
    day_keys_seen = {(today.month, today.day)}

    for row in rows:
        if not row or row[COL_ORDER_ID] in (None, ""):
            continue
        if pending and _status_code(row[COL_CURRENT_STATUS]) not in pending:
            continue
        if exclude_test and _is_test_row(row, test_keywords):
            continue

        month, day = _parse_day(row[COL_CREATED_DATE])
        if day is None:
            continue
        po = str(row[COL_CURRENT_PO] or "").strip() or "(trong)"
        zone = _zone_for(po, zone_cfg)
        order_id = str(row[COL_ORDER_ID]).strip()

        key = (month, day)
        tree[zone][po][order_id] = key
        day_keys_seen.add(key)

    day_keys = sorted(day_keys_seen)
    month_val = day_keys[0][0] if day_keys else None
    return tree, day_keys, month_val


# ---- styling ----
_HDR_FILL  = PatternFill("solid", fgColor="1E293B")
_ZONE_FILL = PatternFill("solid", fgColor="EAEAEA")
_THIN   = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT   = Alignment(horizontal="left", vertical="center")
_RIGHT  = Alignment(horizontal="right", vertical="center")


def export_pivot(tree, day_keys, month_val, pivot_cfg, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = pivot_cfg.get("is_label", "Pivot Report")

    fn = "Segoe UI"
    hdr_font = Font(name=fn, size=10, bold=True, color="FFFFFF")
    zone_font = Font(name=fn, size=10, bold=True, color="0F172A")
    data_font = Font(name=fn, size=10, color="0F172A")
    tot_font = Font(name=fn, size=10, bold=True, color="EF4444")
    tot_fill = PatternFill("solid", fgColor="F1F5F9")

    # Header Row 1
    ws.cell(1, 1, "Zone / Post Office").font = hdr_font
    ws.cell(1, 1).fill = _HDR_FILL
    ws.cell(1, 1).border = _BORDER
    ws.cell(1, 1).alignment = _LEFT

    col_idx_map = {}
    for idx, dk in enumerate(day_keys):
        col_num = 2 + idx
        col_idx_map[dk] = col_num
        cell = ws.cell(1, col_num, f"{dk[1]:02d}")
        cell.font = hdr_font
        cell.fill = _HDR_FILL
        cell.border = _BORDER
        cell.alignment = _CENTER

    tot_col = 2 + len(day_keys)
    ws.cell(1, tot_col, "Grand Total").font = hdr_font
    ws.cell(1, tot_col).fill = _HDR_FILL
    ws.cell(1, tot_col).border = _BORDER
    ws.cell(1, tot_col).alignment = _CENTER

    ws.row_dimensions[1].height = 24

    r = 2
    grand_col_totals = defaultdict(int)
    grand_overall = 0

    for zone in sorted(tree.keys()):
        # Zone Header Row
        ws.cell(r, 1, zone).font = zone_font
        ws.cell(r, 1).fill = _ZONE_FILL
        ws.cell(r, 1).border = _BORDER
        for c in range(2, tot_col + 1):
            cell = ws.cell(r, c)
            cell.fill = _ZONE_FILL
            cell.border = _BORDER
        ws.row_dimensions[r].height = 20
        r += 1

        zone_col_totals = defaultdict(int)
        zone_total = 0

        for po in sorted(tree[zone].keys()):
            ws.cell(r, 1, f"   {po}").font = data_font
            ws.cell(r, 1).border = _BORDER
            ws.cell(r, 1).alignment = _LEFT

            po_sum = 0
            for oid, dk in tree[zone][po].items():
                if dk in col_idx_map:
                    cn = col_idx_map[dk]
                    cur = ws.cell(r, cn).value or 0
                    ws.cell(r, cn, cur + 1)
                    zone_col_totals[dk] += 1
                    grand_col_totals[dk] += 1
                    po_sum += 1

            for c in range(2, tot_col + 1):
                cell = ws.cell(r, c)
                cell.border = _BORDER
                cell.alignment = _CENTER
                cell.font = data_font

            ws.cell(r, tot_col, po_sum).font = data_font
            ws.cell(r, tot_col).border = _BORDER
            ws.cell(r, tot_col).alignment = _CENTER

            zone_total += po_sum
            ws.row_dimensions[r].height = 20
            r += 1

        grand_overall += zone_total

    # Grand Total Row
    ws.cell(r, 1, "Grand Total").font = tot_font
    ws.cell(r, 1).fill = tot_fill
    ws.cell(r, 1).border = _BORDER
    ws.cell(r, 1).alignment = _LEFT

    for dk, cn in col_idx_map.items():
        v = grand_col_totals.get(dk, 0)
        c = ws.cell(r, cn, v if v > 0 else "")
        c.font = tot_font
        c.fill = tot_fill
        c.border = _BORDER
        c.alignment = _CENTER

    c_tot = ws.cell(r, tot_col, grand_overall)
    c_tot.font = tot_font
    c_tot.fill = tot_fill
    c_tot.border = _BORDER
    c_tot.alignment = _CENTER

    ws.row_dimensions[r].height = 22

    # Column widths
    ws.column_dimensions["A"].width = 28
    for c in range(2, tot_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 7

    wb.save(out_path)
    return out_path, grand_overall


def _merge_pivot_cfg(global_pivot, report):
    cfg = dict(global_pivot)
    if "pending_status_codes" in report:
        cfg["pending_status_codes"] = report["pending_status_codes"]
    cfg["is_label"] = report.get("is_label", report.get("title", "Pending"))
    return cfg


def run_report(rows, out_path, config, report):
    """Dung 1 report tu rows da doc san. Tra ve (out_path, total)."""
    pivot_cfg = _merge_pivot_cfg(config.get("pivot", {}), report)
    tree, days, month = build_pivot(rows, pivot_cfg, config["zone_mapping"])
    return export_pivot(tree, days, month, pivot_cfg, out_path)


def build_mega_pivot(rows, pivot_cfg, zone_cfg):
    """
    Builds a 2-level pivot tree for MEGA check matching Metfone OPS format:
      Level 1: CURRENT POST OFFICE (MEGA1 / DVCMEGA1)
      Level 2: DELIVERY PROVINCE (MON, BAT, PNP, SIE, etc.)
    Tracks: order counts, total fee, total cod
    """
    exclude_test     = pivot_cfg.get("exclude_test", False)
    test_keywords    = pivot_cfg.get("test_keywords", ["test"])
    exclude_statuses = {"410", "201", "520", "99", "100", "-99"}

    # tree[hub][prov][(month, day)] = count
    tree = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    fee_tree = defaultdict(lambda: defaultdict(float))
    cod_tree = defaultdict(lambda: defaultdict(float))
    urgent_tree = defaultdict(lambda: defaultdict(int))
    day_keys_seen = set()
    today = datetime.now().date()
    day_keys_seen.add((today.month, today.day))

    for row in rows:
        if not row or len(row) <= COL_CURRENT_PO:
            continue
        if row[COL_ORDER_ID] in (None, ""):
            continue
        status_code = _status_code(row[COL_CURRENT_STATUS])
        if status_code in exclude_statuses:
            continue
        if exclude_test and _is_test_row(row, test_keywords):
            continue

        po = str(row[COL_CURRENT_PO] or "").strip().upper()
        col35 = str(row[COL_ACTION_PO_HUB] if len(row) > COL_ACTION_PO_HUB and row[COL_ACTION_PO_HUB] else "").strip().upper()

        if col35 == "MEGA1" or po == "MEGA1":
            hub_label = "MEGA1"
        elif "DVC" in col35 or "DVC" in po or "MEGA" in col35 or "MEGA" in po or "HUB" in po:
            hub_label = "DVCMEGA1"
        else:
            continue

        prov = str(row[COL_DELIVERY_PROV] if len(row) > COL_DELIVERY_PROV and row[COL_DELIVERY_PROV] else "").strip().upper() or "KHAC"

        month, day = _parse_day(row[COL_CREATED_DATE])
        if day is None:
            continue

        key = (month, day)
        tree[hub_label][prov][key] += 1
        day_keys_seen.add(key)

        # Total fee & COD
        try:
            fee_tree[hub_label][prov] += float(row[COL_TOTAL_FEE] or 0)
        except (ValueError, TypeError):
            pass
        try:
            cod_tree[hub_label][prov] += float(row[COL_COD] or 0)
        except (ValueError, TypeError):
            pass

        created_val  = row[COL_CREATED_DATE]
        created_date = None
        if isinstance(created_val, datetime):
            created_date = created_val.date()
        elif created_val:
            s = str(created_val).strip().split(" ")[0]
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
                try:
                    created_date = datetime.strptime(s, fmt).date()
                    break
                except ValueError:
                    continue
        if created_date and (today - created_date).days > 1:
            urgent_tree[hub_label][prov] += 1

    day_keys = sorted(day_keys_seen)
    return tree, day_keys, (fee_tree, cod_tree, urgent_tree)


def export_mega_pivot(tree, day_keys, out_path, extra_data=None):
    """
    Renders 2-level pivot table with colors, section separation, Fee & COD columns:
      Row 1: PENDING BILL (Title Banner)
      Row 2: Count of ORDER ID | Action day
      Row 3: CURRENT POST OFFICE | DELIVERY PROVINCE | dd/mm Date cols | Total Orders | Total Fees (USD) | Total COD (USD)
      Groups: MEGA1 (Navy) vs DVCMEGA1 (Forest Green) with visual gap separation
    """
    fee_tree = extra_data[0] if extra_data and len(extra_data) > 0 else defaultdict(lambda: defaultdict(float))
    cod_tree = extra_data[1] if extra_data and len(extra_data) > 1 else defaultdict(lambda: defaultdict(float))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PENDING BILL"

    # Color Tokens
    title_fill  = PatternFill("solid", fgColor="1F4E78")   # Dark Navy Blue
    title_font  = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")

    hdr_fill    = PatternFill("solid", fgColor="2F5597")   # Header Blue
    hdr_font    = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")

    m1_hdr_fill = PatternFill("solid", fgColor="3B3838")   # Dark Gray Banner for MEGA1
    m1_sub_fill = PatternFill("solid", fgColor="D9E1F2")   # Light Blue Subtotal
    m1_sub_font = Font(name="Segoe UI", size=10, bold=True, color="1F3864")

    dvc_hdr_fill = PatternFill("solid", fgColor="375623")  # Dark Green Banner for DVCMEGA1
    dvc_sub_fill = PatternFill("solid", fgColor="E2EFDA")  # Light Green Subtotal
    dvc_sub_font = Font(name="Segoe UI", size=10, bold=True, color="276A3C")

    gt_fill     = PatternFill("solid", fgColor="FFF2CC")   # Soft Gold for Grand Total
    gt_font     = Font(name="Segoe UI", size=10, bold=True, color="7F6000")
    fee_font    = Font(name="Segoe UI", size=10, color="196B24", bold=True)
    data_font   = Font(name="Segoe UI", size=10, color="000000")

    # 1. Title Row (Row 1)
    ws.cell(1, 1, "📋 PENDING BILL SUMMARY REPORT").font = title_font
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # 2. Row 2: Sub-headers
    ws.cell(2, 1, "Count of ORDER ID").font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    ws.cell(2, 3, "Action day").font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")

    for r in (1, 2, 3):
        ws.row_dimensions[r].height = 20

    # 3. Row 3: Column Headers
    ws.cell(3, 1, "CURRENT POST OFFICE")
    ws.cell(3, 2, "DELIVERY PROVINCE")

    col_idx_map = {}
    for idx, dk in enumerate(day_keys):
        col_num = 3 + idx
        col_idx_map[dk] = col_num
        ws.cell(3, col_num, f"{dk[1]:02d}/{dk[0]:02d}")

    tot_col = 3 + len(day_keys)
    fee_col = tot_col + 1
    cod_col = fee_col + 1

    ws.cell(3, tot_col, "Total Orders")
    ws.cell(3, fee_col, "Total Fees (USD)")
    ws.cell(3, cod_col, "Total COD (USD)")

    # Merge title row 1 & 2 across total width
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cod_col)

    # Format header row 2 & 3
    for c in range(1, cod_col + 1):
        c1 = ws.cell(1, c)
        c1.fill = title_fill

        c2 = ws.cell(2, c)
        c2.fill = title_fill
        c2.border = _BORDER

        c3 = ws.cell(3, c)
        c3.fill = hdr_fill
        c3.font = hdr_font
        c3.border = _BORDER
        if c >= 3:
            c3.alignment = _CENTER
        else:
            c3.alignment = _LEFT

    r = 4
    grand_col_totals = defaultdict(int)
    grand_overall_orders = 0
    grand_overall_fee = 0.0
    grand_overall_cod = 0.0

    hub_configs = [
        ("MEGA1",    m1_hdr_fill,  m1_sub_fill,  m1_sub_font),
        ("DVCMEGA1", dvc_hdr_fill, dvc_sub_fill, dvc_sub_font),
    ]

    for hub_idx, (hub, h_fill, s_fill, s_font) in enumerate(hub_configs):
        prov_dict = tree.get(hub, {})
        if not prov_dict:
            continue

        # Add visual separation gap row between sections if not first
        if hub_idx > 0:
            ws.row_dimensions[r].height = 10
            r += 1

        # Hub Section Header Banner (e.g. MEGA1 / DVCMEGA1)
        ws.row_dimensions[r].height = 22
        ws.cell(r, 1, f"▶ [{hub}] {hub} HUB").font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        ws.cell(r, 1).fill = h_fill
        ws.cell(r, 1).border = _BORDER
        ws.cell(r, 1).alignment = _LEFT

        for c in range(2, cod_col + 1):
            cell = ws.cell(r, c)
            cell.fill = h_fill
            cell.border = _BORDER

        r += 1

        hub_col_totals = defaultdict(int)
        hub_total_orders = 0
        hub_total_fee = 0.0
        hub_total_cod = 0.0

        # Province Sub-rows
        for prov in sorted(prov_dict.keys()):
            ws.row_dimensions[r].height = 19
            ws.cell(r, 1).border = _BORDER
            ws.cell(r, 2, prov).font = data_font
            ws.cell(r, 2).border = _BORDER
            ws.cell(r, 2).alignment = _LEFT

            row_sum = 0
            for idx, dk in enumerate(day_keys):
                col_num = col_idx_map[dk]
                val = prov_dict[prov].get(dk, 0)
                cell = ws.cell(r, col_num)
                cell.border = _BORDER
                cell.font = data_font
                cell.alignment = _CENTER
                if val > 0:
                    cell.value = val
                    row_sum += val
                    hub_col_totals[dk] += val
                    grand_col_totals[dk] += val

            p_fee = fee_tree[hub].get(prov, 0.0)
            p_cod = cod_tree[hub].get(prov, 0.0)

            ws.cell(r, tot_col, row_sum if row_sum > 0 else "").font = data_font
            ws.cell(r, tot_col).alignment = _CENTER
            ws.cell(r, tot_col).border = _BORDER

            f_cell = ws.cell(r, fee_col, round(p_fee, 2) if p_fee > 0 else "")
            f_cell.font = fee_font
            f_cell.alignment = _RIGHT
            f_cell.border = _BORDER
            if p_fee > 0:
                f_cell.number_format = "$#,##0.00"

            c_cell = ws.cell(r, cod_col, round(p_cod, 2) if p_cod > 0 else "")
            c_cell.font = fee_font
            c_cell.alignment = _RIGHT
            c_cell.border = _BORDER
            if p_cod > 0:
                c_cell.number_format = "$#,##0.00"

            hub_total_orders += row_sum
            hub_total_fee += p_fee
            hub_total_cod += p_cod
            r += 1

        # Hub Subtotal Row (e.g. MEGA1 Total / DVCMEGA1 Total)
        ws.row_dimensions[r].height = 21
        sub_label = f"∑ {hub} Total"
        ws.cell(r, 1, sub_label).font = s_font
        ws.cell(r, 1).fill = s_fill
        ws.cell(r, 1).border = _BORDER
        ws.cell(r, 1).alignment = _LEFT

        ws.cell(r, 2).fill = s_fill
        ws.cell(r, 2).border = _BORDER

        for idx, dk in enumerate(day_keys):
            col_num = col_idx_map[dk]
            val = hub_col_totals.get(dk, 0)
            cell = ws.cell(r, col_num)
            cell.fill = s_fill
            cell.font = s_font
            cell.border = _BORDER
            cell.alignment = _CENTER
            if val > 0:
                cell.value = val

        tot_h = ws.cell(r, tot_col, hub_total_orders)
        tot_h.fill = s_fill
        tot_h.font = s_font
        tot_h.border = _BORDER
        tot_h.alignment = _CENTER

        hf_cell = ws.cell(r, fee_col, round(hub_total_fee, 2) if hub_total_fee > 0 else 0)
        hf_cell.fill = s_fill
        hf_cell.font = fee_font
        hf_cell.alignment = _RIGHT
        hf_cell.border = _BORDER
        hf_cell.number_format = "$#,##0.00"

        hc_cell = ws.cell(r, cod_col, round(hub_total_cod, 2) if hub_total_cod > 0 else 0)
        hc_cell.fill = s_fill
        hc_cell.font = fee_font
        hc_cell.alignment = _RIGHT
        hc_cell.border = _BORDER
        hc_cell.number_format = "$#,##0.00"

        grand_overall_orders += hub_total_orders
        grand_overall_fee += hub_total_fee
        grand_overall_cod += hub_total_cod
        r += 1

    # Gap row before Grand Total
    ws.row_dimensions[r].height = 6
    r += 1

    # Grand Total Row
    ws.row_dimensions[r].height = 24
    ws.cell(r, 1, "🏆 GRAND TOTAL").font = gt_font
    ws.cell(r, 1).fill = gt_fill
    ws.cell(r, 1).border = _BORDER
    ws.cell(r, 1).alignment = _LEFT

    ws.cell(r, 2).fill = gt_fill
    ws.cell(r, 2).border = _BORDER

    for idx, dk in enumerate(day_keys):
        col_num = col_idx_map[dk]
        val = grand_col_totals.get(dk, 0)
        cell = ws.cell(r, col_num)
        cell.fill = gt_fill
        cell.font = gt_font
        cell.border = _BORDER
        cell.alignment = _CENTER
        if val > 0:
            cell.value = val

    gt_c = ws.cell(r, tot_col, grand_overall_orders)
    gt_c.fill = gt_fill
    gt_c.font = gt_font
    gt_c.border = _BORDER
    gt_c.alignment = _CENTER

    gf_c = ws.cell(r, fee_col, round(grand_overall_fee, 2))
    gf_c.fill = gt_fill
    gf_c.font = fee_font
    gf_c.alignment = _RIGHT
    gf_c.border = _BORDER
    gf_c.number_format = "$#,##0.00"

    gc_c = ws.cell(r, cod_col, round(grand_overall_cod, 2))
    gc_c.fill = gt_fill
    gc_c.font = fee_font
    gc_c.alignment = _RIGHT
    gc_c.border = _BORDER
    gc_c.number_format = "$#,##0.00"

    # Column Widths
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20
    for c in range(3, tot_col):
        ws.column_dimensions[get_column_letter(c)].width = 7
    ws.column_dimensions[get_column_letter(tot_col)].width = 14
    ws.column_dimensions[get_column_letter(fee_col)].width = 16
    ws.column_dimensions[get_column_letter(cod_col)].width = 16

    wb.save(out_path)
    return out_path, grand_overall_orders


def run_mega(source_path, out_path, config):
    rows = read_source(source_path)
    tree, day_keys, extra_data = build_mega_pivot(rows, config.get("pivot", {}), config.get("zone_mapping", {}))
    return export_mega_pivot(tree, day_keys, out_path, extra_data=extra_data)


def run_mega_combined(source_path, out_path, config):
    """Build combined Excel."""
    return run_mega(source_path, out_path, config)
