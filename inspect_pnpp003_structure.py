"""Inspect PNPP003 Excel structure in detail"""
import pandas as pd
from openpyxl import load_workbook

file = 'bo/Report_PNPP003_Branch_05_08_2026_000000.xlsx'

print("=" * 100)
print(f"INSPECTING: {file}")
print("=" * 100)

# Load with openpyxl to see actual cell values
wb = load_workbook(file, data_only=True)

for sheet_name in wb.sheetnames:
    print(f"\n📄 Sheet: {sheet_name}")
    ws = wb[sheet_name]
    
    print(f"   Dimensions: {ws.dimensions}")
    print(f"   Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    # Find header row (look for "Age" column)
    age_col_idx = None
    vip_col_idx = None
    receiver_col_idx = None
    
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
        for idx, cell in enumerate(row, 1):
            val = str(cell.value or '').strip()
            if val == 'Age' or 'Age' in val:
                age_col_idx = idx
                print(f"   Found Age column at index: {idx}")
            if val == 'VIP':
                vip_col_idx = idx
                print(f"   Found VIP column at index: {idx}")
            if 'RECEIVER' in val:
                receiver_col_idx = idx
                print(f"   Found RECEIVER column at index: {idx}")
    
    # Check some age values
    if age_col_idx:
        print(f"\n   Sample Age values (column {age_col_idx}):")
        count = 0
        yellow_count = 0
        red_count = 0
        green_count = 0
        
        for row_idx in range(5, min(25, ws.max_row + 1)):
            cell = ws.cell(row_idx, age_col_idx)
            val = str(cell.value or '').strip()
            if val and val != '':
                if '🟢' in val:
                    green_count += 1
                elif '🟡' in val:
                    yellow_count += 1
                    if count < 3:
                        print(f"      Row {row_idx}: {val} ← ❌ YELLOW!")
                elif '🔴' in val:
                    red_count += 1
                    if count < 3:
                        print(f"      Row {row_idx}: {val}")
                count += 1
        
        print(f"\n   Summary: 🟢 {green_count}  🟡 {yellow_count}  🔴 {red_count}")
    
    # Check VIP values
    if vip_col_idx:
        print(f"\n   Sample VIP values (column {vip_col_idx}):")
        vip_count = 0
        for row_idx in range(5, min(25, ws.max_row + 1)):
            cell = ws.cell(row_idx, vip_col_idx)
            val = str(cell.value or '').strip()
            if val == 'VIP':
                vip_count += 1
                # Also show receiver
                if receiver_col_idx:
                    receiver_val = ws.cell(row_idx, receiver_col_idx).value
                    print(f"      Row {row_idx}: VIP → Receiver: {receiver_val}")
        
        print(f"   Total VIPs found: {vip_count}")

print("\n" + "=" * 100)

wb.close()
