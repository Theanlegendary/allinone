"""
generate_summary.py
Builds a single summary image showing totals per handle per report type.
No order detail — just counts.
"""

import io
from datetime import datetime
from PIL import Image, ImageDraw, ImageFont

SUMMARY_HEADER_KHMER = {
    "HANDLE": "ប៉ុស្តិ៍ / ហាង",
    "Pickup": "ត្រូវយក",
    "Delivery": "ត្រូវដឹក",
    "Pending": "កំពុងរង់ចាំ",
    "TOTAL": "សរុប",
    "GRAND TOTAL": "សរុបទាំងអស់"
}

# ── Visual config ──────────────────────────────────────────────────────────────
SCALE       = 3  # High-definition scaling factor
FONT_SIZE   = 11 * SCALE
ROW_H       = 24 * SCALE
PAD_X       = 12 * SCALE
PAD_Y       = 6 * SCALE

# Colours
C_TITLE_BG  = (15,  23,  42)   # #0F172A dark navy
C_TITLE_FG  = (255, 255, 255)
C_HEADER_BG = (30,  41,  59)   # #1E293B dark slate
C_HEADER_FG = (255, 255, 255)
C_ROW_BG    = (255, 255, 255)
C_ROW_ALT   = (248, 250, 252)  # #F8FAFC light stripe
C_TOTAL_BG  = (241, 245, 249)  # #F1F5F9
C_TOTAL_FG  = (239, 68,  68)   # #EF4444 red
C_NUM_FG    = (239, 68,  68)   # red numbers
C_TEXT_FG   = (15,  23,  42)
C_BORDER    = (203, 213, 225)  # #CBD5E1

REPORT_TYPES = ['Pickup', 'Delivery', 'Pending']


def _load_font(size, bold=False):
    # Full system paths first (Windows), then short names as fallback
    _WIN_FONTS = "C:/Windows/Fonts"
    candidates_bold = [
        f"{_WIN_FONTS}/KhmerOSbattambang.ttf",
        f"{_WIN_FONTS}/KhmerOSsiemreap.ttf",
        f"{_WIN_FONTS}/KhmerOScontent.ttf",
        f"{_WIN_FONTS}/KhmerOS.ttf",
        f"{_WIN_FONTS}/KhmerOSsys.ttf",
        f"{_WIN_FONTS}/arialbd.ttf",
        f"{_WIN_FONTS}/DejaVuSans-Bold.ttf",
        'KhmerOSbattambang.ttf', 'KhmerOSsiemreap.ttf', 'KhmerOS.ttf',
        'arialbd.ttf', 'DejaVuSans-Bold.ttf',
    ]
    candidates_regular = [
        f"{_WIN_FONTS}/KhmerOSbattambang.ttf",
        f"{_WIN_FONTS}/KhmerOSsiemreap.ttf",
        f"{_WIN_FONTS}/KhmerOScontent.ttf",
        f"{_WIN_FONTS}/KhmerOS.ttf",
        f"{_WIN_FONTS}/KhmerOSsys.ttf",
        f"{_WIN_FONTS}/arial.ttf",
        f"{_WIN_FONTS}/DejaVuSans.ttf",
        'KhmerOSbattambang.ttf', 'KhmerOSsiemreap.ttf', 'KhmerOS.ttf',
        'arial.ttf', 'DejaVuSans.ttf',
    ]
    for name in (candidates_bold if bold else candidates_regular):
        try:
            return ImageFont.truetype(name, size)
        except Exception:
            pass
    return ImageFont.load_default()


def _text_w(draw, text, font):
    try:
        bb = draw.textbbox((0, 0), text, font=font)
        return bb[2] - bb[0]
    except Exception:
        return len(text) * (FONT_SIZE - 1 * SCALE)


def build_summary_image(handle_results: list, overall: dict, today: datetime = None) -> io.BytesIO:
    """
    handle_results: list of dicts with keys 'handle', 'handle_counts'
    overall: dict {'Pickup': n, 'Delivery': n, 'Pending': n}
    Returns BytesIO PNG.
    """
    now = today or datetime.now()
    n_data_rows = len(handle_results)
    
    # Dynamically adjust SCALE to keep the summary image size within Telegram's limits (under 4000px height)
    if n_data_rows > 75:
        scale = 1
    elif n_data_rows > 35:
        scale = 2
    else:
        scale = 3

    font_size   = 11 * scale
    row_h       = 24 * scale
    pad_x       = 12 * scale
    pad_y       = 6 * scale

    fn      = _load_font(font_size, bold=False)
    fn_bold = _load_font(font_size, bold=True)
    fn_sm   = _load_font(font_size - 1 * scale, bold=False)

    # ── Measure column widths ──────────────────────────────────────────────────
    tmp  = Image.new("RGB", (1, 1))
    draw = ImageDraw.Draw(tmp)

    col_headers = ["HANDLE"] + REPORT_TYPES + ["TOTAL"]

    # Handle col width based on max handle name
    handle_strs = [hr['handle'] for hr in handle_results] + ["GRAND TOTAL"]
    w_handle = max(_text_w(draw, s, fn_bold) for s in handle_strs) + pad_x * 2
    w_handle = max(w_handle, 90 * scale)

    # Number columns — fixed width
    w_num = max(_text_w(draw, h, fn_bold) for h in REPORT_TYPES + ["TOTAL"]) + pad_x * 2
    w_num = max(w_num, 72 * scale)

    col_widths = [w_handle] + [w_num] * (len(REPORT_TYPES) + 1)  # +1 for TOTAL
    total_w    = sum(col_widths) + 1

    # ── Row count ──────────────────────────────────────────────────────────────
    n_rows = 1 + 1 + n_data_rows + 1  # title + header + data + grand total
    total_h = n_rows * row_h + 1

    img  = Image.new("RGB", (total_w, total_h), (255, 255, 255))
    draw = ImageDraw.Draw(img)

    def draw_row(row_idx, cells, bg, fg_list=None, bold=False):
        y = row_idx * row_h
        x = 0
        f = fn_bold if bold else fn
        for ci, (cell_text, cw) in enumerate(zip(cells, col_widths)):
            fg = fg_list[ci] if fg_list else C_TEXT_FG
            draw.rectangle([x, y, x + cw, y + row_h], fill=bg)
            if cell_text:
                tw = _text_w(draw, cell_text, f)
                # Center numbers, left-align handle
                if ci == 0:
                    tx = x + pad_x
                else:
                    tx = x + (cw - tw) // 2
                ty = y + (row_h - font_size) // 2
                draw.text((tx, ty), cell_text, font=f, fill=fg)
            draw.rectangle([x, y, x + cw, y + row_h], outline=C_BORDER, width=1 * scale)
            x += cw

    # Title row
    title = f"📊 DAILY TOTAL  {now.strftime('%d/%m/%Y  %H:%M')}"
    draw.rectangle([0, 0, total_w, row_h], fill=C_TITLE_BG)
    f = fn_bold
    tw = _text_w(draw, title, f)
    draw.text((pad_x, (row_h - font_size) // 2), title, font=f, fill=C_TITLE_FG)
    draw.rectangle([0, 0, total_w, row_h], outline=C_BORDER, width=1 * scale)

    # Header row
    draw_row(1, col_headers, C_HEADER_BG,
             fg_list=[C_HEADER_FG] * len(col_headers), bold=True)

    # Data rows
    for i, hr in enumerate(handle_results):
        counts = hr['handle_counts']
        pickup   = counts.get('Pickup',   0)
        delivery = counts.get('Delivery', 0)
        pending  = counts.get('Pending',  0)
        total    = pickup + delivery + pending
        cells = [
            hr['handle'],
            str(pickup)   if pickup   else '',
            str(delivery) if delivery else '',
            str(pending)  if pending  else '',
            str(total),
        ]
        bg = C_ROW_ALT if i % 2 else C_ROW_BG
        fg = [C_TEXT_FG, C_NUM_FG, C_NUM_FG, C_NUM_FG, C_TOTAL_FG]
        draw_row(2 + i, cells, bg, fg_list=fg)

    # Grand Total row
    g_pickup   = overall.get('Pickup',   0)
    g_delivery = overall.get('Delivery', 0)
    g_pending  = overall.get('Pending',  0)
    g_total    = g_pickup + g_delivery + g_pending
    grand_cells = [
        "GRAND TOTAL",
        str(g_pickup),
        str(g_delivery),
        str(g_pending),
        str(g_total),
    ]
    draw_row(2 + n_data_rows, grand_cells, C_TOTAL_BG,
             fg_list=[C_TOTAL_FG] * len(grand_cells), bold=True)

    buf = io.BytesIO()
    img.save(buf, format='PNG', optimize=True)
    buf.seek(0)
    return buf

# ── Total Excel builder ────────────────────────────────────────────────────────

def build_total_excel(result: dict, out_path: str):
    """
    Build a summary Excel with 3 tables (Pickup / Delivery / Pending) on a SINGLE sheet.
    Each table = ALL branches for that type, sorted by POST OFFICE HANDLE.
    Uses raw DataFrames from result['type_data'] — no re-parsing of Excel files.
    day_cols are datetime.date objects (from generate_report.py).
    Layout per table:
      Row 1 : Title (dark bg, white text, merged)
      Row 2 : Month row  (groups of merged cells, one per month, styled dark)
      Row 3 : Header row (day numbers "01".."31" + index cols + Grand Total)
      Row 4+: Data rows
      Last  : Grand Total footer
    """
    import calendar
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    REPORT_ORDER = ['Pickup', 'Delivery', 'Pending']

    REPORT_COLS = {
        'Pickup':   ['ZONE', 'POST OFFICE HANDLE', 'CURRENT POST OFFICE', 'ORDER ID', 'Cus name', 'Phone'],
        'Delivery': ['ZONE', 'POST OFFICE HANDLE', 'CURRENT POST OFFICE', 'ORDER ID', 'RECEIVER', 'ACTION', 'NEXT_STEP'],
        'Pending':  ['ZONE', 'POST OFFICE HANDLE', 'CURRENT POST OFFICE', 'ORDER ID', 'ACTION', 'NEXT_STEP'],
    }

    type_data   = result.get('type_data', {})
    day_cols    = result.get('day_cols', [])   # list of datetime.date sorted ascending
    date_col    = result.get('cur_time_col') or 'CURRENT TIME'
    now_str     = datetime.now().strftime('%d.%m_%Hh%M')

    fn    = 'Segoe UI'
    RED   = 'EF4444'
    NAVY  = '0F172A'
    SLATE = '1E293B'
    thin  = Side(style='thin', color='BFBFBF')
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report_All"

    current_row = 1

    for rn in REPORT_ORDER:
        df = type_data.get(rn)
        if df is None:
            df = pd.DataFrame(columns=REPORT_COLS[rn])
        elif df.empty:
            df = df.copy()
            for col in REPORT_COLS[rn]:
                if col not in df.columns:
                    df[col] = ''

        idx_cols = REPORT_COLS[rn]
        for col in idx_cols:
            if col not in df.columns:
                df[col] = ''

        # Parse dates → datetime.date for matching with day_cols
        if date_col in df.columns:
            parsed = pd.to_datetime(df[date_col], dayfirst=True, format='mixed', errors='coerce')
            df = df.copy()
            df['_date'] = parsed.dt.date
        else:
            df = df.copy()
            df['_date'] = None

        dates_present = set(df['_date'].dropna().unique())
        active_days = [d for d in day_cols if d in dates_present]

        # Create day indicator columns
        for d in active_days:
            df[d] = (df['_date'] == d).astype(int)
        df['Grand Total'] = 1

        agg = df.groupby(idx_cols, sort=False, dropna=False)[
            active_days + ['Grand Total']
        ].sum().reset_index()

        for d in active_days:
            agg[d] = agg[d].apply(lambda v: int(v) if v > 0 else '')

        if rn in ('Delivery', 'Pending') and 'ACTION' in agg.columns:
            order_created_map = {}
            if 'ORDER ID' in df.columns and 'CREATED DATE' in df.columns:
                parsed_created = pd.to_datetime(df['CREATED DATE'], dayfirst=True, format='mixed', errors='coerce')
                for order_id, dt in zip(df['ORDER ID'].astype(str).str.strip(), parsed_created):
                    if pd.notna(dt):
                        order_created_map[order_id] = dt.date()
            agg = agg.copy()
            agg['_act_w'] = agg['ACTION'].apply(lambda x: 1 if 'ដឹកជញ្ជូន' in str(x) else (2 if 'ពិនិត្យ' in str(x) else (3 if 'ត្រឡប់' in str(x) else 4)))
            from datetime import date as _dt_date
            agg['_created_dt'] = agg['ORDER ID'].apply(lambda x: order_created_map.get(str(x).strip(), _dt_date.max))
            
            sort_by = []
            ascending = []
            if 'ZONE' in agg.columns:
                sort_by.append('ZONE')
                ascending.append(True)
            sort_by.extend(['_act_w', '_created_dt'])
            ascending.extend([True, True])
            
            agg = agg.sort_values(by=sort_by, ascending=ascending).reset_index(drop=True)
            agg = agg.drop(columns=['_act_w', '_created_dt'])
        else:
            sort_cols = [c for c in ['POST OFFICE HANDLE', 'CURRENT POST OFFICE', 'ORDER ID']
                         if c in agg.columns]
            agg = agg.sort_values(sort_cols).reset_index(drop=True)

        all_cols = idx_cols + active_days + ['Grand Total']
        n_idx    = len(idx_cols)
        n_total  = len(all_cols)

        # ── Row 1: Title ──────────────────────────────────────────────────────
        title_row = current_row
        ws.row_dimensions[title_row].height = 22
        tc = ws.cell(title_row, 1, f"{rn.upper()} BILL CHECK  {now_str}  — ALL BRANCHES")
        tc.font      = Font(name=fn, color='FFFFFF', bold=True, size=12)
        tc.fill      = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
        tc.alignment = Alignment(horizontal='left', vertical='center')
        tc.border    = bdr
        for ci in range(2, n_total + 1):
            c = ws.cell(title_row, ci, '')
            c.fill   = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
            c.border = bdr
        if n_total > 1:
            ws.merge_cells(start_row=title_row, end_row=title_row,
                           start_column=1, end_column=n_total)

        # ── Row 2: Month row + Row 3: Header row ─────────────────────────────
        month_row  = current_row + 1
        header_row = current_row + 2
        ws.row_dimensions[month_row].height  = 18
        ws.row_dimensions[header_row].height = 17

        # Pre-style both rows
        for ci in range(1, n_total + 1):
            for ri in (month_row, header_row):
                cell = ws.cell(ri, ci, '')
                cell.fill      = PatternFill(start_color=SLATE, end_color=SLATE, fill_type='solid')
                cell.font      = Font(name=fn, color='FFFFFF', bold=True, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border    = bdr

        # Index cols + Grand Total: vertical merge across month_row & header_row
        for ci, col_name in enumerate(all_cols, start=1):
            is_idx = (ci <= n_idx)
            is_gt  = (ci == n_total)
            if is_idx or is_gt:
                ws.cell(month_row, ci, col_name)
                ws.merge_cells(start_row=month_row, end_row=header_row,
                               start_column=ci, end_column=ci)

        # Day columns: day numbers in header_row
        for ci in range(n_idx + 1, n_total):   # day columns only
            d = all_cols[ci - 1]               # datetime.date object
            ws.cell(header_row, ci, f"{d.day:02d}")

        # Month groups: horizontal merge on month_row
        month_groups = []
        cur_month = None
        grp_start = None
        for ci in range(n_idx + 1, n_total):   # 1-based column index of day cols
            d = all_cols[ci - 1]
            m_val = (d.year, d.month)
            if m_val != cur_month:
                if cur_month is not None:
                    month_groups.append((cur_month, grp_start, ci - 1))
                cur_month = m_val
                grp_start = ci
        if cur_month is not None:
            month_groups.append((cur_month, grp_start, n_total - 1))

        for (yr, mo), start_c, end_c in month_groups:
            ws.cell(month_row, start_c, calendar.month_name[mo])
            if end_c > start_c:
                ws.merge_cells(start_row=month_row, end_row=month_row,
                               start_column=start_c, end_column=end_c)

        # ── Data rows ─────────────────────────────────────────────────────────
        order_created_map = {}
        if 'ORDER ID' in df.columns and 'CREATED DATE' in df.columns:
            parsed_created = pd.to_datetime(df['CREATED DATE'], dayfirst=True, format='mixed', errors='coerce')
            for order_id, dt in zip(df['ORDER ID'].astype(str).str.strip(), parsed_created):
                if pd.notna(dt):
                    order_created_map[order_id] = dt.date()

        day_totals  = {d: 0 for d in active_days}
        grand_total = 0
        data_start  = current_row + 3   # first data row

        for ri, row in agg.iterrows():
            r = data_start + ri
            ws.row_dimensions[r].height = 15
            gt_val = int(row.get('Grand Total', 0))
            grand_total += gt_val

            is_overdue = False
            if 'ORDER ID' in row:
                oid = str(row['ORDER ID']).strip()
                if oid in order_created_map:
                    delta = datetime.now().date() - order_created_map[oid]
                    if delta.days > 1:
                        is_overdue = True

            for ci, col in enumerate(all_cols, start=1):
                val  = row.get(col, '')
                cell = ws.cell(r, ci, val if val != '' else None)
                cell.border = bdr
                cell.font   = Font(name=fn, size=10)
                if col == 'ACTION':
                    act_str = str(val)
                    if 'ដឹកជញ្ជូន' in act_str:
                        cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                    elif 'ពិនិត្យ' in act_str:
                        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                    elif 'ត្រឡប់' in act_str:
                        cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                elif is_overdue:
                    cell.fill = PatternFill(start_color='FFEBEB', end_color='FFEBEB', fill_type='solid')
                if col in active_days:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if isinstance(val, (int, float)) and val:
                        day_totals[col] = day_totals.get(col, 0) + int(val)
                elif col == 'Grand Total':
                    cell.font      = Font(name=fn, color=RED, bold=True, size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        # ── Grand Total footer ────────────────────────────────────────────────
        gt_row = data_start + len(agg)
        ws.row_dimensions[gt_row].height = 17
        for ci, col in enumerate(all_cols, start=1):
            cell = ws.cell(gt_row, ci)
            cell.font      = Font(name=fn, color=RED, bold=True, size=10)
            cell.fill      = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
            cell.border    = bdr
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if ci == 1:
                cell.value = 'Grand Total'
            elif col in active_days:
                cell.value = day_totals.get(col) or None
            elif col == 'Grand Total':
                cell.value = grand_total or None

        # ── Column widths ─────────────────────────────────────────────────────
        for ci, col in enumerate(all_cols, start=1):
            letter = get_column_letter(ci)
            if col == 'Grand Total':
                ws.column_dimensions[letter].width = 20
            elif col == 'ZONE':
                ws.column_dimensions[letter].width = 9
            elif isinstance(col, __import__('datetime').date):
                ws.column_dimensions[letter].width = 7
            elif col in ('Cus name', 'RECEIVER'):
                max_len = max(
                    (len(str(ws.cell(r_iter, ci).value or ''))
                     for r_iter in range(current_row, gt_row + 1)),
                    default=20
                )
                existing = ws.column_dimensions[letter].width or 0
                ws.column_dimensions[letter].width = max(existing, min(max(max_len + 3, 22), 50))
            elif col == 'Phone':
                max_len = max(
                    (len(str(ws.cell(r_iter, ci).value or ''))
                     for r_iter in range(current_row, gt_row + 1)),
                    default=14
                )
                existing = ws.column_dimensions[letter].width or 0
                ws.column_dimensions[letter].width = max(existing, min(max(max_len + 3, 16), 35))
            else:
                max_len = max(
                    (len(str(ws.cell(r_iter, ci).value or ''))
                     for r_iter in range(current_row, gt_row + 1)),
                    default=8
                )
                existing = ws.column_dimensions[letter].width or 0
                ws.column_dimensions[letter].width = max(existing, min(max(max_len + 2, 10), 28))

        current_row = gt_row + 3

    wb.save(out_path)
    return out_path
