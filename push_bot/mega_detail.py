"""mega_detail.py - Detail Excel for orders currently at MEGA/DVC/HUB post offices.

Filter: CURRENT POST OFFICE contains 'MEGA', 'DVC', or 'HUB'
  (parcels physically stuck at MEGA hub right now)

Layout: 2 Separate Excel Sheets
  Sheet 1 — ⚠ Urgent (age >= 1 day for MEGA hub)
  Sheet 2 — ✓ Normal (age < 1 day)

Columns:
  1. មណ្ឌល        (Hub: MEGA1 / DVMEGA)
  2. ស្ថានភាព     (Status KM)
  3. បុគ្គលិក       (Action User - staff ID & name who handled parcel)
  4. ប.ស.ចេញ    (Origin PO  = RECEIVE POST OFFICE)
  5. ប.ស.ទៅ     (Dest PO    = DELIVERY POST OFFICE)
  6. លេខបញ្ជា    (ORDER ID)
  7. ថ្លៃ (USD)   (Total Fee USD)
  8. COD (USD)   (COD USD)
  9. កាលបរិច្ឆេទ   (Created Date)
  10. ថ្ងៃ       (Age days)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

STATUS_KM = {
    "110": "មិនទាន់បញ្ជូន",
    "120": "កំពុងប្រមូល",
    "200": "បានទទួល",
    "210": "កំពុងដឹក",
    "300": "កំពុងដឹក",
    "302": "បែងចែក",
    "306": "ដល់មណ្ឌល",
    "309": "ស្កែន",
    "310": "បានទទួល",
    "311": "បន្តដឹក",
    "400": "រង់ចាំ",
    "401": "ដឹកជូន",
    "402": "ដឹកជូន",
    "420": "រង់នៅហាង",
    "430": "ដឹកឡើងវិញ",
    "460": "ដឹកឡើងវិញ",
    "472": "រក្សាទុក",
    "480": "កែអាសយដ្ឋាន",
    "500": "ត្រឡប់",
    "520": "បានត្រឡប់",
    "540": "ត្រឡប់បរាជ័យ",
}

# Column indices in source Excel (0-based)
CI_CREATED     = 1   # CREATED DATE
CI_ORDER       = 2   # ORDER ID
CI_SENDER      = 3   # SENDER
CI_RECEIVER    = 4   # RECEIVER
CI_RECV_PO     = 11  # RECEIVE POST OFFICE  (origin)
CI_DELIV_PO    = 13  # DELIVERY POST OFFICE (destination)
CI_CURRENT_PO  = 15  # CURRENT POST OFFICE
CI_TOTAL_FEE   = 19  # TOTAL FEE (USD)
CI_COD         = 20  # COD (USD)
CI_STATUS      = 23  # CURRENT STATUS
CI_ACTION_USER = 25  # ACTION USER (Staff ID & Name)
CI_HUB_CODE    = 35  # ACTION POST OFFICE at ORIGIN HUB (MEGA1 / DVCMEGA1)


def _sc(val):
    if val is None:
        return ""
    s = str(val).strip()
    if " - " in s:
        s = s.split(" - ", 1)[0]
    return s.split()[0].strip() if s else ""


def _parse_created(val, dt_cls):
    if val is None:
        return None
    if isinstance(val, dt_cls):
        return val.date()
    s = str(val).strip().split(" ")[0]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return dt_cls.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _fmt_money(val):
    """Return float or '' for display."""
    try:
        f = float(val)
        return round(f, 2) if f != 0 else ""
    except Exception:
        return ""


def build_mega_detail(source_path, out_path, cfg):
    """Build detail Excel with 2 separate sheets (⚠ Urgent and ✓ Normal).
    Returns (total_orders, urgent_count).
    """
    from datetime import datetime as _dt

    exclude_statuses = {"410", "201", "520", "99", "100", "-99"}
    excl_test = cfg.get("pivot", {}).get("exclude_test", False)
    test_kw   = cfg.get("pivot", {}).get("test_keywords", ["test"])

    wb_src = openpyxl.load_workbook(source_path, data_only=True)
    ws_src = wb_src.active
    all_rows = list(ws_src.iter_rows(values_only=True))
    if not all_rows:
        return 0, 0
    data_rows = all_rows[1:]

    today = _dt.now().date()

    urgent_rows = []
    normal_rows = []

    for row in data_rows:
        if not row or len(row) <= CI_STATUS:
            continue
        if row[CI_ORDER] is None or str(row[CI_ORDER]).strip() == "":
            continue
        sc = _sc(row[CI_STATUS])
        if sc in exclude_statuses:
            continue
        if excl_test:
            blob = " ".join(str(row[c] or "") for c in (CI_SENDER, CI_RECEIVER)).lower()
            if any(k.lower() in blob for k in test_kw):
                continue

        # Filter: Middle-mile Hub orders (DVCMEGA1 and MEGA1)
        po = str(row[CI_CURRENT_PO] or "").strip().upper()
        col35 = str(row[CI_HUB_CODE] if len(row) > CI_HUB_CODE and row[CI_HUB_CODE] else "").strip().upper()

        if col35 == "MEGA1" or po == "MEGA1":
            hub_label = "MEGA1"
            if sc not in ("309", "306"):
                continue
        elif "DVC" in col35 or "DVC" in po or "MEGA" in col35 or "MEGA" in po or "HUB" in po:
            hub_label = "DVMEGA"
            if sc != "306":
                continue
        else:
            continue

        # Date parsing: Use Action Date (Col 24) or Created Date
        action_val = row[24] if len(row) > 24 and row[24] else row[CI_CREATED]
        action_date = _parse_created(action_val, _dt)
        if not action_date:
            action_date = _parse_created(row[CI_CREATED], _dt)

        age = (today - action_date).days if action_date else 0
        # MEGA SLA Rule: Hold >= 1 day at MEGA hub is Urgent
        is_urgent = age >= 1

        action_user = str(row[CI_ACTION_USER] or "").strip() if len(row) > CI_ACTION_USER else ""

        record = {
            "hub":         hub_label,
            "status_km":   STATUS_KM.get(sc, sc),
            "action_user": action_user,
            "origin":      str(row[CI_RECV_PO]  or "").strip(),
            "dest":        str(row[CI_DELIV_PO] or "").strip(),
            "order_id":    str(row[CI_ORDER]    or "").strip(),
            "fee":         _fmt_money(row[CI_TOTAL_FEE] if len(row) > CI_TOTAL_FEE else None),
            "cod":         _fmt_money(row[CI_COD]       if len(row) > CI_COD       else None),
            "created":     str(row[CI_CREATED]  or "").strip(),
            "age":         age,
        }

        if is_urgent:
            urgent_rows.append(record)
        else:
            normal_rows.append(record)

    # Sort: Hub (MEGA1 first, DVMEGA second), then age desc, then order_id
    _hub_rank = {"MEGA1": 0, "DVMEGA": 1}
    urgent_rows.sort(key=lambda r: (_hub_rank.get(r["hub"], 9), -r["age"], r["order_id"]))
    normal_rows.sort(key=lambda r: (_hub_rank.get(r["hub"], 9), -r["age"], r["order_id"]))

    # ── Build Excel with 2 Sheets ─────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # Sheet 1: Urgent
    ws_urg = wb.active
    ws_urg.title = "⚠ Urgent"

    # Sheet 2: Normal
    ws_norm = wb.create_sheet(title="✓ Normal")

    # Styles
    thin        = Side(style="thin", color="BFBFBF")
    bdr         = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr         = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align  = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    right_align = Alignment(horizontal="right",  vertical="center", wrap_text=False)

    hdr_fill   = PatternFill("solid", fgColor="1F4E78")
    urgent_sec = PatternFill("solid", fgColor="C00000")
    normal_sec = PatternFill("solid", fgColor="375623")
    urgent_row = PatternFill("solid", fgColor="FFEBEB")
    normal_bg  = PatternFill("solid", fgColor="FFFFFF")

    hdr_font  = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    sec_font  = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10, color="1F1F1F")
    id_font   = Font(name="Segoe UI", size=10, color="1F3864", bold=True)
    red_font  = Font(name="Segoe UI", size=10, color="C00000", bold=True)
    gray_font = Font(name="Segoe UI", size=10, color="595959")
    fee_font  = Font(name="Segoe UI", size=10, color="196B24")  # green for fee/cod

    HDR = [
        "មណ្ឌល",         # 1 — Hub (MEGA1 / DVMEGA)
        "ស្ថានភាព",      # 2 — Status KM
        "បុគ្គលិក",       # 3 — Action User (Staff ID & Name)
        "ប.ស.ចេញ",       # 4 — Origin PO
        "ប.ស.ទៅ",        # 5 — Dest PO
        "លេខបញ្ជា",      # 6 — Order ID
        "ថ្លៃ (USD)",    # 7 — Fee
        "COD (USD)",     # 8 — COD
        "កាលបរិច្ឆេទ",  # 9 — Created Date
        "ថ្ងៃ",          # 10 — Age
    ]
    COL_WIDTHS = [12, 26, 32, 12, 12, 16, 10, 10, 18, 6]
    NCOLS = len(HDR)

    def _populate_sheet(ws, rows_data, sec_label, sec_fill, is_urg):
        ws.views.sheetView[0].showGridLines = True
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
        c = ws.cell(1, 1, sec_label)
        c.fill = sec_fill
        c.font = sec_font
        c.alignment = left_align
        c.border = bdr
        ws.row_dimensions[1].height = 20

        ws.row_dimensions[2].height = 22
        for ci, h in enumerate(HDR, 1):
            cell = ws.cell(2, ci, h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = ctr
            cell.border = bdr

        curr = 3
        bg = urgent_row if is_urg else normal_bg
        for rec in rows_data:
            ws.row_dimensions[curr].height = 18
            vals = [
                rec["hub"],
                rec["status_km"],
                rec["action_user"],
                rec["origin"],
                rec["dest"],
                rec["order_id"],
                rec["fee"],
                rec["cod"],
                rec["created"],
                rec["age"],
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(curr, ci, val)
                cell.fill = bg
                cell.border = bdr

                if ci == 1:       # Hub
                    cell.font = data_font; cell.alignment = ctr
                elif ci == 2:     # Status
                    cell.font = data_font; cell.alignment = left_align
                elif ci == 3:     # Action User
                    cell.font = gray_font; cell.alignment = left_align
                elif ci in (4, 5):# Origin / Dest
                    cell.font = data_font; cell.alignment = ctr
                elif ci == 6:     # Order ID
                    cell.font = id_font; cell.alignment = ctr
                elif ci in (7, 8):# Fee / COD
                    cell.font = fee_font; cell.alignment = right_align
                    if val != "":
                        cell.number_format = "$#,##0.00"
                elif ci == 9:     # Created Date
                    cell.font = gray_font; cell.alignment = ctr
                elif ci == 10:    # Age
                    cell.font = red_font if is_urg else data_font
                    cell.alignment = ctr

            curr += 1

        for ci, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    _populate_sheet(ws_urg, urgent_rows, f"⚠ ប្រញាប់ (Urgent) — {len(urgent_rows)} records", urgent_sec, True)
    _populate_sheet(ws_norm, normal_rows, f"✓ ធម្មតា (Normal) — {len(normal_rows)} records", normal_sec, False)

    wb.save(out_path)
    total_orders = len(urgent_rows) + len(normal_rows)
    return total_orders, len(urgent_rows)
