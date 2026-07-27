"""
pivot.py
Doc file Excel chi tiet don (export-detail) va dung bang pivot
"PENDING BILL CHECK" giong mau:
  - Loc: CURRENT STATUS thuoc danh sach pending_status_codes
         (+ loai test neu exclude_test)
  - Hang (rows): ZONE > CURRENT POST OFFICE > ORDER ID
  - Cot (columns): MONTH / DAY theo CREATED DATE
  - Gia tri: Count of ORDER ID
  - Co dong/cot Grand Total
"""

from datetime import datetime
from collections import defaultdict, OrderedDict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---- vi tri cot trong file nguon (0-based) ----
COL_CREATED_DATE    = 1   # CREATED DATE  (dd/mm/yyyy HH:MM:SS)
COL_ORDER_ID        = 2   # ORDER ID
COL_CURRENT_PO      = 15  # CURRENT POST OFFICE
COL_CURRENT_STATUS  = 23  # CURRENT STATUS  ("110 - Chua tiep nhan")
COL_SENDER          = 3
COL_RECEIVER        = 4
COL_ACTION_PO_HUB   = 35  # ACTION POST OFFICE at ORIGIN HUB (col 35)
COL_TOTAL_FEE       = 19  # TOTAL FEE (USD)
COL_COD             = 20  # COD (USD)

# Hub origin code -> row label in MEGA report
# MEGA1 and DVCMEGA1 are the same physical hub; both map to "MEGA"
MEGA_HUB_LABELS = {
    "MEGA1":    "MEGA",
    "DVCMEGA1": "MEGA",
}


def _status_code(value):
    """Lay ma so dau chuoi trang thai: '110 - Chua tiep nhan' -> '110'."""
    if value is None:
        return ""
    s = str(value).strip()
    if " - " in s:
        s = s.split(" - ", 1)[0]
    return s.split()[0].strip() if s else ""


def _parse_day(value):
    """Tra ve (month, day) tu CREATED DATE. Ho tro dd/mm/yyyy ..."""
    if value is None:
        return None, None
    if isinstance(value, datetime):
        return value.month, value.day
    s = str(value).strip()
    date_part = s.split(" ")[0]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            dt = datetime.strptime(date_part, fmt)
            return dt.month, dt.day
        except ValueError:
            continue
    return None, None


def _zone_for(po_code, zone_cfg):
    if not po_code:
        return zone_cfg.get("default_zone", "Khac")
    po = str(po_code).strip()
    by_po = zone_cfg.get("by_post_office", {})
    if po in by_po:
        return by_po[po]
    by_prefix = zone_cfg.get("by_prefix", {})
    for prefix, zone in by_prefix.items():
        if po.upper().startswith(prefix.upper()):
            return zone
    return zone_cfg.get("default_zone", "Khac")


def _is_test_row(row, test_keywords):
    blob = " ".join(str(row[c] or "") for c in (COL_SENDER, COL_RECEIVER)).lower()
    return any(k.lower() in blob for k in test_keywords)


def read_source(path):
    """Doc file Excel nguon -> list rows (tuple), bo qua header."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    return rows[1:]  # bo header

def build_pivot(rows, pivot_cfg, zone_cfg):
    """
    Tra ve cau truc pivot:
      tree: dict[zone][po][order_id] = (month, day) tuple
      day_keys: sorted list of (month, day) tuples (chronological)
    """
    import calendar
    pending = set(str(c).strip() for c in pivot_cfg.get("pending_status_codes", []))
    exclude_test = pivot_cfg.get("exclude_test", False)
    test_keywords = pivot_cfg.get("test_keywords", ["test"])

    tree = defaultdict(lambda: defaultdict(dict))  # zone -> po -> {order_id: (month, day)}
    day_keys_seen = set()

    for row in rows:
        if not row or row[COL_ORDER_ID] in (None, ""):
            continue
        if pending and _status_code(row[COL_CURRENT_STATUS]) not in pending:
            continue
        if exclude_test and _is_test_row(row, test_keywords):
            continue

        month, day = _parse_day(row[COL_CREATED_DATE])
        if day is None:
            continue
        po = str(row[COL_CURRENT_PO] or "").strip() or "(trong)"
        zone = _zone_for(po, zone_cfg)
        order_id = str(row[COL_ORDER_ID]).strip()

        key = (month, day)
        tree[zone][po][order_id] = key
        day_keys_seen.add(key)

    # Sort chronologically: (month, day) tuples sort naturally
    day_keys = sorted(day_keys_seen)
    # For backward compat, also return the first month
    month_val = day_keys[0][0] if day_keys else None
    return tree, day_keys, month_val


# ---- styling ----
_HDR_FILL  = PatternFill("solid", fgColor="1E293B")   # dark slate for headers
_ZONE_FILL = PatternFill("solid", fgColor="EAEAEA")
_THIN   = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT   = Alignment(horizontal="left",   vertical="center")
_RED    = Font(color="C00000", bold=True)


def export_pivot(tree, day_keys, month, pivot_cfg, out_path):
    """Xuat pivot ra file Excel with multi-month merged headers."""
    import calendar as _cal
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (pivot_cfg.get("title", "PENDING BILL CHECK") or "REPORT")[:31]

    # day_keys is list of (month, day) tuples sorted chronologically
    # Build unique label per key: "DD" — but use the tuple as the key for matching
    n_days = len(day_keys)
    first_day_col = 4   # columns: A=ZONE, B=CURRENT POST OFFICE, C=ORDER ID, D..=days
    gt_col = first_day_col + n_days

    # --- filter rows at top (unchanged layout) ---
    ws.cell(1, 1, "Phan loai").font = Font(bold=True)
    ws.cell(1, 2, pivot_cfg.get("classification", "Post office"))
    ws.cell(2, 1, "Is test").font   = Font(bold=True)
    ws.cell(2, 2, "#N/A" if pivot_cfg.get("exclude_test") else "(tat ca)")
    ws.cell(3, 1, "is ton ket noi").font = Font(bold=True)
    ws.cell(3, 2, pivot_cfg.get("is_label", "Pending"))
    title_cell = ws.cell(4, 1, f"{pivot_cfg.get('title','PENDING BILL CHECK')}  "
                               f"{datetime.now():%d.%m_%HH%M}")
    title_cell.font = Font(bold=True, color="C00000", size=12)

    # --- Header rows ---
    r_count = 5   # "Count of ORDER ID" label row
    r_month = 6   # Month name row (merged per month group)
    r_head  = 7   # Day number header row

    ws.row_dimensions[r_month].height = 18
    ws.row_dimensions[r_head].height  = 17

    ws.cell(r_count, 1,  "Count of ORDER ID").font = Font(bold=True)
    ws.cell(r_count, gt_col, "Grand Total").font   = Font(bold=True)

    # Style helper for header cells
    def _hdr(r, c, val=None):
        cell = ws.cell(r, c, val)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = _HDR_FILL
        cell.alignment = _CENTER
        cell.border    = _BORDER
        return cell

    # Index column headers — vertical merge across month_row and head_row
    for label, col in (("ZONE", 1), ("CURRENT POST OFFICE", 2), ("ORDER ID", 3)):
        _hdr(r_month, col, label)
        _hdr(r_head,  col)
        ws.merge_cells(start_row=r_month, end_row=r_head,
                       start_column=col, end_column=col)

    # Grand Total header — vertical merge
    _hdr(r_month, gt_col, "Grand Total")
    _hdr(r_head,  gt_col)
    ws.merge_cells(start_row=r_month, end_row=r_head,
                   start_column=gt_col, end_column=gt_col)

    # Day columns: day numbers in r_head
    for i, (mo, dy) in enumerate(day_keys):
        col_i = first_day_col + i
        _hdr(r_month, col_i)       # pre-style month row cell
        _hdr(r_head,  col_i, f"{dy:02d}")

    # Month groups: horizontally merge r_month cells per month
    month_groups = []
    cur_mo = None
    grp_start_ci = None
    for i, (mo, dy) in enumerate(day_keys):
        col_i = first_day_col + i
        if mo != cur_mo:
            if cur_mo is not None:
                month_groups.append((cur_mo, grp_start_ci, col_i - 1))
            cur_mo = mo
            grp_start_ci = col_i
    if cur_mo is not None:
        month_groups.append((cur_mo, grp_start_ci, first_day_col + n_days - 1))

    for mo, start_c, end_c in month_groups:
        ws.cell(r_month, start_c, _cal.month_name[mo])
        ws.cell(r_month, start_c).font      = Font(bold=True, color="FFFFFF", size=10)
        ws.cell(r_month, start_c).fill      = _HDR_FILL
        ws.cell(r_month, start_c).alignment = _CENTER
        ws.cell(r_month, start_c).border    = _BORDER
        if end_c > start_c:
            ws.merge_cells(start_row=r_month, end_row=r_month,
                           start_column=start_c, end_column=end_c)

    # --- body ---
    r = r_head + 1
    col_totals = defaultdict(int)
    grand_total = 0

    for zone in sorted(tree.keys()):
        zone_written = False
        for po in sorted(tree[zone].keys()):
            po_written = False
            for order_id in sorted(tree[zone][po].keys()):
                order_key = tree[zone][po][order_id]   # (month, day) tuple
                zc = ws.cell(r, 1, zone if not zone_written else None)
                pc = ws.cell(r, 2, po   if not po_written   else None)
                oc = ws.cell(r, 3, order_id)
                zc.fill = _ZONE_FILL
                for c in (zc, pc, oc):
                    c.border = _BORDER
                oc.alignment = _LEFT
                # Mark the matching day column
                for i, dk in enumerate(day_keys):
                    cell = ws.cell(r, first_day_col + i)
                    cell.border    = _BORDER
                    cell.alignment = _CENTER
                    if dk == order_key:
                        cell.value = 1
                        cell.font  = _RED
                        col_totals[dk] += 1
                tot = ws.cell(r, gt_col, 1)
                tot.alignment = _CENTER
                tot.border    = _BORDER
                grand_total += 1
                zone_written = True
                po_written   = True
                r += 1

    # --- Grand Total row ---
    gt_row = r
    ws.cell(gt_row, 1, "Grand Total").font = Font(bold=True)
    ws.cell(gt_row, 1).fill = PatternFill("solid", fgColor="DCE6F1")
    for i, dk in enumerate(day_keys):
        c = ws.cell(gt_row, first_day_col + i, col_totals.get(dk, 0))
        c.font = Font(bold=True); c.alignment = _CENTER
        c.fill = PatternFill("solid", fgColor="DCE6F1"); c.border = _BORDER
    c = ws.cell(gt_row, gt_col, grand_total)
    c.font = Font(bold=True); c.alignment = _CENTER
    c.fill = PatternFill("solid", fgColor="DCE6F1"); c.border = _BORDER

    # --- column widths ---
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    for i in range(n_days):
        ws.column_dimensions[get_column_letter(first_day_col + i)].width = 7
    ws.column_dimensions[get_column_letter(gt_col)].width = 12
    ws.freeze_panes = ws.cell(r_head + 1, 4)

    wb.save(out_path)
    return out_path, grand_total


def run(source_path, out_path, config):
    rows = read_source(source_path)
    tree, day_keys, month = build_pivot(rows, config["pivot"], config["zone_mapping"])
    return export_pivot(tree, day_keys, month, config["pivot"], out_path)



def _merge_pivot_cfg(base_pivot, report):
    """Gop cau hinh pivot goc voi 1 report (ghi de status codes / title / label)."""
    cfg = dict(base_pivot or {})
    cfg["pending_status_codes"] = report.get("status_codes",
                                             cfg.get("pending_status_codes", []))
    cfg["title"] = report.get("title", cfg.get("title", "REPORT"))
    cfg["is_label"] = report.get("is_label", report.get("title", "Pending"))
    return cfg


def run_report(rows, out_path, config, report):
    """Dung 1 report tu rows da doc san. Tra ve (out_path, total)."""
    pivot_cfg = _merge_pivot_cfg(config.get("pivot", {}), report)
    tree, days, month = build_pivot(rows, pivot_cfg, config["zone_mapping"])
    return export_pivot(tree, days, month, pivot_cfg, out_path)


def build_mega_pivot(rows, pivot_cfg, zone_cfg):
    """
    Builds a pivot tree for MEGA check with 2 rows:
      - MEGA1  : parcels originating from MEGA1 hub or currently at MEGA1
      - DVMEGA : parcels originating from DVCMEGA1 hub or currently at DVC/HUB
    Only active parcels physically at a MEGA/DVC/HUB post office are counted.
    """
    exclude_test     = pivot_cfg.get("exclude_test", False)
    test_keywords    = pivot_cfg.get("test_keywords", ["test"])
    exclude_statuses = {"410", "201", "520", "99", "100", "-99"}

    tree         = defaultdict(lambda: defaultdict(int))
    urgent_tree  = defaultdict(int)
    day_keys_seen = set()
    today = datetime.now().date()
    day_keys_seen.add((today.month, today.day))   # always include today

    for row in rows:
        if not row or len(row) <= COL_CURRENT_PO:
            continue
        if row[COL_ORDER_ID] in (None, ""):
            continue
        status_code = _status_code(row[COL_CURRENT_STATUS])
        if status_code in exclude_statuses:
            continue
        if exclude_test and _is_test_row(row, test_keywords):
            continue

        # Filter: only parcels CURRENTLY at a hub (MEGA/DVC/HUB code)
        po = str(row[COL_CURRENT_PO] or "").strip().upper()
        if not ("MEGA" in po or "HUB" in po or "DVC" in po):
            continue

        # Classify into MEGA1 vs DVMEGA row
        col35 = str(row[COL_ACTION_PO_HUB] if len(row) > COL_ACTION_PO_HUB and row[COL_ACTION_PO_HUB] else "").strip().upper()
        if "MEGA1" in col35 or "MEGA1" in po:
            label = "MEGA1"
        else:
            label = "DVMEGA"

        month, day = _parse_day(row[COL_CREATED_DATE])
        if day is None:
            continue

        key = (month, day)
        tree[label][key] += 1
        day_keys_seen.add(key)

        # Check urgent: created > 1 day ago
        created_val  = row[COL_CREATED_DATE]
        created_date = None
        if isinstance(created_val, datetime):
            created_date = created_val.date()
        elif created_val:
            s = str(created_val).strip().split(" ")[0]
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
                try:
                    created_date = datetime.strptime(s, fmt).date()
                    break
                except ValueError:
                    continue
        if created_date and (today - created_date).days > 1:
            urgent_tree[label] += 1

    day_keys = sorted(day_keys_seen)
    return tree, day_keys, urgent_tree

def export_mega_pivot(tree, day_keys, out_path, urgent_tree=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TỒN MEGA CHECK"

    # Header style: light blue fill, bold black font, thin border, centered
    hdr_fill = PatternFill("solid", fgColor="BDD7EE")
    hdr_font = Font(name="Calibri", size=11, bold=True, color="000000")
    total_font = Font(name="Calibri", size=11, bold=True, color="000000")
    urgent_font = Font(name="Calibri", size=11, bold=True, color="FF0000")
    data_font = Font(name="Calibri", size=11, color="000000")
    
    # 1. Title Row (Row 1)
    title_time = datetime.now().strftime("%d.%m_%HH%M")
    title_text = f"TỒN MEGA CHECK {title_time}"
    
    ws.cell(1, 1, title_text)
    ws.cell(1, 1).font = Font(name="Calibri", size=11, bold=True, color="FF0000")
    ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    # 2. Row 2: "Count of ORDER ID", "Colu"
    ws.cell(2, 1, "Count of ORDER ID").font = Font(name="Calibri", size=11, bold=True)
    ws.cell(2, 2, "Colu").font = Font(name="Calibri", size=11, bold=True)
    
    for r in (2, 3):
        ws.row_dimensions[r].height = 20

    # 3. Row 3: "Row Labels", Day numbers, "Total", "Urgent"
    ws.cell(3, 1, "Row Labels")
    
    col_idx_map = {}
    for idx, dk in enumerate(day_keys):
        col_num = 2 + idx
        col_idx_map[dk] = col_num
        ws.cell(3, col_num, f"{dk[1]:02d}")
    
    total_col_num = 2 + len(day_keys)
    urgent_col_num = total_col_num + 1
    ws.cell(3, total_col_num, "Total")
    ws.cell(3, urgent_col_num, "Urgent")

    # Apply header formatting for Row 2 & Row 3
    for r in (2, 3):
        for c in range(1, urgent_col_num + 1):
            cell = ws.cell(r, c)
            cell.fill = hdr_fill
            cell.border = _BORDER
            if r == 2 and c > 2:
                pass
            else:
                cell.font = hdr_font
                if c > 1:
                    cell.alignment = _CENTER
                else:
                    cell.alignment = _LEFT
    # Urgent header in red
    ws.cell(3, urgent_col_num).font = urgent_font

    # 4. Data rows — always MEGA1 first, then DVMEGA
    r = 4
    col_totals = defaultdict(int)
    grand_total = 0
    total_urgent = 0

    for hub in ["DVMEGA", "MEGA1"]:
        ws.row_dimensions[r].height = 20
        ws.cell(r, 1, hub).font = data_font
        ws.cell(r, 1).border = _BORDER
        ws.cell(r, 1).alignment = _LEFT
        
        row_sum = 0
        for idx, dk in enumerate(day_keys):
            col_num = col_idx_map[dk]
            val = tree[hub].get(dk, 0)
            cell = ws.cell(r, col_num)
            cell.border = _BORDER
            cell.font = data_font
            cell.alignment = _CENTER
            if val > 0:
                cell.value = val
                row_sum += val
                col_totals[dk] += val
        
        # Row Total
        tot_cell = ws.cell(r, total_col_num, row_sum)
        tot_cell.font = data_font
        tot_cell.alignment = _CENTER
        tot_cell.border = _BORDER
        grand_total += row_sum

        # Row Urgent
        hub_urgent = urgent_tree.get(hub, 0) if urgent_tree else 0
        urg_cell = ws.cell(r, urgent_col_num)
        urg_cell.border = _BORDER
        urg_cell.alignment = _CENTER
        if hub_urgent > 0:
            urg_cell.value = hub_urgent
            urg_cell.font = urgent_font
        else:
            urg_cell.font = data_font
        total_urgent += hub_urgent
        r += 1

    # 5. Grand Total row
    gt_row = r
    ws.row_dimensions[gt_row].height = 20
    ws.cell(gt_row, 1, "Total").font = total_font
    ws.cell(gt_row, 1).fill = hdr_fill
    ws.cell(gt_row, 1).border = _BORDER
    ws.cell(gt_row, 1).alignment = _LEFT

    for idx, dk in enumerate(day_keys):
        col_num = col_idx_map[dk]
        val = col_totals.get(dk, 0)
        cell = ws.cell(gt_row, col_num)
        cell.fill = hdr_fill
        cell.font = total_font
        cell.border = _BORDER
        cell.alignment = _CENTER
        if val > 0:
            cell.value = val

    gt_cell = ws.cell(gt_row, total_col_num, grand_total)
    gt_cell.fill = hdr_fill
    gt_cell.font = total_font
    gt_cell.border = _BORDER
    gt_cell.alignment = _CENTER

    # Urgent total in red
    urg_gt = ws.cell(gt_row, urgent_col_num, total_urgent)
    urg_gt.fill = hdr_fill
    urg_gt.font = urgent_font
    urg_gt.border = _BORDER
    urg_gt.alignment = _CENTER

    # 6. Column widths
    ws.column_dimensions["A"].width = 16
    for c in range(2, urgent_col_num + 1):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = 7
    # Urgent column slightly wider
    ws.column_dimensions[get_column_letter(urgent_col_num)].width = 9
    
    wb.save(out_path)
    return out_path, grand_total


def run_mega(source_path, out_path, config):
    rows = read_source(source_path)
    tree, day_keys, urgent_tree = build_mega_pivot(rows, config.get("pivot", {}), config.get("zone_mapping", {}))
    return export_mega_pivot(tree, day_keys, out_path, urgent_tree=urgent_tree)




def run_mega_combined(source_path, out_path, config):
    """Build combined Excel: Zone pivot on top + MEGA pivot below."""
    rows = read_source(source_path)
    zone_cfg = config.get("zone_mapping", {})
    pivot_cfg = config.get("pivot", {})

    # Build zone pivot (all pending statuses by zone)
    zone_tree, zone_day_keys, zone_month = build_pivot(rows, pivot_cfg, zone_cfg)

    # Build mega pivot
    mega_tree, mega_day_keys, mega_urgent = build_mega_pivot(rows, pivot_cfg, zone_cfg)

    # Combined day keys
    all_day_keys = sorted(set(zone_day_keys + mega_day_keys))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MEGA + Zone"

    hdr_fill = PatternFill("solid", fgColor="BDD7EE")
    hdr_font = Font(name="Calibri", size=11, bold=True, color="000000")
    total_font = Font(name="Calibri", size=11, bold=True, color="000000")
    urgent_font = Font(name="Calibri", size=11, bold=True, color="FF0000")
    data_font = Font(name="Calibri", size=11, color="000000")
    title_font = Font(name="Calibri", size=11, bold=True, color="FF0000")

    # ── ZONE TABLE (top) ──────────────────────────────────────────────
    r = 1
    title_time = datetime.now().strftime("%d.%m_%HH%M")
    ws.cell(r, 1, f"BILLS {title_time}").font = title_font
    ws.row_dimensions[r].height = 20
    r += 1

    # Zone header row
    ws.cell(r, 1, "Zone").font = hdr_font
    ws.cell(r, 1).fill = hdr_fill
    ws.cell(r, 1).border = _BORDER
    col_idx_map = {}
    for idx, dk in enumerate(all_day_keys):
        col_num = 2 + idx
        col_idx_map[dk] = col_num
        cell = ws.cell(r, col_num, f"{dk[1]:02d}")
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = _BORDER
        cell.alignment = _CENTER
    total_col = 2 + len(all_day_keys)
    ws.cell(r, total_col, "Total").font = hdr_font
    ws.cell(r, total_col).fill = hdr_fill
    ws.cell(r, total_col).border = _BORDER
    ws.cell(r, total_col).alignment = _CENTER
    r += 1

    # Zone data rows
    zone_col_totals = defaultdict(int)
    zone_grand = 0
    for zone_name in sorted(zone_tree.keys()):
        ws.cell(r, 1, zone_name).font = data_font
        ws.cell(r, 1).border = _BORDER
        row_sum = 0
        for po in zone_tree[zone_name]:
            for oid, dk in zone_tree[zone_name][po].items():
                if dk in col_idx_map:
                    col_num = col_idx_map[dk]
                    cur = ws.cell(r, col_num).value or 0
                    ws.cell(r, col_num, cur + 1)
                    zone_col_totals[dk] += 1
                    row_sum += 1
        for c in range(2, total_col + 1):
            cell = ws.cell(r, c)
            cell.border = _BORDER
            cell.alignment = _CENTER
            cell.font = data_font
            if cell.value == 0 or cell.value is None:
                cell.value = None
        ws.cell(r, total_col, row_sum).font = data_font
        ws.cell(r, total_col).border = _BORDER
        ws.cell(r, total_col).alignment = _CENTER
        zone_grand += row_sum
        ws.row_dimensions[r].height = 20
        r += 1

    # Zone total row
    ws.cell(r, 1, "Total").font = total_font
    ws.cell(r, 1).fill = hdr_fill
    ws.cell(r, 1).border = _BORDER
    for dk, col_num in col_idx_map.items():
        val = zone_col_totals.get(dk, 0)
        cell = ws.cell(r, col_num)
        cell.value = val if val > 0 else None
        cell.font = total_font
        cell.fill = hdr_fill
        cell.border = _BORDER
        cell.alignment = _CENTER
    ws.cell(r, total_col, zone_grand).font = total_font
    ws.cell(r, total_col).fill = hdr_fill
    ws.cell(r, total_col).border = _BORDER
    ws.cell(r, total_col).alignment = _CENTER
    ws.row_dimensions[r].height = 20
    r += 2  # gap

    # ── MEGA TABLE (bottom) ───────────────────────────────────────────
    ws.cell(r, 1, f"TỒN MEGA CHECK {title_time}").font = title_font
    ws.row_dimensions[r].height = 20
    r += 1

    # Mega header
    ws.cell(r, 1, "Row Labels").font = hdr_font
    ws.cell(r, 1).fill = hdr_fill
    ws.cell(r, 1).border = _BORDER
    for dk, col_num in col_idx_map.items():
        cell = ws.cell(r, col_num, f"{dk[1]:02d}")
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = _BORDER
        cell.alignment = _CENTER
    ws.cell(r, total_col, "Total").font = hdr_font
    ws.cell(r, total_col).fill = hdr_fill
    ws.cell(r, total_col).border = _BORDER
    ws.cell(r, total_col).alignment = _CENTER
    urgent_col = total_col + 1
    ws.cell(r, urgent_col, "Urgent").font = urgent_font
    ws.cell(r, urgent_col).fill = hdr_fill
    ws.cell(r, urgent_col).border = _BORDER
    ws.cell(r, urgent_col).alignment = _CENTER
    r += 1

    # Mega data rows
    mega_col_totals = defaultdict(int)
    mega_grand = 0
    mega_urgent_total = 0
    for hub in sorted(mega_tree.keys()):
        ws.cell(r, 1, hub).font = data_font
        ws.cell(r, 1).border = _BORDER
        row_sum = 0
        for dk in all_day_keys:
            col_num = col_idx_map[dk]
            val = mega_tree[hub].get(dk, 0)
            cell = ws.cell(r, col_num)
            cell.border = _BORDER
            cell.alignment = _CENTER
            cell.font = data_font
            if val > 0:
                cell.value = val
                row_sum += val
                mega_col_totals[dk] += val
        ws.cell(r, total_col, row_sum).font = data_font
        ws.cell(r, total_col).border = _BORDER
        ws.cell(r, total_col).alignment = _CENTER
        # Urgent
        hub_urgent = mega_urgent.get(hub, 0)
        urg_cell = ws.cell(r, urgent_col)
        urg_cell.border = _BORDER
        urg_cell.alignment = _CENTER
        if hub_urgent > 0:
            urg_cell.value = hub_urgent
            urg_cell.font = urgent_font
        else:
            urg_cell.font = data_font
        mega_urgent_total += hub_urgent
        mega_grand += row_sum
        ws.row_dimensions[r].height = 20
        r += 1

    # Mega total row
    ws.cell(r, 1, "Total").font = total_font
    ws.cell(r, 1).fill = hdr_fill
    ws.cell(r, 1).border = _BORDER
    for dk, col_num in col_idx_map.items():
        val = mega_col_totals.get(dk, 0)
        cell = ws.cell(r, col_num)
        cell.value = val if val > 0 else None
        cell.font = total_font
        cell.fill = hdr_fill
        cell.border = _BORDER
        cell.alignment = _CENTER
    ws.cell(r, total_col, mega_grand).font = urgent_font
    ws.cell(r, total_col).fill = hdr_fill
    ws.cell(r, total_col).border = _BORDER
    ws.cell(r, total_col).alignment = _CENTER
    ws.cell(r, urgent_col, mega_urgent_total).font = urgent_font
    ws.cell(r, urgent_col).fill = hdr_fill
    ws.cell(r, urgent_col).border = _BORDER
    ws.cell(r, urgent_col).alignment = _CENTER

    # Column widths
    ws.column_dimensions["A"].width = 16
    for c in range(2, urgent_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 7
    ws.column_dimensions[get_column_letter(urgent_col)].width = 9

    wb.save(out_path)
    return out_path, mega_grand


if __name__ == "__main__":
    import sys, json
    cfg = json.load(open(sys.argv[3])) if len(sys.argv) > 3 else json.load(
        open("config.json"))
    out, total = run(sys.argv[1], sys.argv[2], cfg)
    print(f"Da xuat: {out}  (tong {total} don pending)")

