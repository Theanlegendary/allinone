# -*- coding: utf-8 -*-
"""
branch_today.py - Generate Branch Today Performance Summary Report
Tracks 2 key metrics per branch/district:
1. Picked Up from MEGA Today (Status 306)
2. Success Delivery Today (Status 410)
"""

import os
import openpyxl
import copy
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

POST_OFFICE_DISTRICT_MAP = {
    # Phnom Penh (PNP)
    'PNPA002': 'Boeng Keng Kang', 'PNPP001': 'Boeng Keng Kang', 'PNPP007': 'Boeng Keng Kang', 'PNPS007': 'Boeng Keng Kang',
    'PNPP005': 'Chbar Ampov', 'PNPP010': 'Chraoy Chongvar', 'PNPP011': 'Dangkao', 'PNPP014': 'Doun Penh',
    'PNPA016': 'Kamboul', 'PNPP012': 'Kamboul', 'PNPA029': 'Mean Chey', 'PNPA055': 'Mean Chey', 'PNPP002': 'Mean Chey', 'PNPP003': 'Mean Chey',
    'PNPA028': 'Pou Saen Chey', 'PNPP008': 'Pou Saen Chey', 'PNPP009': 'Pou Saen Chey', 'PNPP004': 'Preaek Pnov',
    'PNPA040': 'Saen Sokh', 'PNPP013': 'Saen Sokh', 'PNPA036': 'Tuol Kouk', 'PNPP006': 'Tuol Kouk',

    # Kandal (KAN)
    'KANA024': 'Kandal Stueng', 'KANA028': 'Kandal Stueng', 'KANA049': 'Kandal Stueng',
    'KANS003': 'Kaoh Thum', 'KANA031': 'Khsach Kandal', 'KANA012': 'Kien Svay', 'KANA013': 'Kien Svay',
    'KANA008': 'Leuk Daek', 'KANS004': 'Mukh Kampul', 'KANA019': 'Ponhea Lueu',
    'KANA007': "S'ang", 'KANA020': "S'ang", 'KANA026': "S'ang", 'KANA040': 'Sampov Pun',
    'KANA023': 'Ta Khmau', 'KANP001': 'Ta Khmau',

    # Prey Veng (PRE)
    'PREA024': 'Ba Phnum', 'PREA023': 'Peam Ro', 'PREA020': 'Preah Sdach', 'PREA029': 'Preah Sdach', 'PREA035': 'Preah Sdach',
    'PREP001': 'Prey Veng', 'PRES001': 'Prey Veng', 'PREA002': 'Pur Rieng', 'PREA028': 'Sithor Kandal',

    # Svay Rieng (SVA)
    'SVAP001': 'Svay Rieng', 'SVAS002': 'Bavet', 'SVAA001': 'Bavet', 'SVAA002': 'Bavet', 'SVAA003': 'Romeas Haek', 'SVAA004': 'Rumduol', 'SVAA005': 'Svay Chrum',

    # Battambang (BAT)
    'BATA003': 'Banan', 'BATA016': 'Banan',
    'BATA001': 'Battambang', 'BATA009': 'Battambang', 'BATA040': 'Battambang', 'BATA042': 'Battambang', 'BATP001': 'Battambang',
    'BATA011': 'Kamrieng', 'BATS007': 'Moung Ruessei', 'BATA017': 'Samlout',
    'BATA010': 'Sampov Lun', 'BATA028': 'Sampov Lun', 'BATA004': 'Sangkae', 'BATA008': 'Sangkae',
    'BATA023': 'Thma Koul', 'BATA025': 'Thma Koul',

    # Siem Reap (SIE)
    'SIEP001': 'Siem Reap', 'SIEA001': 'Angkor Chum', 'SIEA002': 'Angkor Thon', 'SIEA003': 'Banteay Srei', 'SIEA004': 'Chi Kraeng',

    # Sihanoukville (SIH)
    'SIHP001': 'Sihanoukville', 'SIHA001': 'Preah Sihanouk', 'SIHA002': 'Stung Hav', 'SIHA003': 'Kampong Seila'
}


def build_branch_today_report(src_excel, out_xlsx, target_label="ALL"):
    """Builds Branch Today Performance Report Excel file."""
    df = pd.read_excel(src_excel)

    # Standardize target label
    tgt = str(target_label).strip().upper()
    target_clean = tgt.replace(" ", "_")

    zone_by_prefix = {
        "KAN": "ZONE1", "PNP": "ZONE1", "PRE": "ZONE1", "SVA": "ZONE1",
        "KAM": "ZONE2", "KOH": "ZONE2", "SIH": "ZONE2", "SPE": "ZONE2", "TAK": "ZONE2",
        "BAN": "ZONE3", "BAT": "ZONE3", "CHH": "ZONE3", "PUR": "ZONE3",
        "ODD": "ZONE4", "PRH": "ZONE4", "SIE": "ZONE4", "THO": "ZONE4",
        "CHA": "ZONE5", "KRA": "ZONE5", "TBK": "ZONE5", "ROT": "ZONE5", "MON": "ZONE5", "STU": "ZONE5"
    }

    summary_data = {} # (zone_str, br, dist) -> {"picked_up": c1, "delivered": c2}
    total_picked_up = 0
    total_delivered = 0

    for idx, row in df.iterrows():
        dest_po = str(row.get("DELIVERY POST OFFICE", "")).strip().upper()
        dest_prov = str(row.get("DELIVERY PROVINCE", "")).strip().upper()
        curr_status = str(row.get("CURRENT STATUS", "")).strip().upper()

        if not dest_po or dest_po == "NAN":
            continue

        dest_br = dest_po[:3] if len(dest_po) >= 3 else dest_prov

        # Filtering target
        tgt_norm = tgt.replace(" ", "")
        if tgt_norm.startswith("ZONE"):
            target_zone_name = tgt_norm
            item_zone = zone_by_prefix.get(dest_prov, "ZONE1")
            if item_zone != target_zone_name:
                continue

        elif tgt not in ("ALL", "TOTAL", "MEGA"):
            branch_prefix = tgt[:3]
            if dest_prov != branch_prefix and dest_po != tgt:
                continue

        # District mapping
        if dest_po in POST_OFFICE_DISTRICT_MAP:
            dist_name = POST_OFFICE_DISTRICT_MAP[dest_po]
        else:
            if dest_br == 'SVA':
                dist_name = 'Svay Rieng'
            elif dest_br == 'PNP':
                dist_name = 'Phnom Penh'
            elif dest_br == 'KAN':
                dist_name = 'General District'
            elif dest_br == 'BAT':
                dist_name = 'Battambang'
            elif dest_br == 'PRE':
                dist_name = 'Prey Veng'
            elif dest_br == 'SIE':
                dist_name = 'Siem Reap'
            elif dest_br == 'SIH':
                dist_name = 'Sihanoukville'
            else:
                dist_name = dest_prov if dest_prov != dest_po else 'General District'

        zone_str = zone_by_prefix.get(dest_prov, "Zone 1").replace("ZONE", "Zone ")

        # Status check
        is_picked_up = "306" in curr_status or pd.notna(row.get("STATUS 306 AT STORE / AGENT FROM HUB (FIRST TIME)"))
        is_delivered = "410" in curr_status or "GIAO THÀNH CÔNG" in curr_status

        key = (zone_str, dest_br, dist_name)
        if key not in summary_data:
            summary_data[key] = {"picked_up": 0, "delivered": 0}

        if is_picked_up:
            summary_data[key]["picked_up"] += 1
            total_picked_up += 1
        if is_delivered:
            summary_data[key]["delivered"] += 1
            total_delivered += 1

    # Create Workbook
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "BRANCH TODAY PERFORMANCE"
    ws1.views.sheetView[0].showGridLines = True

    # Styling Tokens
    font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    font_hdr   = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_data  = Font(name="Segoe UI", size=10, color="0F172A")
    font_tot   = Font(name="Segoe UI", size=11, bold=True, color="0F172A")
    font_tot_red = Font(name="Segoe UI", size=11, bold=True, color="991B1B")

    fill_hdr      = PatternFill("solid", fgColor="0F766E")
    fill_sub      = PatternFill("solid", fgColor="E0F2FE")
    fill_tot      = PatternFill("solid", fgColor="CCFBF1")

    border_clean = Border(
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0")
    )
    border_tot_acc = Border(
        top=Side(style="thin", color="0F766E"),
        bottom=Side(style="double", color="0F766E"),
        left=Side(style="thin", color="0F766E"),
        right=Side(style="thin", color="0F766E")
    )

    # Title Banner (Row 1)
    ws1.row_dimensions[1].height = 35.0
    ws1.merge_cells("A1:E1")
    title_cell = ws1.cell(1, 1, f"BRANCH TODAY PERFORMANCE ({target_label})")
    title_cell.font = font_title
    title_cell.fill = fill_hdr
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Table Headers (Row 2)
    headers = [
        "ZONE\n(តំបន់)",
        "DESTINATION_BRANCH\n(សាខា)",
        "District\n(ស្រុក/ខណ្ឌ)",
        "Picked Up from MEGA\n(ទទួលពី MEGA)",
        "Success Delivery Today\n(ដឹកជោគជ័យថ្ងៃនេះ)"
    ]
    ws1.row_dimensions[2].height = 35.0
    for c_idx, h in enumerate(headers, 1):
        cell = ws1.cell(2, c_idx, h)
        cell.font = font_hdr
        cell.fill = fill_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data & Subtotal Rows
    r_curr = 3
    branch_groups = {}
    for (zone_str, br, dist), stats in sorted(summary_data.items()):
        if br not in branch_groups:
            branch_groups[br] = []
        branch_groups[br].append((zone_str, br, dist, stats))

    for br in sorted(branch_groups.keys()):
        br_items = branch_groups[br]
        br_picked = 0
        br_deliv  = 0

        for zone_str, b_code, dist, stats in br_items:
            ws1.row_dimensions[r_curr].height = 20.0
            row_vals = [zone_str, b_code, dist, stats["picked_up"], stats["delivered"]]
            for ci, val in enumerate(row_vals, 1):
                cell = ws1.cell(r_curr, ci, val)
                cell.font = font_data
                cell.border = border_clean
                if ci in (1, 2, 3):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "#,##0"

            br_picked += stats["picked_up"]
            br_deliv  += stats["delivered"]
            r_curr += 1

        # Branch Subtotal Row
        ws1.row_dimensions[r_curr].height = 22.0
        ws1.merge_cells(start_row=r_curr, start_column=1, end_row=r_curr, end_column=3)
        lbl_c = ws1.cell(r_curr, 1, f"{br} Total")
        lbl_c.font = font_tot
        lbl_c.alignment = Alignment(horizontal="left", vertical="center")

        for c in range(1, 4):
            cell = ws1.cell(r_curr, c)
            cell.fill = fill_sub
            cell.border = border_clean

        p_c = ws1.cell(r_curr, 4, br_picked)
        p_c.font = font_tot
        p_c.fill = fill_sub
        p_c.border = border_clean
        p_c.alignment = Alignment(horizontal="right", vertical="center")
        p_c.number_format = "#,##0"

        d_c = ws1.cell(r_curr, 5, br_deliv)
        d_c.font = font_tot_red
        d_c.fill = fill_sub
        d_c.border = border_clean
        d_c.alignment = Alignment(horizontal="right", vertical="center")
        d_c.number_format = "#,##0"

        r_curr += 1

    # Grand Total Row
    ws1.row_dimensions[r_curr].height = 25.0
    ws1.merge_cells(start_row=r_curr, start_column=1, end_row=r_curr, end_column=3)
    gt_lbl = ws1.cell(r_curr, 1, f"{target_clean[:3]} Total")
    gt_lbl.font = font_tot
    gt_lbl.alignment = Alignment(horizontal="left", vertical="center")

    for c in range(1, 4):
        cell = ws1.cell(r_curr, c)
        cell.fill = fill_tot
        cell.border = border_tot_acc

    gt_p = ws1.cell(r_curr, 4, total_picked_up)
    gt_p.font = font_tot
    gt_p.fill = fill_tot
    gt_p.border = border_tot_acc
    gt_p.alignment = Alignment(horizontal="right", vertical="center")
    gt_p.number_format = "#,##0"

    gt_d = ws1.cell(r_curr, 5, total_delivered)
    gt_d.font = font_tot_red
    gt_d.fill = fill_tot
    gt_d.border = border_tot_acc
    gt_d.alignment = Alignment(horizontal="right", vertical="center")
    gt_d.number_format = "#,##0"

    # Column Widths
    col_widths = [14, 22, 18, 22, 24]
    for ci, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    wb.save(out_xlsx)
    return total_picked_up, total_delivered


def render_today_summary_image(out_xlsx):
    """Renders pixel-perfect PNG image of Today Performance Summary table."""
    import excel_to_image
    return excel_to_image.excel_to_image(out_xlsx)
