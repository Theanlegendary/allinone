import pandas as pd
import os
import json
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ── Khmer Helpers ─────────────────────────────────────────────────────────────
HEADER_KHMER_MAP = {
    'ZONE': 'តំបន់',
    'POST OFFICE HANDLE': 'ប៉ុស្តិ៍ទទួលខុសត្រូវ',
    'CURRENT POST OFFICE': 'ប៉ុស្តិ៍បច្ចុប្បន្ន',
    'ORDER ID': 'លេខបុង / លេខកូដបញ្ជាទិញ',
    'Cus name': 'ឈ្មោះអតិថិជន',
    'Phone': 'លេខទូរស័ព្ទ',
    'RECEIVER': 'ឈ្មោះអតិថិជន',
    'ACTION': 'សកម្មភាព',
    'NEXT_STEP': 'ជំហានបន្ទាប់',
    'REMARK': 'សកម្មភាព ត្រូវធ្វើ',
    'NEXT_ACTION': 'អ្នកគ្រប់គ្រង់ប្រតិបត្តិការ',
    'Grand Total': 'សរុប',
    'Pending': 'អីវ៉ាន់កំពុងរង់ចាំ (Pending)',
    'Pickup': 'អីវ៉ាន់ត្រូវយក (Pickup)',
    'Delivery': 'អីវ៉ាន់ត្រូវដឹក (Delivery)',
    'Transit': 'ឡានដឹកធំ (Mega Truck)',
    'Branch': 'ចាត់តាំងដឹក (Assign Deliver)',
    'TOTAL FEE (USD)': 'ថ្លៃដឹក (USD)',
    'COD (USD)': 'COD (USD)',
    'STATUS_CODE': 'Status',
}

def translate_header(col_name):
    return HEADER_KHMER_MAP.get(str(col_name).strip(), col_name)

def get_khmer_month_name(mo):
    months = {
        1: 'មករា', 2: 'កុម្ភៈ', 3: 'មីនា', 4: 'មេសា', 5: 'ឧសភា', 6: 'មិថុនា',
        7: 'កក្កដា', 8: 'សីហា', 9: 'កញ្ញា', 10: 'តុលា', 11: 'វិច្ឆិកា', 12: 'ធ្នូ'
    }
    return months.get(mo, "")

# ── Status codes ───────────────────────────────────────────────────────────────
CLASSIFY_LABEL = {'Pickup': 'Pickup', 'Delivery': 'Delivery', 'Pending': 'Pending', 'Transit': 'Transit', 'Branch': 'Branch'}

# All active report tabs in order
ALL_TABS = ['Pickup', 'Delivery', 'Transit', 'Branch']

# Max index cols across all report types (for stacked/long alignment)
MAX_INDEX = 8  # Delivery has most with Fee+COD

REPORT_COLS = {
    'Pickup':   ['ZONE', 'POST OFFICE HANDLE', 'CURRENT POST OFFICE', 'ORDER ID', 'STATUS_CODE', 'Cus name', 'Phone'],
    'Delivery': ['ZONE', 'POST OFFICE HANDLE', 'CURRENT POST OFFICE', 'ORDER ID', 'RECEIVER', 'STATUS_CODE', 'ACTION', 'NEXT_STEP', 'TOTAL FEE (USD)', 'COD (USD)'],
    'Transit':  ['ZONE', 'POST OFFICE HANDLE', 'CURRENT POST OFFICE', 'ORDER ID', 'STATUS_CODE', 'ACTION', 'NEXT_STEP'],
    'Branch':   ['ZONE', 'POST OFFICE HANDLE', 'CURRENT POST OFFICE', 'ORDER ID', 'RECEIVER', 'STATUS_CODE', 'ACTION', 'NEXT_STEP', 'TOTAL FEE (USD)', 'COD (USD)'],
}

REPORT_FILTER_COLS = {
    'Pickup':   'POST OFFICE HANDLE',
    'Delivery': 'POST OFFICE HANDLE',
    'Transit':  'POST OFFICE HANDLE',
    'Branch':   'POST OFFICE HANDLE',
}

GAP_COLS = 1  # gap between side-by-side tables in wide mode

# Fixed column widths (Excel units)
W_DAY   = 7.0   # day columns e.g. "01","02"
W_ZONE  = 9.0   # ZONE
W_GT    = 20.0  # Grand Total
W_MIN   = 8.0
W_MAX   = 38.0

# Highlight threshold: rows older than this many days get highlighted
HIGHLIGHT_OVER_DAYS = 1
HIGHLIGHT_COLOR = 'FFEBEB'

PENDING_REMARK_MAP = {
    '306': 'ដឹកជញ្ជូន (Deliver)',
    '309': 'ដឹកជញ្ជូន (Deliver)',
    '311': 'ដឹកជញ្ជូន (Deliver)',
    '210': 'ត្រួតពិនិត្យ (Check)',
    '300': 'ត្រួតពិនិត្យ (Check)',
    '302': 'ត្រួតពិនិត្យ (Check)',
    '310': 'ត្រួតពិនិត្យ (Check)',
    '500': 'ផ្ញើត្រឡប់ (Return)',
}

DELIVERY_ACTION_MAP = {
    '400': ('ដឹកជញ្ជូន', 'ចាត់អ្នកដឹក'),
    '401': ('ដឹកជញ្ជូន', 'ដឹកជូនថ្ងៃនេះ'),
    '402': ('ដឹកជញ្ជូន', 'ដឹកជូនឡើងវិញ'),
    '420': ('ដឹកជញ្ជូន', 'ជូនដំណឹងភ្ញៀវមកទទួល'),
    '430': ('ដឹកជញ្ជូន', 'ទាក់ទងអ្នកទទួល'),
    '460': ('ត្រឡប់', 'បញ្ជូនត្រឡប់ទៅហាងផ្ញើ'),
    '470': ('ត្រឡប់', 'បន្តដំណើរការត្រឡប់'),
    '471': ('ពិនិត្យ', 'ពិនិត្យព័ត៌មាន'),
    '472': ('ពិនិត្យ', 'ដោះស្រាយបញ្ហា'),
    '480': ('ពិនិត្យ', 'បញ្ជាក់អាសយដ្ឋានថ្មី'),
    '500': ('ត្រឡប់', 'ផ្ញើត្រឡប់ទៅហាងផ្ញើ'),
}

TRANSIT_ACTION_MAP = {
    '210': ('ពិនិត្យ', 'ផ្ទេរទៅឡានដឹក'),
    '230': ('ពិនិត្យ', 'ចាកចេញពីស្ថានីយ៍រង'),
    '300': ('ពិនិត្យ', 'ទទួលការផ្ទេរ'),
    '302': ('ពិនិត្យ', 'ទទួលការផ្ទេរ'),
    '306': ('ពិនិត្យ', 'បញ្ជូនទៅឃ្លាំងធំ'),
    '311': ('ពិនិត្យ', 'ទទួលការបញ្ជូន'),
}

BRANCH_ACTION_MAP = {
    '309': ('ដឹកជញ្ជូន', 'ចាត់អ្នកដឹក'),
    '310': ('ដឹកជញ្ជូន', 'ចាត់អ្នកដឹក'),
}



def get_next_action(status_code):
    sc = str(status_code).strip()
    if sc in ('306', '309', '311'):
        return 'ដឹកជញ្ជូន (Deliver)'
    elif sc in ('300', '302', '210', '310'):
        return 'ត្រួតពិនិត្យ (Check)'
    elif sc in ('500',):
        return 'ផ្ញើត្រឡប់ (Return)'
    return ''


def get_responsible_party(status_code, cur_po, recv_po):
    sc = str(status_code).strip()
    cur_po_upper = str(cur_po).strip().upper()
    recv_po_upper = str(recv_po).strip().upper()
    
    if sc in ('110', '120', '200', '311', '201'):
        return 'ហាងផ្ញើ / Sender Store'
    elif sc == '302':
        if cur_po_upper == recv_po_upper:
            return 'ហាងផ្ញើ / Sender Store'
        else:
            return 'ហាងទទួល / Receiver Store'
    elif sc in ('210', '300'):
        return 'អ្នកបើកបរដឹកជញ្ជូន / Transit Driver'
    elif sc == '306':
        if any(x in cur_po_upper for x in ('MEGA', 'HUB', 'DVC')):
            return 'មជ្ឈមណ្ឌលចែកចាយ / Transit Hub'
        else:
            return 'ហាងទទួល / Receiver Store'
    elif sc in ('310', '309', '400', '420', '472', '480'):
        return 'ហាងទទួល / Receiver Store'
    elif sc in ('401', '402', '430', '460'):
        return 'អ្នកដឹកជញ្ជូន / Delivery Rider'
    elif sc in ('500', '510', '511', '512', '520', '540'):
        return 'ហាងផ្ញើ (ត្រឡប់វិញ) / Return Sender Store'
    return 'Unknown'


# ── Helpers ────────────────────────────────────────────────────────────────────

def normalize_code(v):
    return '' if pd.isna(v) else str(v).strip().upper()

def clean_text(v):
    if pd.isna(v):
        return ''
    s = str(v).strip()
    return '' if s.lower() == '(blank)' else s

def normalize_id(v):
    if pd.isna(v):
        return ''
    s = str(v).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s

def reference_column(df, keys):
    return next((c for c in df.columns if any(k in c for k in keys)), None)

def load_test_order_ids(cfg):
    test_ids = set()
    
    # Try loading from text file
    txt_path = "test_bills.txt"
    if os.path.exists(txt_path):
        try:
            with open(txt_path, "r", encoding="utf-8") as f:
                for line in f:
                    val = line.strip()
                    if val:
                        test_ids.add(val)
        except Exception as e:
            print(f"Error reading test_bills.txt: {e}")

    # Try loading from delayed bills JSON
    delay_path = "delayed_bills.json"
    if os.path.exists(delay_path):
        try:
            with open(delay_path, "r", encoding="utf-8") as f:
                delayed = json.load(f)
            
            today = datetime.now().date()
            updated_delayed = {}
            for bill_id, exp_date_str in delayed.items():
                try:
                    exp_date = datetime.strptime(exp_date_str, "%Y-%m-%d").date()
                    if today < exp_date:
                        test_ids.add(str(bill_id))
                        updated_delayed[bill_id] = exp_date_str
                except Exception:
                    pass
            
            # Prune expired bills
            if len(updated_delayed) != len(delayed):
                with open(delay_path, "w", encoding="utf-8") as f:
                    json.dump(updated_delayed, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error reading delayed_bills.json: {e}")

    tr_cfg = cfg.get("test_receipts", {})
    if not tr_cfg.get("enabled"):
        return test_ids

    path = tr_cfg.get("path")
    if not path or not os.path.exists(path):
        return test_ids

    if path.lower().endswith((".xlsx", ".xls")):
        df_test = pd.read_excel(path, dtype=str)
    else:
        df_test = pd.read_csv(path, dtype=str, keep_default_na=False)
    df_test = df_test.fillna("")

    order_col = next(
        (
            c for c in df_test.columns
            if "order" in str(c).lower()
            or "phi" in str(c).lower()
            or "shipment" in str(c).lower()
        ),
        df_test.columns[0] if len(df_test.columns) else None,
    )
    if not order_col:
        return test_ids

    excel_ids = {
        str(v).strip()
        for v in df_test[order_col].tolist()
        if str(v).strip() and str(v).strip().lower() != "nan"
    }
    
    test_ids.update(excel_ids)
    return test_ids

def build_simple_reference(df_ref):
    key_col      = reference_column(df_ref, ['current_post_office', 'dealer', 'code'])
    handle_col   = reference_column(df_ref, ['post_office_handle', 'responsible', 'handle'])
    customer_col = reference_column(df_ref, ['customer_name', 'agent', 'showroom'])
    phone_col    = reference_column(df_ref, ['phone'])
    if not key_col:
        raise ValueError("Reference file missing current_post_office column.")
    ref = pd.DataFrame({
        'current_post_office': df_ref[key_col].apply(normalize_code),
        'post_office_handle':  df_ref[handle_col].apply(normalize_code) if handle_col else '',
        'customer_name':       df_ref[customer_col].apply(clean_text)   if customer_col else '',
        'phone':               df_ref[phone_col].apply(clean_text)       if phone_col else '',
    })
    ref = ref[ref['current_post_office'] != ''].copy()
    ref['_s'] = ref.apply(lambda r: sum(1 for v in r if clean_text(v)), axis=1)
    ref = ref.sort_values('_s', ascending=False).drop_duplicates('current_post_office', keep='first').drop(columns=['_s'])
    ref['_ref_post_office_key'] = ref['current_post_office']
    return ref

def get_zone(po, zone_mapping):
    po = str(po).strip().upper()
    if po in zone_mapping.get('by_post_office', {}):
        return zone_mapping['by_post_office'][po]
    for prefix, zone in zone_mapping.get('by_prefix', {}).items():
        if po.startswith(prefix.upper()):
            return zone
    return zone_mapping.get('default_zone', 'Zone?')

def load_reference(ref_path):
    if ref_path.lower().endswith('.csv'):
        df = pd.read_csv(ref_path, dtype=str, keep_default_na=False)
    else:
        df = pd.read_excel(ref_path, dtype=str)
    df.columns = [str(c).strip().lower() for c in df.columns]
    df = df.dropna(how='all').fillna('').replace(r'^\s*\(blank\)\s*$', '', regex=True)
    return build_simple_reference(df)


# ── Row builder ────────────────────────────────────────────────────────────────

def build_section_rows(df_h, index_cols, day_cols, date_col):
    """
    Returns (rows, total, active_day_cols).
    active_day_cols = only days that have >=1 order in this section.
    """
    df = df_h.copy()
    for col in index_cols:
        if col not in df.columns:
            df[col] = ''

    if date_col and date_col in df.columns:
        parsed = pd.to_datetime(df[date_col], dayfirst=True, format='mixed', errors='coerce')
        df['_date'] = parsed.dt.date
    else:
        df['_date'] = None

    dates_present = set(df['_date'].dropna().unique())
    active_days = [d for d in day_cols if d in dates_present]

    if not active_days:
        footer = {col: '' for col in index_cols}
        footer[index_cols[0]] = 'Grand Total'
        footer['Grand Total'] = 0
        return [footer], 0, []

    for d in active_days:
        df[d] = (df['_date'] == d).astype(int)
    df['Grand Total'] = 1

    agg = df.groupby(index_cols, sort=False, dropna=False)[active_days + ['Grand Total']].sum().reset_index()
    for d in active_days:
        agg[d] = agg[d].apply(lambda v: int(v) if v > 0 else '')

    total = int(agg['Grand Total'].sum())

    footer = {col: '' for col in index_cols}
    footer[index_cols[0]] = 'Grand Total'
    for d in active_days:
        footer[d] = int(sum(v for v in agg[d] if isinstance(v, (int, float)) and v != ''))
    footer['Grand Total'] = total

    # Sum numeric columns (Fee, COD, etc.)
    numeric_cols = ['TOTAL FEE (USD)', 'COD (USD)']
    for nc in numeric_cols:
        if nc in agg.columns:
            col_sum = pd.to_numeric(agg[nc], errors='coerce').sum()
            if col_sum > 0:
                footer[nc] = round(col_sum, 2)

    rows = agg.to_dict('records')
    rows.append(footer)
    return rows, total, active_days


# ── Style helpers ──────────────────────────────────────────────────────────────

def _fill(h):
    return PatternFill(start_color=h, end_color=h, fill_type='solid')

def _font(name, color='000000', bold=False, size=10):
    return Font(name=name, color=color, bold=bold, size=size)

def _align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border():
    t = Side(style='thin', color='BFBFBF')
    return Border(left=t, right=t, top=t, bottom=t)


# ── Table writer ───────────────────────────────────────────────────────────────

def _write_table(ws, start_row, start_col, report_name, rows, index_cols, active_days, dc, handle="", show_top_title=False, max_index=0, order_created_map=None, order_status_map=None):
    """
    Write one report section at (start_row, start_col).
    Returns (next_free_row, next_free_col_after_block).
    """
    fn     = dc.get('font_name',         'Segoe UI')
    t_fg   = dc.get('title_color',       'FFFFFF')
    t_bg   = dc.get('title_fill_color',  '0F172A')
    h_bg   = dc.get('header_fill_color', '1E293B')
    h_fg   = dc.get('header_font_color', 'FFFFFF')
    z_bg   = dc.get('zone_fill_color',   'FFFFFF')
    z_fg   = dc.get('zone_font_color',   '0F172A')
    g_bg   = dc.get('group_fill_color',  'F8FAFC')
    g_fg   = dc.get('group_font_color',  '0F172A')
    tot_bg = dc.get('total_fill_color',  'F1F5F9')
    RED    = 'EF4444'
    bdr    = _border()

    # Pad index_cols so day columns align perfectly across tables
    padded_index = list(index_cols)
    while len(padded_index) < max_index:
        padded_index.append('')

    all_cols = padded_index + active_days + ['Grand Total']
    n        = len(all_cols)
    end_col  = start_col + n - 1

    # Row 1 — Title Row
    r = start_row
    ws.row_dimensions[r].height = 24
    today_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    report_name_kh = translate_header(report_name)
    title_text = f"{report_name_kh} — {handle} — {today_str}" if handle else f"{report_name_kh} — {today_str}"
    tc = ws.cell(r, start_col, title_text)
    tc.font      = _font(fn, t_fg, bold=True, size=11)
    tc.fill      = _fill(t_bg)
    tc.alignment = _align('center')
    tc.border    = bdr
    for ci in range(1, n):
        c = ws.cell(r, start_col + ci, '')
        c.fill   = _fill(t_bg)
        c.border = bdr
    if n > 1:
        ws.merge_cells(start_row=r, end_row=r, start_column=start_col, end_column=start_col + n - 1)
    r += 1

    # Row 2 — Month Row (merged per month group)
    # Row 3 — Day Number Row
    ws.row_dimensions[r].height = 20
    ws.row_dimensions[r + 1].height = 20

    # 1. Pre-fill and style all cells in both header rows
    for ci in range(n):
        col_idx = start_col + ci
        for row_idx in (r, r + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.fill = _fill(h_bg)
            cell.font = _font(fn, h_fg, bold=True)
            cell.alignment = _align('center')
            cell.border = bdr

    # 2. Vertically merge index columns and Grand Total column across both rows
    for ci, col_name in enumerate(all_cols):
        col_idx = start_col + ci
        is_index = (ci < len(padded_index))
        is_gt = (ci == n - 1)
        if is_index or is_gt:
            ws.cell(r, col_idx, translate_header(col_name))
            ws.merge_cells(start_row=r, end_row=r + 1, start_column=col_idx, end_column=col_idx)

    # 3. Write day numbers on Row r + 1 (day number row) for day columns
    for ci in range(len(padded_index), n - 1):
        col_idx = start_col + ci
        col_name = all_cols[ci]  # datetime.date object
        ws.cell(r + 1, col_idx, f"{col_name.day:02d}")

    # 4. Group day columns by month and horizontally merge on Row r (month row)
    month_groups = []
    current_month = None
    group_start = None
    for ci in range(len(padded_index), n - 1):
        col_name = all_cols[ci]
        m_val = (col_name.year, col_name.month)
        if m_val != current_month:
            if current_month is not None:
                month_groups.append((current_month, group_start, start_col + ci - 1))
            current_month = m_val
            group_start = start_col + ci
    if current_month is not None:
        month_groups.append((current_month, group_start, start_col + n - 2))

    for (yr, mo), start_c, end_c in month_groups:
        ws.cell(r, start_c, get_khmer_month_name(mo))
        if end_c > start_c:
            ws.merge_cells(start_row=r, end_row=r, start_column=start_c, end_column=end_c)

    r += 2

    # Data rows
    today = datetime.now().date()
    highlight_fill = _fill(HIGHLIGHT_COLOR)

    for row_dict in rows:
        ws.row_dimensions[r].height = 20
        is_total = str(row_dict.get(index_cols[0], '')).strip() == 'Grand Total'

        # Check overdue and status code
        is_overdue = False
        is_overdue_7days = False
        status_code = None
        if not is_total:
            order_id = normalize_id(row_dict.get('ORDER ID', ''))
            if order_created_map:
                created_date = order_created_map.get(order_id)
                if created_date:
                    age_days = (today - created_date).days
                    if age_days > HIGHLIGHT_OVER_DAYS:
                        is_overdue = True
                    if age_days > 7:
                        is_overdue_7days = True
            if order_status_map:
                status_code = order_status_map.get(order_id)

        for ci, col_name in enumerate(all_cols):
            val  = row_dict.get(col_name, '')
            cell = ws.cell(r, start_col + ci, val if val != '' else None)
            cell.border = bdr

            # Row-level fill (overdue / status-based only — boss wants 3 colors: orange, red, white)
            row_fill = None
            if is_total:
                row_fill = _fill(tot_bg)
            elif status_code in ("420", "472") and is_overdue_7days:
                row_fill = _fill("FFEBEB")  # Light red
            elif status_code in ("420", "472"):
                row_fill = _fill("E2EFDA")  # Light green (store waiting)
            elif is_overdue:
                row_fill = _fill("FFEBEB")  # Light orange (overdue)
            elif status_code in ("500", "510", "511", "512", "520", "540"):
                row_fill = _fill("FFEBEB")  # Light red/pink (return)

            if is_total:
                if row_fill:
                    cell.fill = row_fill
                cell.font      = _font(fn, RED, bold=True)
                cell.alignment = _align('center')
            elif status_code in ("420", "472") or is_overdue or status_code in ("500", "510", "511", "512", "520", "540"):
                if row_fill:
                    cell.fill = row_fill
                cell.font      = _font(fn, '000000', bold=False)
                cell.alignment = _align('center')
            elif col_name == index_cols[0]:
                cell.fill      = _fill(z_bg)
                cell.font      = _font(fn, z_fg, bold=False)
                cell.alignment = _align('center')
            elif len(index_cols) > 1 and col_name == index_cols[1]:
                cell.fill      = _fill(g_bg)
                cell.font      = _font(fn, g_fg, bold=False)
                cell.alignment = _align('center')
            elif col_name in index_cols:
                # All other index columns (CURRENT POST OFFICE, ORDER ID, etc.) → center
                if col_name == 'ORDER ID' and not is_total:
                    cell.font      = _font(fn, '1E293B', bold=False)
                cell.alignment = _align('center')
            elif col_name in active_days:
                cell.font      = _font(fn, '1E293B', bold=False)
                cell.alignment = _align('center')
            elif col_name == 'Grand Total':
                cell.font      = _font(fn, RED, bold=True)
                cell.alignment = _align('center')
            else:
                cell.font      = _font(fn, '1E293B', bold=False)
                cell.alignment = _align('center')
        r += 1

    return r, end_col + 1  # (next_free_row, next_free_col)


# ── Column width setter ────────────────────────────────────────────────────────

def _set_col_widths(ws):
    """
    Set fixed widths:
    - Day cols (2-digit header 01-31) → W_DAY (all same, uniform)
    - ZONE col → W_ZONE
    - Grand Total col → W_GT
    - Empty cols → 1.0 (hidden/minimal)
    - Other text cols → auto-fit capped at W_MAX
    """
    # Scan all header rows to classify columns
    day_cols_ci   = set()
    zone_cols_ci  = set()
    gt_cols_ci    = set()
    empty_cols_ci = set()

    for r in range(1, min(ws.max_row + 1, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            try:
                val = str(ws.cell(r, c).value or '').strip()
            except Exception:
                val = ''
            if val.isdigit() and len(val) == 2 and 1 <= int(val) <= 31:
                day_cols_ci.add(c)
            elif val == 'ZONE':
                zone_cols_ci.add(c)
            elif val == 'Grand Total':
                gt_cols_ci.add(c)

    # Detect fully empty columns
    for c in range(1, ws.max_column + 1):
        if all(
            str(ws.cell(r, c).value or '').strip() == ''
            for r in range(1, ws.max_row + 1)
        ):
            empty_cols_ci.add(c)

    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        if c in day_cols_ci:
            ws.column_dimensions[letter].width = W_DAY
        elif c in zone_cols_ci:
            ws.column_dimensions[letter].width = W_ZONE
        elif c in gt_cols_ci:
            ws.column_dimensions[letter].width = W_GT
        elif c in empty_cols_ci:
            ws.column_dimensions[letter].width = 1.0
        else:
            # Determine header value to apply specific rules
            header_val = ''
            for r_scan in range(1, min(10, ws.max_row + 1)):
                cell_v = str(ws.cell(r_scan, c).value or '').strip()
                if cell_v in ('Cus name', 'RECEIVER', 'Phone', 'ORDER ID', 'CURRENT POST OFFICE', 'POST OFFICE HANDLE', 'REMARK'):
                    header_val = cell_v
                    break
            
            max_len = 0
            for r in range(1, ws.max_row + 1):
                try:
                    v = ws.cell(r, c).value
                    if v:
                        max_len = max(max_len, len(str(v)))
                except Exception:
                    pass
            
            if header_val in ('Cus name', 'RECEIVER'):
                ws.column_dimensions[letter].width = min(max(max_len + 4, 25), 55)
            elif header_val == 'Phone':
                ws.column_dimensions[letter].width = min(max(max_len + 4, 19), 38)
            elif header_val == 'ORDER ID':
                ws.column_dimensions[letter].width = min(max(max_len + 3, 16), 24)
            else:
                ws.column_dimensions[letter].width = min(max(max_len + 2, W_MIN), W_MAX)


# ── Excel builders ─────────────────────────────────────────────────────────────

def build_handle_excel(handle, sections, day_cols, dc, out_path, mode='wide', order_created_map=None, order_status_map=None):
    """
    sections: list of (report_name, rows, total, index_cols, active_days)
    mode: 'wide' = side by side, 'long' = stacked
    """
    wb = Workbook()
    ws = wb.active
    ws.title = handle[:31]
    fn  = dc.get('font_name', 'Segoe UI')

    # Shared active days = union of all section active days, in sorted order
    shared_days = sorted(set(d for _, _, _, _, ad in sections for d in ad))

    # Compute total cols for merge using the widest index cols set
    max_index = max((len(ic) for _, _, _, ic, _ in sections), default=3)
    total_cols = max_index + len(shared_days) + 1  # +1 for Grand Total

    if mode == 'wide':
        cur_col = 1
        for report_name, rows, total, index_cols, active_days in sections:
            _, next_col = _write_table(ws, 1, cur_col, report_name,
                                       rows, index_cols, shared_days, dc, handle=handle, max_index=max_index, order_created_map=order_created_map, order_status_map=order_status_map)
            cur_col = next_col + GAP_COLS
    else:
        # Long mode: Pickup → Delivery → Pending stacked
        r = 1
        for i, (report_name, rows, total, index_cols, active_days) in enumerate(sections):
            next_row, _ = _write_table(ws, r, 1, report_name,
                                       rows, index_cols, shared_days, dc, handle=handle, show_top_title=(i==0), max_index=max_index, order_created_map=order_created_map, order_status_map=order_status_map)
            r = next_row + 1  # 1 blank row gap

    _set_col_widths(ws)
    wb.save(out_path)


def build_final_excel(all_handle_sections, day_cols, dc, out_path, mode='wide', order_created_map=None, order_status_map=None, handle_title='ALL BRANCHES'):
    from collections import defaultdict
    wb = Workbook()
    
    report_types = ALL_TABS
    
    # Global shared days across ALL handles
    shared_days = sorted(set(
        d for _, sections in all_handle_sections
        for _, _, _, _, ad in sections
        for d in ad
    ))

    # Collect all data rows (excluding branch total rows) for each report type
    combined_data = defaultdict(list)
    index_cols_map = {}
    
    for handle, sections in all_handle_sections:
        for rn, rows, total, index_cols, active_days in sections:
            index_cols_map[rn] = index_cols
            non_footer_rows = [r for r in rows if r.get(index_cols[0]) != 'Grand Total']
            combined_data[rn].extend(non_footer_rows)

    for idx, rn in enumerate(report_types):
        if idx == 0:
            ws = wb.active
            ws.title = rn
        else:
            ws = wb.create_sheet(title=rn)
            
        rows = combined_data[rn]
        icols = index_cols_map.get(rn)
        
        if not icols:
            icols = REPORT_COLS[rn]

        if not rows:
            footer = {col: '' for col in icols}
            footer[icols[0]] = 'Grand Total'
            footer['Grand Total'] = 0
            rows = [footer]
            combined_active_days = []
        else:
            # Sort the combined rows
            if rn in ('Delivery', 'Transit', 'Branch') and 'ACTION' in icols:
                def get_combined_sort_key(row):
                    zone = str(row.get('ZONE', '') or '').strip().upper()
                    action = str(row.get('ACTION', '') or '').strip()
                    if 'ដឹកជញ្ជូន' in action:
                        act_w = 1
                    elif 'ពិនិត្យ' in action:
                        act_w = 2
                    elif 'ត្រឡប់' in action:
                        act_w = 3
                    else:
                        act_w = 4
                    
                    oid = str(row.get('ORDER ID', '') or '').strip()
                    dt = order_created_map.get(oid) if order_created_map else None
                    if dt is None:
                        import datetime as dt_mod
                        dt = dt_mod.date.max
                    return (zone, act_w, dt)
                rows = sorted(rows, key=get_combined_sort_key)
            else:
                sort_keys = [c for c in ['ZONE', 'POST OFFICE HANDLE', 'CURRENT POST OFFICE', 'ORDER ID'] if c in icols]
                def get_sort_key(row):
                    return tuple(str(row.get(k, '') or '').strip().upper() for k in sort_keys)
                rows = sorted(rows, key=get_sort_key)
            
            combined_active_days = shared_days

            # Compute column totals and Grand Total for the combined set
            col_totals = defaultdict(int)
            grand_total = 0
            for row in rows:
                grand_total += 1
                for d in combined_active_days:
                    val = row.get(d)
                    if isinstance(val, (int, float)) and val != '':
                        col_totals[d] += int(val)

            # Build the global footer row
            footer = {col: '' for col in icols}
            footer[icols[0]] = 'Grand Total'
            for d in combined_active_days:
                footer[d] = col_totals[d] if col_totals[d] > 0 else ''
            footer['Grand Total'] = grand_total
            rows.append(footer)

        _write_table(ws, 1, 1, rn, rows, icols, combined_active_days, dc,
                     handle=handle_title, show_top_title=False, max_index=len(icols),
                     order_created_map=order_created_map, order_status_map=order_status_map)
        
        _set_col_widths(ws)
        
    wb.save(out_path)


# ── Main ───────────────────────────────────────────────────────────────────────

def generate_reports_from_data(export_path, ref_path, output_dir,
                                return_metadata=False, mode='wide', target_handles=None):
    config_path = os.path.join(os.path.dirname(__file__), 'config.json')
    with open(config_path, encoding='utf-8') as f:
        cfg = json.load(f)
    zone_mapping = cfg.get('zone_mapping', {})
    excel_design = cfg.get('excel_design', {})


    df_ref = load_reference(ref_path)

    # Auto-detect sheet
    xl = pd.ExcelFile(export_path)
    sheet = xl.sheet_names[0]
    if len(xl.sheet_names) > 1:
        for sname in xl.sheet_names:
            try:
                s = xl.parse(sname, nrows=3)
                if 'ORDER ID' in s.columns and (
                    'CURRENT STATUS' in s.columns or
                    'is tồn kết nối' in s.columns
                ):
                    sheet = sname
                    break
            except Exception:
                continue
    df = xl.parse(sheet)

    # Date column for day-of-month grouping
    date_col = (
        'CREATED DATE' if 'CREATED DATE' in df.columns else
        'CURRENT TIME' if 'CURRENT TIME' in df.columns else None
    )

    order_created_map = {}
    order_status_map = {}
    if 'ORDER ID' in df.columns:
        if date_col:
            parsed_created = pd.to_datetime(df[date_col], dayfirst=True, format='mixed', errors='coerce')
            for order_id, dt in zip(df['ORDER ID'], parsed_created):
                if pd.notna(dt):
                    order_created_map[normalize_id(order_id)] = dt.date()
        if 'CURRENT STATUS' in df.columns:
            for order_id, status_val in zip(df['ORDER ID'], df['CURRENT STATUS']):
                if pd.notna(order_id) and pd.notna(status_val):
                    sc_val = str(status_val).strip()
                    import re
                    match = re.match(r'^(-?\d+)', sc_val)
                    if match:
                        order_status_map[normalize_id(order_id)] = match.group(1)

    today = datetime.now().date()

    # Build day_cols = only days that have data in the whole export, + today
    if date_col:
        parsed = pd.to_datetime(df[date_col], dayfirst=True, format='mixed', errors='coerce')
        dates_with_data = set(parsed.dropna().dt.date.tolist())
    else:
        dates_with_data = set()

    dates_with_data.add(today)
    day_cols = sorted(list(dates_with_data))

    # Exclude test orders
    test_col = next(
        (c for c in df.columns if str(c).strip().lower() in ('is test', 'đơn test', 'don test')),
        None
    )
    if test_col:
        df = df[
            df[test_col].isna() |
            df[test_col].astype(str).str.strip().isin(['', 'nan', 'NaN', '#N/A'])
        ].copy()

    test_order_ids = load_test_order_ids(cfg)
    if test_order_ids and 'ORDER ID' in df.columns:
        df = df[~df['ORDER ID'].astype(str).str.strip().isin(test_order_ids)].copy()

    # Exclude test orders based on keywords in name/note columns
    if cfg.get("pivot", {}).get("exclude_test", False):
        keywords = cfg["pivot"].get("test_keywords", ["test"])
        if keywords:
            check_cols = [c for c in df.columns if any(k in str(c).lower() for k in ('name', 'note', 'customer', 'sender', 'receiver', 'remark', 'address'))]
            for kw in keywords:
                kw_lower = str(kw).lower().strip()
                for col in check_cols:
                    # Drop rows where the keyword is found in the column
                    df = df[~df[col].astype(str).str.lower().str.contains(kw_lower, na=False)].copy()

    if df.empty:
        raise ValueError("No data left after test/exclusion filters.")

    # Classification column
    classify_col = next(
        (c for c in df.columns if 'tồn kết nối' in str(c).lower()), None
    )

    # Status code
    if 'CURRENT STATUS' in df.columns:
        df['CURRENT STATUS'] = df['CURRENT STATUS'].astype(str).str.strip()
        df['STATUS_CODE'] = df['CURRENT STATUS'].str.extract(r'^(\d{3})')[0]
    else:
        df['STATUS_CODE'] = ''

    if 'CURRENT POST OFFICE' in df.columns:
        df['_po_key'] = df['CURRENT POST OFFICE'].apply(normalize_code)
    else:
        df['_po_key'] = ''

    # POST OFFICE HANDLE + ZONE
    if 'POST OFFICE HANDLE' in df.columns:
        dm = df.copy()
        dm['POST OFFICE HANDLE'] = dm['POST OFFICE HANDLE'].apply(normalize_code)
        if 'CURRENT POST OFFICE' in dm.columns:
            dm['CURRENT POST OFFICE'] = dm['CURRENT POST OFFICE'].apply(normalize_code)
    else:
        dm = pd.merge(df, df_ref, left_on='_po_key', right_on='_ref_post_office_key', how='left')
        if 'CURRENT POST OFFICE' in dm.columns:
            dm['CURRENT POST OFFICE'] = dm['CURRENT POST OFFICE'].apply(normalize_code)
        dm['POST OFFICE HANDLE'] = dm['post_office_handle'].apply(normalize_code)
        if 'CURRENT POST OFFICE' in dm.columns:
            dm.loc[dm['POST OFFICE HANDLE'] == '', 'POST OFFICE HANDLE'] = dm['CURRENT POST OFFICE']

    if 'ZONE' not in dm.columns or dm['ZONE'].isna().all():
        dm['ZONE'] = dm['POST OFFICE HANDLE'].apply(lambda x: get_zone(x, zone_mapping))
    else:
        dm['ZONE'] = dm['ZONE'].apply(clean_text)

    if 'ORDER ID' in dm.columns:
        dm['ORDER ID'] = dm['ORDER ID'].astype(str).str.strip()

    for col in ("RECEIVE POST OFFICE", "DELIVERY POST OFFICE", "CURRENT POST OFFICE"):
        if col not in dm.columns:
            dm[col] = ""
        dm[col] = dm[col].apply(normalize_code)

    # RECEIVER
    if 'RECEIVER' not in dm.columns:
        if 'RECEIVER' in df.columns:
            dm['RECEIVER'] = df['RECEIVER'].apply(
                lambda v: str(v).split(' - ', 1)[1].strip() if ' - ' in str(v) else clean_text(v))
        else:
            dm['RECEIVER'] = ''

    # Cus name / Phone
    cus_src = next((c for c in dm.columns if c.strip() == 'Cus name'), None)
    if cus_src:
        dm['Cus name'] = dm[cus_src].apply(clean_text)
    elif 'SENDER' in dm.columns:
        dm['Cus name'] = dm['SENDER'].apply(
            lambda v: str(v).split(' - ', 1)[1].strip() if ' - ' in str(v) else clean_text(v))
    else:
        dm['Cus name'] = ''

    if 'Phone' not in dm.columns:
        if 'SENDER' in dm.columns:
            dm['Phone'] = dm['SENDER'].apply(
                lambda v: str(v).split(' - ', 1)[0].strip() if ' - ' in str(v) else '')
        else:
            dm['Phone'] = ''

    os.makedirs(output_dir, exist_ok=True)

    # Build status map from config.json
    status_map = {}
    for r in cfg.get("reports", []):
        key = r.get('key', '').lower()
        if 'pickup' in key:
            label = 'Pickup'
        elif 'delivery' in key:
            label = 'Delivery'
        elif 'transit' in key:
            label = 'Transit'
        elif 'branch' in key:
            label = 'Branch'
        else:
            label = r.get('is_label', '')
        for sc in r.get("status_codes", []):
            status_map[str(sc).strip()] = label

    # Globally drop completed/done statuses from all reports
    if 'STATUS_CODE' in dm.columns:
        dm = dm[~dm['STATUS_CODE'].isin(['99', '100', '410', '201', '520'])].copy()

    # Normalise Fee/COD columns if present
    fee_col_raw = next((c for c in dm.columns if 'TOTAL FEE' in c.upper()), None)
    cod_col_raw = next((c for c in dm.columns if c.upper().startswith('COD')), None)
    pay_col_raw = next((c for c in dm.columns if 'PAYMENT' in c.upper() and 'METHOD' in c.upper()), None)

    if fee_col_raw:
        dm['TOTAL FEE (USD)'] = pd.to_numeric(dm[fee_col_raw], errors='coerce')
    else:
        dm['TOTAL FEE (USD)'] = None

    # Zero out fee when SENDER pays — delivery staff only collect from receiver
    # 'Người gửi' = sender pays → show 0
    # 'Người nhận' / 'Trả sau' = receiver pays → show real fee
    if pay_col_raw and 'TOTAL FEE (USD)' in dm.columns:
        sender_pays_mask = dm[pay_col_raw].astype(str).str.contains('g\u1eedi', case=False, na=False)
        dm.loc[sender_pays_mask, 'TOTAL FEE (USD)'] = 0

    if cod_col_raw:
        dm['COD (USD)'] = pd.to_numeric(dm[cod_col_raw], errors='coerce')
    else:
        dm['COD (USD)'] = None

    # Filter per report type
    type_data = {}
    if classify_col:
        dm["_report_class"] = dm[classify_col].astype(str).str.strip()
    else:
        dm["_report_class"] = dm['STATUS_CODE'].map(status_map).fillna("Unknown")

    if target_handles:
        target_handles = [h.upper() for h in target_handles if h]

    for rn in ALL_TABS:
        df_t = dm[dm["_report_class"] == rn].copy()
        if target_handles:
            filter_col = REPORT_FILTER_COLS[rn]
            if filter_col in df_t.columns:
                df_t = df_t[df_t[filter_col].isin(target_handles)]
        # Exclude MEGA/HUB/DVC from Transit/Branch (these show in /total mega instead)
        if rn in ('Transit', 'Branch') and 'CURRENT POST OFFICE' in df_t.columns:
            mega_mask = df_t['CURRENT POST OFFICE'].str.contains('MEGA|HUB|DVC', case=False, na=False)
            df_t = df_t[~mega_mask].copy()
        # Add ACTION columns
        if rn == 'Delivery' and 'STATUS_CODE' in df_t.columns:
            def _delivery_action(row):
                sc = str(row.get('STATUS_CODE', '')).strip()
                return DELIVERY_ACTION_MAP.get(sc, ('', ''))[0]
            def _delivery_next_step(row):
                sc = str(row.get('STATUS_CODE', '')).strip()
                return DELIVERY_ACTION_MAP.get(sc, ('', ''))[1]
            df_t['ACTION'] = df_t.apply(_delivery_action, axis=1)
            df_t['NEXT_STEP'] = df_t.apply(_delivery_next_step, axis=1)
        if rn == 'Transit' and 'STATUS_CODE' in df_t.columns:
            def _transit_action(row):
                sc = str(row.get('STATUS_CODE', '')).strip()
                return TRANSIT_ACTION_MAP.get(sc, ('', ''))[0]
            def _transit_next_step(row):
                sc = str(row.get('STATUS_CODE', '')).strip()
                return TRANSIT_ACTION_MAP.get(sc, ('', ''))[1]
            df_t['ACTION'] = df_t.apply(_transit_action, axis=1)
            df_t['NEXT_STEP'] = df_t.apply(_transit_next_step, axis=1)
        if rn == 'Branch' and 'STATUS_CODE' in df_t.columns:
            def _branch_action(row):
                sc = str(row.get('STATUS_CODE', '')).strip()
                return BRANCH_ACTION_MAP.get(sc, ('', ''))[0]
            def _branch_next_step(row):
                sc = str(row.get('STATUS_CODE', '')).strip()
                return BRANCH_ACTION_MAP.get(sc, ('', ''))[1]
            df_t['ACTION'] = df_t.apply(_branch_action, axis=1)
            df_t['NEXT_STEP'] = df_t.apply(_branch_next_step, axis=1)
        type_data[rn] = df_t

    # Gather all handles
    all_handles = set()
    for rn in ALL_TABS:
        df_t = type_data[rn]
        filter_col = REPORT_FILTER_COLS[rn]
        if filter_col in df_t.columns:
            all_handles.update(df_t[filter_col].dropna().unique())

    unique_handles = sorted(list(h for h in all_handles if str(h).strip()))
    if target_handles:
        for th in target_handles:
            if th not in unique_handles:
                unique_handles.append(th)
        unique_handles = sorted(unique_handles)

    handle_results = []
    all_handle_sections = []
    overall = {rn: 0 for rn in ALL_TABS}

    for handle in unique_handles:
        sections = []
        counts = {}
        for rn in ALL_TABS:
            df_t = type_data[rn]
            filter_col = REPORT_FILTER_COLS[rn]
            if filter_col in df_t.columns:
                df_h = df_t[df_t[filter_col] == handle].copy()
            else:
                df_h = pd.DataFrame()

            icols = REPORT_COLS[rn]
            if df_h.empty:
                counts[rn] = 0
                continue
                
            if rn in ('Delivery', 'Transit', 'Branch') and 'ACTION' in df_h.columns:
                def get_action_weight(val):
                    val_str = str(val)
                    if 'ដឹកជញ្ជូន' in val_str:
                        return 1
                    elif 'ពិនិត្យ' in val_str:
                        return 2
                    elif 'ត្រឡប់' in val_str:
                        return 3
                    return 4
                df_h = df_h.copy()
                df_h['_action_weight'] = df_h['ACTION'].apply(get_action_weight)
                
                created_col = 'CREATED DATE'
                if created_col in df_h.columns:
                    df_h['_created_dt'] = pd.to_datetime(df_h[created_col], dayfirst=True, format='mixed', errors='coerce')
                else:
                    df_h['_created_dt'] = pd.NaT
                
                sort_by = []
                ascending = []
                if 'ZONE' in df_h.columns:
                    sort_by.append('ZONE')
                    ascending.append(True)
                sort_by.extend(['_action_weight', '_created_dt'])
                ascending.extend([True, True])
                
                df_h = df_h.sort_values(by=sort_by, ascending=ascending)
            else:
                sort_cols = [c for c in ['ZONE', filter_col, 'ORDER ID'] if c in df_h.columns]
                if sort_cols:
                    df_h = df_h.sort_values(by=sort_cols)

            rows, total, active_days = build_section_rows(df_h, icols, day_cols, date_col)
            counts[rn] = total
            overall[rn] += total
            if total > 0:
                sections.append((rn, rows, total, icols, active_days))

        total_handle = sum(counts.values())
        if total_handle == 0 and not (target_handles and handle in target_handles):
            continue

        if total_handle > 0:
            all_handle_sections.append((handle, sections))

        handle_files = []
        for rn, rows, total, icols, active_days in sections:
            tmp_xlsx = os.path.join(output_dir, f"Report_{handle}_{rn}_{today.strftime('%d_%m_%Y_%H%M%S')}.xlsx")
            build_handle_excel(f"Report_{handle}_{rn}", [(rn, rows, total, icols, active_days)], day_cols, excel_design, tmp_xlsx, mode=mode, order_created_map=order_created_map, order_status_map=order_status_map)
            handle_files.append({'path': tmp_xlsx, 'handle': handle})

        # Build exactly ONE multi-tab Excel file for this branch containing all tabs
        combined_xlsx = os.path.join(output_dir, f"Report_{handle}_{today.strftime('%d_%m_%Y')}.xlsx")
        build_final_excel([(handle, sections)], day_cols, excel_design, combined_xlsx, mode=mode, order_created_map=order_created_map, order_status_map=order_status_map, handle_title=handle)

        remark = (
            f"{handle}  |  "
            + "  |  ".join(f"{t}: {counts.get(t,0)}" for t in ALL_TABS)
            + f"  |  Total: {total_handle}"
        )

        handle_results.append({
            'handle':            handle,
            'handle_counts':     counts,
            'handle_files':      handle_files,
            'handle_excel_path': combined_xlsx,
            'remark':            remark,
            'sections':          sections,
        })

    final_xlsx = os.path.join(output_dir, f"Report_All_{today.strftime('%d_%m_%Y')}.xlsx")
    if all_handle_sections:
        build_final_excel(all_handle_sections, day_cols, excel_design, final_xlsx, mode=mode, order_created_map=order_created_map, order_status_map=order_status_map)
    else:
        wb = Workbook()
        wb.save(final_xlsx)
        
    if not handle_results:
        handle_str = 'ALL' if not target_handles else ', '.join(target_handles)
        handle_results = [{
            'handle':        handle_str,
            'handle_counts': overall,
            'handle_files':  [{'path': final_xlsx, 'handle': handle_str}],
            'remark':        f"No data found for {handle_str}",
            'sections':      [],
        }]

    grand_total = sum(overall.values())
    summary = "\n".join([
        f"📋 Daily Report  {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        f"Pickup: {overall.get('Pickup',0)}  |  Delivery: {overall.get('Delivery',0)}  |  Transit: {overall.get('Transit',0)}  |  Branch: {overall.get('Branch',0)}",
        f"Grand Total: {grand_total}",
    ])

    result = {
        'handle_results':  handle_results,
        'final_xlsx':      final_xlsx,
        'summary_caption': summary,
        'overall_counts':  overall,
        'type_data':       type_data,
        'day_cols':        day_cols,
        'cur_time_col':    date_col,
        'order_status_map': order_status_map,
    }
    return result if return_metadata else [final_xlsx]

