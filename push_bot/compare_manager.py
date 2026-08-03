import os
import json
import re
import glob
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COMPARE_DIR = "compare"
os.makedirs(COMPARE_DIR, exist_ok=True)

SNAPSHOT_FILE = os.path.join(COMPARE_DIR, "compare_snapshots.json")
EXCLUDE_KEYWORDS = ["TRAINER", "GLOBAL", "EXTERNAL", "TEST"]

CATEGORIES = ["Pickup", "Delivery", "Transit", "Not Assign"]
SHIFTS = ["9AM", "2PM", "5PM"]

PREFIX_ZONE_MAP = {
    "PNP": "Phnom Penh",
    "BAN": "Banteay Meanchey",
    "BAT": "Battambang",
    "KMP": "Kandal",
    "TAK": "Takeo",
    "KPC": "Kampong Cham",
    "KPT": "Kampot",
    "SRP": "Siem Reap",
    "KPS": "Preah Sihanouk",
    "KPE": "Kampong Speu",
    "KTH": "Kampong Thom",
    "SVR": "Svay Rieng",
    "PVG": "Prey Veng",
    "PUR": "Pursat",
    "KRT": "Kratie",
    "STU": "Stung Treng",
    "RAT": "Ratanakiri",
    "MON": "Mondulkiri",
    "PRE": "Preah Vihear",
    "ODD": "Oddar Meanchey",
    "PVI": "Pailin",
    "TKG": "Tbong Khmum",
    "KEP": "Kep",
    "KOH": "Koh Kong"
}

def clean_cache_3_times_daily():
    """Cleans old files in compare/ folder 3 times per day."""
    try:
        if os.path.exists(SNAPSHOT_FILE):
            with open(SNAPSHOT_FILE, "r", encoding="utf-8") as f:
                snapshots = json.load(f)
            
            cutoff_date = datetime.now() - timedelta(days=3)
            cleaned = {}
            for d_str, val in snapshots.items():
                try:
                    dt = datetime.strptime(d_str, "%d/%m/%Y")
                    if dt >= cutoff_date:
                        cleaned[d_str] = val
                except Exception:
                    cleaned[d_str] = val

            with open(SNAPSHOT_FILE, "w", encoding="utf-8") as f:
                json.dump(cleaned, f, ensure_ascii=False, indent=2)

        for fpath in glob.glob(os.path.join(COMPARE_DIR, "*.xlsx")):
            try:
                mtime = datetime.fromtimestamp(os.path.getmtime(fpath))
                if datetime.now() - mtime > timedelta(hours=24):
                    os.remove(fpath)
            except Exception:
                pass
    except Exception:
        pass

def load_post_office_lookup_map():
    po_map = {}
    if os.path.exists("post_office_lookup.csv"):
        try:
            df_po = pd.read_csv("post_office_lookup.csv", encoding="utf-8-sig")
            df_po.columns = [str(c).strip().lower() for c in df_po.columns]
            if "current_post_office" in df_po.columns and "post_office_handle" in df_po.columns:
                for _, r in df_po.iterrows():
                    cur = str(r["current_post_office"]).strip().upper()
                    hnd = str(r["post_office_handle"]).strip().upper()
                    if cur and hnd and cur != "NAN" and hnd != "NAN":
                        po_map[cur] = hnd
        except Exception:
            pass
    return po_map

PO_LOOKUP_MAP = load_post_office_lookup_map()

def resolve_post_office_handle(raw_po):
    po = str(raw_po).strip().upper()
    if not po or po == "NAN" or any(kw in po for kw in EXCLUDE_KEYWORDS):
        return None

    SPECIAL_MAP = {
        "STU": "STUP001",
        "MON": "MONP001",
        "PAI": "PAIP001",
        "ROT": "RATP001",
        "PRH": "PREP001",
        "TBO": "TBKP001",
        "MOS": "MONP001",
        "CHA": "KCPC001",
        "CHH": "KCPC001",
    }

    prefix3 = po[:3]
    if prefix3 in SPECIAL_MAP:
        return SPECIAL_MAP[prefix3]

    m = re.match(r'^([A-Z]{3,4})[AS](\d{1,4})$', po)
    if m:
        prefix, num = m.groups()
        if len(num) == 3:
            return f"{prefix}P{num}"
        return f"{prefix}P001"

    global PO_LOOKUP_MAP
    if not PO_LOOKUP_MAP:
        PO_LOOKUP_MAP = load_post_office_lookup_map()

    if po in PO_LOOKUP_MAP:
        return PO_LOOKUP_MAP[po]

    return po

def resolve_branch_zone(branch_code, df_detail=None):
    b_code = str(branch_code).strip().upper()

    if df_detail is not None:
        df = df_detail.copy()
        df.columns = [str(c).strip().upper() for c in df.columns]
        h_col = "POST OFFICE HANDLE" if "POST OFFICE HANDLE" in df.columns else ("CURRENT POST OFFICE" if "CURRENT POST OFFICE" in df.columns else None)
        z_col = "ZONE" if "ZONE" in df.columns else ("RECEIVE PROVINCE" if "RECEIVE PROVINCE" in df.columns else None)
        if h_col and z_col:
            m = df[df[h_col].astype(str).str.strip().str.upper() == b_code]
            if not m.empty:
                zv = str(m.iloc[0][z_col]).strip()
                if zv and zv.lower() != "nan":
                    return zv

    if os.path.exists("post_office_lookup.csv"):
        try:
            df_po = pd.read_csv("post_office_lookup.csv", encoding="utf-8-sig")
            df_po.columns = [str(c).strip().lower() for c in df_po.columns]
            if "post_office_handle" in df_po.columns and "zone" in df_po.columns:
                m = df_po[df_po["post_office_handle"].astype(str).str.strip().str.upper() == b_code]
                if not m.empty:
                    zv = str(m.iloc[0]["zone"]).strip()
                    if zv and zv.lower() != "nan":
                        return zv
        except Exception:
            pass

    prefix3 = b_code[:3]
    if prefix3 in PREFIX_ZONE_MAP:
        return PREFIX_ZONE_MAP[prefix3]

    return "Other Zone"

def get_all_known_branches(df_detail=None):
    branches = set()
    if os.path.exists("post_office_lookup.csv"):
        try:
            df_po = pd.read_csv("post_office_lookup.csv", encoding="utf-8-sig")
            df_po.columns = [str(c).strip().lower() for c in df_po.columns]
            if "post_office_handle" in df_po.columns:
                for h in df_po["post_office_handle"].dropna().unique():
                    h_str = str(h).strip().upper()
                    if h_str and h_str != "NAN" and not any(kw in h_str for kw in EXCLUDE_KEYWORDS):
                        branches.add(h_str)
        except Exception:
            pass

    if df_detail is not None:
        df = df_detail.copy()
        df.columns = [str(c).strip().upper() for c in df.columns]
        handle_col = "POST OFFICE HANDLE" if "POST OFFICE HANDLE" in df.columns else ("CURRENT POST OFFICE" if "CURRENT POST OFFICE" in df.columns else None)
        if handle_col:
            for h in df[handle_col].dropna().unique():
                h_str = resolve_post_office_handle(h)
                if h_str:
                    branches.add(h_str)

    return sorted(list(branches))

def determine_shift(now=None):
    """
    7:00 AM - 11:59 AM -> 9AM
    12:00 PM - 3:59 PM -> 2PM
    4:00 PM onwards    -> 5PM
    """
    if now is None:
        now = datetime.now()
    hour = now.hour
    if hour < 12:
        return "9AM"
    elif hour < 16:
        return "2PM"
    else:
        return "5PM"

def load_snapshots():
    clean_cache_3_times_daily()
    if os.path.exists(SNAPSHOT_FILE):
        try:
            with open(SNAPSHOT_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_snapshots(data):
    with open(SNAPSHOT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def is_urgent_bill(r):
    sc = str(r.get("STATUS_CODE", "") or r.get("STATUS", "") or "").lstrip("S").strip()
    if sc and " " in sc:
        sc = sc.split()[0].strip()

    if sc in ("99", "100", "201", "410", "420", "472", "520"):
        return False

    if "_IS_OVERDUE" in r and pd.notna(r["_IS_OVERDUE"]):
        return bool(r["_IS_OVERDUE"])

    age_val = str(r.get("AGE", "") or "")
    if "🔴" in age_val:
        return True

    cat_raw = str(r.get("REPORT TYPE", "") or r.get("TYPE", "") or "").upper()
    is_transit = ("TRANSIT" in cat_raw or "MEGA" in cat_raw or sc in ("306", "309"))
    threshold = 48 if is_transit else 24

    scan_time = ""
    for tc in ["CURRENT TIME", "STATUS 306 AT STORE / AGENT (LAST TIME)", "STATUS 306 AT STORE / AGENT FROM HUB (FIRST TIME)", "SCAN TIME", "CREATED DATE"]:
        if tc in r and pd.notna(r[tc]):
            val_str = str(r[tc]).strip()
            if val_str and val_str.lower() != "nan":
                scan_time = val_str
                break

    if scan_time:
        match = re.search(r'(\d+)\s*h', scan_time, re.IGNORECASE)
        if match:
            return float(match.group(1)) >= threshold
        try:
            parsed_dt = pd.to_datetime(scan_time, dayfirst=True, format='mixed', errors='coerce')
            if pd.notna(parsed_dt):
                return ((datetime.now() - parsed_dt).total_seconds() / 3600.0) >= threshold
        except Exception:
            pass

    return True

def extract_total_report_counts(df_detail):
    """Extracts urgent counts and section counts per post office handle exactly like /total."""
    df = df_detail.copy()
    df.columns = [str(c).strip().upper() for c in df.columns]

    handle_col = "POST OFFICE HANDLE" if "POST OFFICE HANDLE" in df.columns else ("CURRENT POST OFFICE" if "CURRENT POST OFFICE" in df.columns else None)
    sc_col = "STATUS_CODE" if "STATUS_CODE" in df.columns else ("STATUS" if "STATUS" in df.columns else None)

    handle_map = {}
    if not handle_col or not sc_col:
        return handle_map

    for _, r in df.iterrows():
        raw_h = str(r.get(handle_col, "") or "").strip()
        hnd = resolve_post_office_handle(raw_h)
        if not hnd:
            continue

        sc = str(r.get(sc_col, "") or "").lstrip("S").strip()
        if sc and " " in sc:
            sc = sc.split()[0].strip()

        if sc in ("99", "100", "201", "410", "420", "520"):
            continue

        if hnd not in handle_map:
            handle_map[hnd] = {"urgent": 0, "pickup": 0, "delivery": 0, "transit": 0, "branch": 0}

        if is_urgent_bill(r):
            handle_map[hnd]["urgent"] += 1

        cat_raw = str(r.get("REPORT TYPE", "") or r.get("TYPE", "") or "").upper()
        if "PICKUP" in cat_raw or sc in ("200", "210", "302", "310"):
            handle_map[hnd]["pickup"] += 1
        elif "DELIVERY" in cat_raw or sc in ("311", "401", "402", "470", "472", "480", "500"):
            handle_map[hnd]["delivery"] += 1
        elif "TRANSIT" in cat_raw or "MEGA" in cat_raw or sc in ("306", "309"):
            handle_map[hnd]["transit"] += 1
        else:
            handle_map[hnd]["branch"] += 1

    return handle_map

def get_next_shift_to_record(date_str):
    snapshots = load_snapshots()
    day_data = snapshots.get(date_str, {})
    if "9AM" not in day_data:
        return "9AM"
    elif "2PM" not in day_data:
        return "2PM"
    else:
        return "5PM"

def record_total_snapshot(date_str, shift_name, df_detail):
    snapshots = load_snapshots()
    if date_str not in snapshots:
        snapshots[date_str] = {}

    target_shift = get_next_shift_to_record(date_str)
    handle_map = extract_total_report_counts(df_detail)
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    snapshots[date_str][target_shift] = {
        "captured_at": now_str,
        "handles": handle_map
    }

    save_snapshots(snapshots)
    return snapshots

def build_comparison_summary(date_str, df_detail=None):
    if df_detail is not None:
        target_shift = get_next_shift_to_record(date_str)
        record_total_snapshot(date_str, target_shift, df_detail)

    snapshots = load_snapshots()
    day_data = snapshots.get(date_str, {})

    h_9am = day_data.get("9AM", {}).get("handles", {})
    h_2pm = day_data.get("2PM", {}).get("handles", {})
    h_5pm = day_data.get("5PM", {}).get("handles", {})

    has_2pm = bool(h_2pm)
    has_5pm = bool(h_5pm)

    all_handles = get_all_known_branches(df_detail)

    rows = []
    tot_urg_9am = tot_urg_2pm = tot_urg_5pm = 0

    sorted_branch_tuples = []
    for h in all_handles:
        z = resolve_branch_zone(h, df_detail)
        sorted_branch_tuples.append((z, h))

    sorted_branch_tuples.sort(key=lambda x: (x[0], x[1]))

    for z, h in sorted_branch_tuples:
        u9 = h_9am.get(h, {}).get("urgent", 0)
        u2 = h_2pm.get(h, {}).get("urgent", 0) if has_2pm else 0
        u5 = h_5pm.get(h, {}).get("urgent", 0) if has_5pm else 0

        if has_5pm:
            urg_change = u5 - u9
            clear_pct = ((u9 - u5) / u9 * 100.0) if u9 > 0 else 100.0
        elif has_2pm:
            urg_change = u2 - u9
            clear_pct = ((u9 - u2) / u9 * 100.0) if u9 > 0 else 100.0
        else:
            urg_change = 0
            clear_pct = 0.0

        urg_change_str = f"{urg_change}" if urg_change == 0 else (f"+{urg_change}" if urg_change > 0 else f"{urg_change}")

        rows.append({
            "ZONE": z,
            "POST OFFICE HANDLE": h,
            "URGENT_9AM": u9,
            "URGENT_2PM": u2,
            "URGENT_5PM": u5,
            "URGENT_CHANGE": urg_change_str,
            "CLEARANCE_PCT": clear_pct
        })

        tot_urg_9am += u9
        tot_urg_2pm += u2
        tot_urg_5pm += u5

    if has_5pm:
        grand_change = tot_urg_5pm - tot_urg_9am
        grand_pct = ((tot_urg_9am - tot_urg_5pm) / tot_urg_9am * 100.0) if tot_urg_9am > 0 else 100.0
    elif has_2pm:
        grand_change = tot_urg_2pm - tot_urg_9am
        grand_pct = ((tot_urg_9am - tot_urg_2pm) / tot_urg_9am * 100.0) if tot_urg_9am > 0 else 100.0
    else:
        grand_change = 0
        grand_pct = 0.0

    grand_change_str = f"{grand_change}" if grand_change == 0 else (f"+{grand_change}" if grand_change > 0 else f"{grand_change}")

    totals = {
        "ZONE": "ALL ZONES",
        "POST OFFICE HANDLE": "TOTAL",
        "URGENT_9AM": tot_urg_9am,
        "URGENT_2PM": tot_urg_2pm,
        "URGENT_5PM": tot_urg_5pm,
        "URGENT_CHANGE": grand_change_str,
        "CLEARANCE_PCT": grand_pct
    }

    return rows, totals, []

def build_compare_excel(date_str, rows, totals, itemized_transitions=None, out_filepath=None):
    if out_filepath is None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        out_filepath = os.path.join(COMPARE_DIR, f"Urgent_Clearance_Compare_{stamp}.xlsx")

    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Total Urgent Shift Compare"

    font_family = "Segoe UI"
    f_title = Font(name=font_family, size=13, bold=True, color="FFFFFF")
    f_hdr = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    f_data = Font(name=font_family, size=9)
    f_total = Font(name=font_family, size=10, bold=True, color="0F172A")
    f_good = Font(name=font_family, size=9, bold=True, color="065F46")

    fill_title = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_hdr = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    fill_alt = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    fill_s1 = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid") # Blue 9AM
    font_s1 = Font(name=font_family, size=9, bold=True, color="1E40AF")
    hdr_s1  = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

    fill_s2 = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid") # Orange 2PM
    font_s2 = Font(name=font_family, size=9, bold=True, color="9A3412")
    hdr_s2  = PatternFill(start_color="C2410C", end_color="C2410C", fill_type="solid")

    fill_s3 = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Green 5PM
    font_s3 = Font(name=font_family, size=9, bold=True, color="065F46")
    hdr_s3  = PatternFill(start_color="047857", end_color="047857", fill_type="solid")

    thin = Side(border_style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws_sum.merge_cells("A1:G1")
    t_cell = ws_sum.cell(1, 1, f"TOTAL URGENT SHIFT COMPARISON REPORT — {date_str}")
    t_cell.font = f_title
    t_cell.fill = fill_title
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 30

    headers = [
        "ZONE", "POST OFFICE HANDLE",
        "URGENT (9AM)", "URGENT (2PM)", "URGENT (5PM)",
        "URGENT CHANGE", "CLEAR %"
    ]
    ws_sum.row_dimensions[2].height = 24

    for c_i, h_text in enumerate(headers, 1):
        cell = ws_sum.cell(2, c_i, h_text)
        cell.font = f_hdr

        if c_i == 3:
            cell.fill = hdr_s1 # Blue 9AM
        elif c_i == 4:
            cell.fill = hdr_s2 # Orange 2PM
        elif c_i == 5:
            cell.fill = hdr_s3 # Green 5PM
        else:
            cell.fill = fill_hdr

        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    r_idx = 3
    for r in rows:
        row_fill = fill_alt if r_idx % 2 == 0 else None
        ws_sum.row_dimensions[r_idx].height = 19

        vals = [
            r["ZONE"], r["POST OFFICE HANDLE"],
            r["URGENT_9AM"], r["URGENT_2PM"], r["URGENT_5PM"],
            r["URGENT_CHANGE"], f"{r['CLEARANCE_PCT']:.1f}%"
        ]

        for c_i, val in enumerate(vals, 1):
            cell = ws_sum.cell(r_idx, c_i, val)
            cell.border = border

            if c_i == 3:
                cell.fill = fill_s1
                cell.font = font_s1
            elif c_i == 4:
                cell.fill = fill_s2
                cell.font = font_s2
            elif c_i == 5:
                cell.fill = fill_s3
                cell.font = font_s3
            elif c_i in (6, 7):
                cell.font = f_good
                if row_fill: cell.fill = row_fill
            else:
                cell.font = f_data
                if row_fill: cell.fill = row_fill

            cell.alignment = Alignment(horizontal="center" if c_i > 2 else "left", vertical="center")
        r_idx += 1

    ws_sum.row_dimensions[r_idx].height = 22
    tot_vals = [
        totals["ZONE"], totals["POST OFFICE HANDLE"],
        totals["URGENT_9AM"], totals["URGENT_2PM"], totals["URGENT_5PM"],
        totals["URGENT_CHANGE"], f"{totals['CLEARANCE_PCT']:.1f}%"
    ]

    for c_i, val in enumerate(tot_vals, 1):
        cell = ws_sum.cell(r_idx, c_i, val)
        cell.font = f_total
        cell.fill = fill_total
        cell.border = border
        cell.alignment = Alignment(horizontal="center" if c_i > 2 else "left", vertical="center")

    for col in ws_sum.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_sum.column_dimensions[col_letter].width = max(max_len + 4, 14)

    wb.save(out_filepath)
    return out_filepath
