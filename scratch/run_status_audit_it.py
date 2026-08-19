import json
import os
import sys
import pandas as pd
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import generate_report

with open('config.json', 'r', encoding='utf-8') as f:
    cfg = json.load(f)

token = cfg["api"]["bearer_token"]
headers = {
    "Authorization": f"Bearer {token}",
    "Accept": "application/json, text/plain, */*",
    "User-Agent": "Mozilla/5.0"
}

fresh_src = 'scratch/fresh_run_test/fresh_export.xlsx'
output_dir = 'scratch/status_audit'
os.makedirs(output_dir, exist_ok=True)

print("1. Generating active report metadata...")
res = generate_report.generate_reports_from_data(
    fresh_src,
    'post_office_lookup.csv',
    output_dir,
    return_metadata=True,
    mode='wide'
)

all_hr = res['handle_results']
active_orders = []
seen_oids = set()

for hr in all_hr:
    h = hr['handle']
    zone = hr.get('zone', '')
    sections = hr.get('sections', [])
    for sec in sections:
        if isinstance(sec, tuple) and len(sec) >= 2:
            tab_name = sec[0]
            rows = sec[1]
            for r in rows:
                oid = str(r.get('ORDER ID', '') or r.get('Bill code', '')).strip()
                if oid and oid != 'nan' and oid not in seen_oids:
                    seen_oids.add(oid)
                    active_orders.append({
                        'order_id': oid,
                        'handle': h,
                        'zone': zone,
                        'report_tab': tab_name,
                        'created_date': r.get('CREATED DATE', ''),
                        'export_status': r.get('CURRENT STATUS', ''),
                        'export_current_po': r.get('CURRENT POST OFFICE', ''),
                        'delivery_po': r.get('DELIVERY POST OFFICE', ''),
                        'sender': r.get('SENDER', '') or r.get('Cus name', ''),
                        'receiver': r.get('RECEIVER', ''),
                        'cod': r.get('COD (USD)', 0),
                        'fee': r.get('TOTAL FEE (USD)', 0),
                    })

print(f"Total active orders currently counted in /total: {len(active_orders)}")

print("2. Fetching live Web TMS status for all active orders in parallel...")

def check_order_web(ord_info):
    oid = ord_info['order_id']
    url_track = "https://gw-express.metfone.com.kh/tms-tracking/api/v1/order-tracking"
    web_status_code = ""
    web_status_name = "NOT FOUND / ERROR"
    web_desc = ""
    web_updated_at = ""
    web_shipper = ""
    
    try:
        r = requests.get(url_track, params={"order_id": oid}, headers=headers, timeout=8)
        if r.status_code == 200:
            data = r.json()
            trips = data.get('trackingTrips', [])
            if trips:
                top = trips[0]
                web_status_code = str(top.get('status', '') or '').strip()
                web_status_name = str(top.get('statusName', '') or '').strip()
                web_desc = str(top.get('desc', '') or '').strip()
                web_updated_at = str(top.get('updatedAt', '') or '').strip()
                shipper = top.get('shipperName') or (top.get('updatedBy') or {}).get('name') or ''
                shipper_phone = (top.get('updatedBy') or {}).get('phone') or ''
                web_shipper = f"{shipper} ({shipper_phone})".strip() if shipper_phone else shipper
    except Exception as e:
        web_desc = f"API error: {e}"
        
    return {
        **ord_info,
        'web_status_code': web_status_code,
        'web_status_name': web_status_name,
        'web_desc': web_desc,
        'web_updated_at': web_updated_at,
        'web_shipper': web_shipper
    }

results = []
with ThreadPoolExecutor(max_workers=30) as executor:
    futures = [executor.submit(check_order_web, o) for o in active_orders]
    count = 0
    for f in as_completed(futures):
        results.append(f.result())
        count += 1
        if count % 300 == 0 or count == len(active_orders):
            print(f"  Checked {count}/{len(active_orders)} orders...")

print("\n3. Classifying discrepancies and errors...")
error_rows = []
all_rows = []

for r in results:
    oid = r['order_id']
    tab = r['report_tab']
    h = r['handle']
    z = r['zone']
    exp_st = str(r['export_status'])
    w_code = str(r['web_status_code']).upper()
    w_name = str(r['web_status_name']).upper()
    w_desc = str(r['web_desc'])
    w_time = str(r['web_updated_at'])
    w_shipper = str(r['web_shipper'])
    
    # Check if Shipped / Delivered on Web
    is_shipped_on_web = ('410' in w_code or 'SHIPPED' in w_name or 'DELIVERED' in w_name or 'DELIVERED' in w_desc.upper() or 'GIAO THÀNH CÔNG' in w_desc.upper() or 'ĐÃ GIAO' in w_desc.upper())
    is_returned_on_web = ('520' in w_code or 'RETURN' in w_name or 'ĐÃ TRẢ HÀNG' in w_desc.upper())
    
    # Clean up status code from export
    exp_code = ""
    for piece in exp_st.split(' - '):
        if piece.strip().isdigit():
            exp_code = piece.strip()
            break
            
    is_code_mismatch = False
    if w_code and exp_code:
        norm_w = w_code.lstrip('S')
        norm_e = exp_code.lstrip('S')
        if norm_w != norm_e:
            is_code_mismatch = True

    issue_type = ""
    action_it = ""
    
    if is_shipped_on_web:
        issue_type = "🔴 SHIPPED ON WEB (S410) BUT STILL IN /TOTAL"
        action_it = "IT Action: Fix export-detail sync lag. Order is already delivered on TMS Web/App, but export-detail still lists as active."
    elif is_returned_on_web:
        issue_type = "🟠 RETURNED ON WEB (S520) BUT STILL IN /TOTAL"
        action_it = "IT Action: Order return is completed on Web/App, but export-detail still lists as active."
    elif is_code_mismatch:
        issue_type = f"🟡 STATUS MISMATCH (Web: {w_code} vs Export: {exp_code})"
        action_it = f"IT Action: Realtime status mismatch. Web shows {w_code} ({w_name}) while export reports {exp_st}."
    else:
        issue_type = "🟢 Status Matched"
        action_it = "None"

    row_data = {
        'Order ID': oid,
        'Zone': z,
        'Handle Post Office': h,
        'Report Tab (/total)': tab,
        'Issue / Error Type': issue_type,
        'Action for IT': action_it,
        'Export Report Status': exp_st,
        'Web Live Status Code': w_code,
        'Web Live Status Name': w_name,
        'Web Live Description': w_desc,
        'Web Latest Update Time': w_time,
        'Web Shipper / Staff': w_shipper,
        'Created Date': r['created_date'],
        'Delivery Post Office': r['delivery_po'],
        'COD (USD)': r['cod'],
        'Fee (USD)': r['fee'],
        'Sender': r['sender'],
        'Receiver': r['receiver']
    }
    
    all_rows.append(row_data)
    if issue_type != "🟢 Status Matched":
        error_rows.append(row_data)

print(f"\nAudit complete!")
print(f"Total Active Orders Checked: {len(all_rows)}")
print(f"Total Discrepancies / Errors Found: {len(error_rows)}")

shipped_errors = [r for r in error_rows if 'SHIPPED ON WEB' in r['Issue / Error Type']]
returned_errors = [r for r in error_rows if 'RETURNED ON WEB' in r['Issue / Error Type']]
mismatch_errors = [r for r in error_rows if 'STATUS MISMATCH' in r['Issue / Error Type']]

print(f"  - Shipped on Web (S410) but stuck in report: {len(shipped_errors)}")
print(f"  - Returned on Web (S520) but stuck in report: {len(returned_errors)}")
print(f"  - Status Code Mismatch (Web vs Export)     : {len(mismatch_errors)}")

# 4. Generate formatted Excel
excel_out = 'Metfone_Order_Status_Discrepancies_IT_Report.xlsx'
wb = openpyxl.Workbook()

# Sheet 1: Error List for IT
ws_err = wb.active
ws_err.title = "IT Error List (Discrepancies)"

headers_list = list(all_rows[0].keys())
ws_err.append(headers_list)

header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

for col_idx, cell in enumerate(ws_err[1], 1):
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Add error rows
red_fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
orange_fill = PatternFill(start_color="FFEAD2", end_color="FFEAD2", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF9D2", end_color="FFF9D2", fill_type="solid")

for r in error_rows:
    row_vals = [r[k] for k in headers_list]
    ws_err.append(row_vals)
    curr_row = ws_err.max_row
    
    fill_to_use = None
    if 'SHIPPED ON WEB' in r['Issue / Error Type']:
        fill_to_use = red_fill
    elif 'RETURNED ON WEB' in r['Issue / Error Type']:
        fill_to_use = orange_fill
    elif 'STATUS MISMATCH' in r['Issue / Error Type']:
        fill_to_use = yellow_fill

    for c_idx in range(1, len(headers_list) + 1):
        c = ws_err.cell(row=curr_row, column=c_idx)
        c.font = Font(name="Calibri", size=10)
        c.border = thin_border
        if fill_to_use:
            c.fill = fill_to_use
        if headers_list[c_idx - 1] in ('Order ID', 'Zone', 'Handle Post Office', 'Report Tab (/total)', 'Web Live Status Code', 'Created Date'):
            c.alignment = Alignment(horizontal="center", vertical="center")

# Sheet 2: All Active Orders Full Audit
ws_all = wb.create_sheet(title="All Active Orders Audit")
ws_all.append(headers_list)
for col_idx, cell in enumerate(ws_all[1], 1):
    cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for r in all_rows:
    row_vals = [r[k] for k in headers_list]
    ws_all.append(row_vals)
    curr_row = ws_all.max_row
    for c_idx in range(1, len(headers_list) + 1):
        c = ws_all.cell(row=curr_row, column=c_idx)
        c.font = Font(name="Calibri", size=10)
        c.border = thin_border
        if headers_list[c_idx - 1] in ('Order ID', 'Zone', 'Handle Post Office', 'Report Tab (/total)', 'Web Live Status Code', 'Created Date'):
            c.alignment = Alignment(horizontal="center", vertical="center")

# Auto fit column widths
for ws in [ws_err, ws_all]:
    ws.row_dimensions[1].height = 28
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

wb.save(excel_out)
print(f"\nSaved formatted Excel report to: {excel_out}")
