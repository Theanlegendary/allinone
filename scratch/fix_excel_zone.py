import json
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('config.json', 'r', encoding='utf-8') as f:
    cfg = json.load(f)

zone_mapping = cfg.get("total_zones", {})

def get_zone_name(handle):
    h = str(handle).upper().strip()
    for zk, handles in zone_mapping.items():
        if h in [x.upper().strip() for x in handles]:
            return zk.upper()
    return "UNKNOWN"

excel_file = 'Metfone_Order_Status_Discrepancies_IT_Report.xlsx'
wb = openpyxl.load_workbook(excel_file)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    # Find Zone column and Handle column
    zone_col = None
    handle_col = None
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == 'Zone':
            zone_col = col_idx
        elif cell.value == 'Handle Post Office':
            handle_col = col_idx
            
    if zone_col and handle_col:
        for r_idx in range(2, ws.max_row + 1):
            h_val = ws.cell(row=r_idx, column=handle_col).value
            if h_val:
                z_name = get_zone_name(h_val)
                ws.cell(row=r_idx, column=zone_col).value = z_name

wb.save(excel_file)
print("Updated Zone values in Excel file!")
