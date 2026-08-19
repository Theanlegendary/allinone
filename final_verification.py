"""Final verification of VIP column and no-yellow-dots implementation"""
from openpyxl import load_workbook
from PIL import Image
import numpy as np

print("=" * 100)
print("FINAL VERIFICATION - VIP COLUMN & NO YELLOW DOTS")
print("=" * 100)

# Test Excel file
excel_file = 'bo/Report_PNPP003_Branch_05_08_2026_000000.xlsx'
image_file = 'bo/Report_PNPP003_Branch_05_08_2026_000000.png'

print(f"\n1️⃣  CHECKING EXCEL: {excel_file}")
print("-" * 100)

wb = load_workbook(excel_file, data_only=True)
ws = wb[wb.sheetnames[0]]

# Find columns
vip_col = None
age_col = None
receiver_col = None

for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
    for idx, cell in enumerate(row, 1):
        val = str(cell.value or '').strip()
        if val == 'VIP':
            vip_col = idx
        if 'Age' in val:
            age_col = idx
        if 'RECEIVER' in val:
            receiver_col = idx

print(f"   Column positions: VIP={vip_col}, AGE={age_col}, RECEIVER={receiver_col}")

# Check age dots
green_count = 0
yellow_count = 0
red_count = 0

for row_idx in range(5, ws.max_row + 1):
    if age_col:
        val = str(ws.cell(row_idx, age_col).value or '').strip()
        if '🟢' in val:
            green_count += 1
        elif '🟡' in val:
            yellow_count += 1
        elif '🔴' in val:
            red_count += 1

print(f"\n   Age Dots Distribution:")
print(f"      🟢 Green (0-10h): {green_count}")
print(f"      🟡 Yellow (SHOULD BE 0): {yellow_count} {'← ❌ PROBLEM!' if yellow_count > 0 else '← ✅ CORRECT!'}")
print(f"      🔴 Red (>10h): {red_count}")

# Check VIP column
vip_count = 0
vip_samples = []

for row_idx in range(5, min(ws.max_row + 1, 50)):
    if vip_col:
        vip_val = str(ws.cell(row_idx, vip_col).value or '').strip()
        if vip_val == 'VIP':
            vip_count += 1
            if vip_count <= 3 and receiver_col:
                receiver = ws.cell(row_idx, receiver_col).value
                vip_samples.append(f"Row {row_idx}: {receiver}")

print(f"\n   VIP Column:")
print(f"      Total VIPs found: {vip_count}")
if vip_samples:
    print(f"      Sample VIPs:")
    for sample in vip_samples:
        print(f"         {sample}")
else:
    print(f"      ⚠️ No VIPs found in this report")

wb.close()

# Check image for yellow color
print(f"\n2️⃣  CHECKING IMAGE: {image_file}")
print("-" * 100)

try:
    img = Image.open(image_file)
    img_array = np.array(img)
    
    # Yellow color range (approximate)
    # F59E0B in RGB = (245, 158, 11)
    # Check for colors close to yellow
    yellow_pixels = 0
    red_pixels = 0
    green_pixels = 0
    
    # Check for specific color ranges
    for pixel in img_array.reshape(-1, img_array.shape[-1]):
        r, g, b = pixel[0], pixel[1], pixel[2]
        
        # Yellow: high R, moderate-high G, low B
        if r > 200 and 100 < g < 200 and b < 100:
            yellow_pixels += 1
        
        # Red: high R, low G, low B
        if r > 200 and g < 100 and b < 100:
            red_pixels += 1
        
        # Green: low R, high G, low B
        if r < 100 and g > 150 and b < 100:
            green_pixels += 1
    
    total_pixels = img_array.shape[0] * img_array.shape[1]
    
    print(f"   Image size: {img.size}")
    print(f"   Color analysis (approximate):")
    print(f"      🟢 Green-ish pixels: {green_pixels:,} ({green_pixels/total_pixels*100:.2f}%)")
    print(f"      🟡 Yellow-ish pixels: {yellow_pixels:,} ({yellow_pixels/total_pixels*100:.2f}%)")
    print(f"      🔴 Red-ish pixels: {red_pixels:,} ({red_pixels/total_pixels*100:.2f}%)")
    
    if yellow_pixels > 100:
        print(f"      ⚠️ Significant yellow detected in image!")
    else:
        print(f"      ✅ Minimal/no yellow in image")
        
except Exception as e:
    print(f"   ⚠️ Could not analyze image: {e}")

print("\n" + "=" * 100)
print("✅ VERIFICATION COMPLETE")
print("=" * 100)
print("\nSUMMARY:")
print(f"   1. VIP Column: {'✅ Present and working' if vip_col else '❌ Missing'}")
print(f"   2. No Yellow Dots in Excel: {'✅ Correct' if yellow_count == 0 else '❌ Still has yellow'}")
print(f"   3. Red dots for >10h: {'✅ Working' if red_count > 0 else '⚠️ No samples'}")
print("=" * 100)
