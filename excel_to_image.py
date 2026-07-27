"""
excel_to_image.py
Renders an openpyxl-formatted Excel sheet to a PNG image using Pillow.
- Day columns: fixed uniform width
- Hidden/empty columns: skipped entirely
- Merged cells: top-left value shown, others blank
"""

import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

# ── Constants ──────────────────────────────────────────────────────────────────
SCALE       = 3             # Increased from 2 to 3 for HD resolution quality

FONT_SIZE   = 10 * SCALE
ROW_H       = 22 * SCALE    # px — uniform row height
PAD_X       = 8 * SCALE     # horizontal text padding

# Fixed pixel widths per column type
PX_DAY      = 34 * SCALE    # day columns "01"-"31"  — all same size
PX_ZONE     = 72 * SCALE    # ZONE
PX_GT       = 140 * SCALE   # Grand Total
PX_MIN      = 72 * SCALE    # minimum for auto-fit text columns (ORDER ID, names etc.)
PX_MAX      = 230 * SCALE   # maximum for auto-fit
PX_GAP      = 6 * SCALE     # gap/empty separator columns

BG_WHITE    = (255, 255, 255)
BORDER_COL  = (190, 190, 190)
TEXT_DEF    = (0,   0,   0)


def _hex(h):
    h = h.lstrip('#')
    if len(h) == 6:
        return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))
    return TEXT_DEF


def _cell_bg(cell):
    try:
        f = cell.fill
        if f and f.fill_type == 'solid':
            fc = f.fgColor
            if fc.type == 'rgb' and fc.rgb not in ('00000000', 'FFFFFFFF', 'FF000000'):
                return _hex(fc.rgb[-6:])
    except Exception:
        pass
    return None


def _cell_fg(cell):
    try:
        ft = cell.font
        if ft and ft.color and ft.color.type == 'rgb':
            if ft.color.rgb not in ('00000000', 'FF000000'):
                return _hex(ft.color.rgb[-6:])
    except Exception:
        pass
    return TEXT_DEF


def _cell_bold(cell):
    try:
        return bool(cell.font and cell.font.bold)
    except Exception:
        return False


def _cell_align(cell):
    try:
        a = cell.alignment
        if a and a.horizontal in ('center', 'right'):
            return a.horizontal
    except Exception:
        pass
    return 'left'


_WIN_FONTS = "C:/Windows/Fonts"

def _has_khmer(text: str) -> bool:
    """Return True if text contains any Khmer Unicode characters (U+1780–U+17FF)."""
    return any('\u1780' <= ch <= '\u17FF' for ch in text)

def _load_font(size, bold=False):
    # Try system paths first
    for name in (['arialbd.ttf', 'Arial Bold.ttf', 'DejaVuSans-Bold.ttf'] if bold
                 else ['arial.ttf', 'Arial.ttf', 'DejaVuSans.ttf']):
        for prefix in (f"{_WIN_FONTS}/", ""):
            try:
                return ImageFont.truetype(prefix + name, size)
            except Exception:
                pass
    return ImageFont.load_default()

def _load_khmer_font(size):
    """Load Khmer OS Battambang font for Khmer text."""
    for name in ['KhmerOSbattambang.ttf', 'KhmerOSsiemreap.ttf', 'KhmerOScontent.ttf', 'KhmerOS.ttf']:
        for prefix in (f"{_WIN_FONTS}/", ""):
            try:
                return ImageFont.truetype(prefix + name, size)
            except Exception:
                pass
    return _load_font(size, bold=False)

def _get_font(text: str, size: int, bold: bool = False):
    """Return Khmer font if text has Khmer chars, otherwise return normal font."""
    if _has_khmer(text):
        # Khmer fonts are slightly smaller, boost size dynamically based on scale
        return _load_khmer_font(size + int(1.5 * SCALE))
    return _load_font(size, bold)


def excel_to_image(xlsx_path: str) -> io.BytesIO:
    # ── Try Excel COM rendering first (for perfect Khmer text shaping and native styling on Windows) ──
    try:
        import win32com.client
        import time
        import os
        from PIL import ImageGrab
        
        abs_path = os.path.abspath(xlsx_path)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        wb = None
        try:
            wb = excel.Workbooks.Open(abs_path)
            ws = wb.ActiveSheet
            
            img = None
            temp_png = None
            
            # --- 1. Try High Quality Chart Export Method (Crisp HD 2.0x rendering) ---
            try:
                rng = ws.UsedRange
                excel_scale = 3.0  # Scale factor for HD rendering (higher = sharper text)
                chart_width = rng.Width * excel_scale
                chart_height = rng.Height * excel_scale
                
                # Copy as picture using xlScreen=1, xlPicture=-4147 for maximum vector detail
                rng.CopyPicture(1, -4147)
                
                # Add temporary chart
                chart_obj = ws.ChartObjects().Add(Left=rng.Left, Top=rng.Top, Width=chart_width, Height=chart_height)
                chart_obj.Activate()
                chart = chart_obj.Chart
                chart.Paste()
                
                # Scale the pasted picture shape to fill the chart and position it at top-left (0,0)
                if chart.Shapes.Count > 0:
                    shape = chart.Shapes(1)
                    shape.Left = 0
                    shape.Top = 0
                    shape.Width = chart_width
                    shape.Height = chart_height
                
                # Remove chart borders and fill to prevent extra margins/background padding
                chart.ChartArea.Format.Line.Visible = 0
                chart.ChartArea.Format.Fill.Visible = 0
                
                # Export to temp PNG
                temp_png = os.path.abspath(os.path.join(os.path.dirname(xlsx_path), f"temp_excel_hd_{int(time.time())}.png"))
                chart.Export(temp_png, "PNG")
                chart_obj.Delete()
                
                if os.path.exists(temp_png):
                    img = Image.open(temp_png)
                    # Load image fully into memory and then close file handle so we can delete it
                    img.load()
            except Exception:
                img = None
            finally:
                if temp_png and os.path.exists(temp_png):
                    try:
                        os.remove(temp_png)
                    except Exception:
                        pass
            
            # --- 2. Clipboard Fallback Method (if Chart method failed or wasn't used) ---
            if img is None:
                ws.UsedRange.CopyPicture(1, 2)
                time.sleep(0.5) # wait for clipboard
                for _ in range(5):
                    img = ImageGrab.grabclipboard()
                    if img:
                        break
                    time.sleep(0.5)
            
            if img:
                # Max dimension & aspect ratio safety check for Telegram
                w, h = img.size
                max_dim = 2400
                if w > max_dim or h > max_dim:
                    scale_factor = min(max_dim / float(w), max_dim / float(h))
                    new_w = max(1, int(w * scale_factor))
                    new_h = max(1, int(h * scale_factor))
                    img = img.resize((new_w, new_h), Image.Resampling.LANCZOS)

                w, h = img.size
                max_ratio = 18.0
                new_w, new_h = w, h
                if h > 0 and w / h > max_ratio:
                    new_h = int(w / max_ratio)
                elif w > 0 and h / w > max_ratio:
                    new_w = int(h / max_ratio)

                if (new_w, new_h) != (w, h):
                    padded_img = Image.new('RGB', (new_w, new_h), (255, 255, 255))
                    padded_img.paste(img, (0, 0))
                    img = padded_img

                buf = io.BytesIO()
                img.save(buf, format='PNG', optimize=True)
                buf.seek(0)
                return buf
        finally:
            if wb:
                wb.Close(SaveChanges=False)
            excel.Quit()
    except Exception as e:
        # Fall back to Pillow rendering if Excel COM fails or is not available
        import logging
        logging.warning("Excel COM rendering failed, falling back to Pillow: %s", e, exc_info=True)
        pass

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column
    if not max_row or not max_col:
        raise ValueError("Empty sheet")

    # ── 1. Build value grid (handle merged cells) ──────────────────────────────
    grid = [[''] * (max_col + 1) for _ in range(max_row + 1)]
    skip = set()
    for mc in ws.merged_cells.ranges:
        v = ws.cell(mc.min_row, mc.min_col).value
        grid[mc.min_row][mc.min_col] = str(v) if v is not None else ''
        for r in range(mc.min_row, mc.max_row + 1):
            for c in range(mc.min_col, mc.max_col + 1):
                if not (r == mc.min_row and c == mc.min_col):
                    skip.add((r, c))

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if (r, c) in skip:
                continue
            try:
                v = ws.cell(r, c).value
            except Exception:
                v = None
            if grid[r][c] == '':
                grid[r][c] = str(v) if v is not None else ''

    # ── 2. Classify columns ────────────────────────────────────────────────────
    # Only scan the first 5 rows to find headers — avoids data rows like
    # "Grand Total" row (which has a number before it) being misclassified
    day_ci  = set()
    zone_ci = set()
    gt_ci   = set()

    HEADER_SCAN_ROWS = min(max_row, 8)
    for r in range(1, HEADER_SCAN_ROWS + 1):
        for c in range(1, max_col + 1):
            val = grid[r][c].strip()
            # Day col: exactly 2-digit 01-31 in a header row
            # Exclude if previous or next cell in same row has content (avoids data Grand Total rows)
            if val.isdigit() and len(val) == 2 and 1 <= int(val) <= 31:
                day_ci.add(c)
            elif val == 'ZONE':
                zone_ci.add(c)
            elif val == 'Grand Total':
                # Only count as GT column if it's a header row (check row has other headers)
                row_vals = [grid[r][cc].strip() for cc in range(1, max_col + 1)]
                if any(v in ('ZONE', 'ORDER ID', 'POST OFFICE HANDLE', 'CURRENT POST OFFICE')
                       for v in row_vals):
                    gt_ci.add(c)

    # Detect empty columns (gap separators between side-by-side tables)
    empty_ci = set(
        c for c in range(1, max_col + 1)
        if not any(grid[r][c].strip() for r in range(1, max_row + 1))
    )

    # ── 3. Compute per-column pixel widths ─────────────────────────────────────
    fn   = _load_font(FONT_SIZE, bold=False)
    fn_b = _load_font(FONT_SIZE, bold=True)
    tmp  = Image.new('RGB', (1, 1))
    d    = ImageDraw.Draw(tmp)

    col_px = []   # pixel width per column (0 = skip/hidden)
    for c in range(1, max_col + 1):
        if c in empty_ci:
            col_px.append(PX_GAP)   # tiny but visible separator
            continue

        # Check if column is hidden in Excel
        letter = get_column_letter(c)
        cd = ws.column_dimensions.get(letter)
        if cd and (cd.hidden or (cd.width is not None and cd.width < 0.5)):
            col_px.append(0)        # fully hidden — skip in image too
            continue

        if c in day_ci:
            col_px.append(PX_DAY)   # ALL day columns same fixed width
        elif c in zone_ci:
            col_px.append(PX_ZONE)
        elif c in gt_ci:
            col_px.append(PX_GT)
        else:
            # Auto-fit text columns
            max_w = PX_MIN
            for r in range(1, max_row + 1):
                text = grid[r][c]
                if not text:
                    continue
                bold = _cell_bold(ws.cell(r, c))
                f = _get_font(text, FONT_SIZE, bold)
                stroke_w = 0
                if bold and _has_khmer(text):
                    # Simulate bold for Khmer OS fonts using a stroke outline
                    stroke_w = max(1, int(0.4 * SCALE))
                try:
                    bb = d.textbbox((0, 0), text, font=f, stroke_width=stroke_w)
                    w  = bb[2] - bb[0]
                except Exception:
                    w = len(text) * (FONT_SIZE - 2) + stroke_w * 2
                max_w = max(max_w, w + PAD_X * 2)
            col_px.append(min(max_w, PX_MAX))

    # Columns to actually render (skip 0-width hidden ones)
    render_cols = [c for c in range(1, max_col + 1) if col_px[c - 1] > 0]

    # ── 3.5 Compute per-row heights ────────────────────────────────────────────
    # Rows that contain any Khmer text get ROW_H_KHMER, others get ROW_H
    ROW_H_KHMER = 34 * SCALE
    ROW_H = 22 * SCALE
    row_heights = []
    for r in range(1, max_row + 1):
        has_khmer_row = any(
            _has_khmer(grid[r][c]) for c in range(1, max_col + 1)
        )
        row_heights.append(ROW_H_KHMER if has_khmer_row else ROW_H)

    # ── 4. Build canvas ────────────────────────────────────────────────────────
    total_w = sum(col_px[c - 1] for c in render_cols) + 1
    total_h = sum(row_heights) + 1
    img  = Image.new('RGB', (total_w, total_h), BG_WHITE)
    draw = ImageDraw.Draw(img)

    # ── 4.5 Pre-calculate merged cell pixel dimensions ───────────────
    skip = set()
    merged_rects = {}
    for mc in ws.merged_cells.ranges:
        mw = 0
        for cc in range(mc.min_col, mc.max_col + 1):
            if cc - 1 < len(col_px) and col_px[cc - 1] > 0:
                mw += col_px[cc - 1]
        mh = 0
        for rr in range(mc.min_row, mc.max_row + 1):
            if rr - 1 < len(row_heights):
                mh += row_heights[rr - 1]

        for cc in range(mc.min_col, mc.max_col + 1):
            if (mc.min_row, cc) != (mc.min_row, mc.min_col):
                skip.add((mc.min_row, cc))
        for r_m in range(mc.min_row + 1, mc.max_row + 1):
            for c_m in range(mc.min_col, mc.max_col + 1):
                skip.add((r_m, c_m))
        merged_rects[(mc.min_row, mc.min_col)] = (mw, mh)

    # ── 5. Draw ────────────────────────────────────────────────────────────────
    y = 0
    for r in range(1, max_row + 1):
        rh = row_heights[r - 1]   # per-row height
        x = 0
        for c in render_cols:
            cw   = col_px[c - 1]
            if (r, c) in skip:
                x += cw
                continue

            mw_mh = merged_rects.get((r, c))
            if mw_mh:
                mw, mh = mw_mh
            else:
                mw, mh = cw, rh
            text = grid[r][c]

            try:
                cell  = ws.cell(r, c)
                bg    = _cell_bg(cell) or BG_WHITE
                fg    = _cell_fg(cell)
                bold  = _cell_bold(cell)
                align = _cell_align(cell)
            except Exception:
                bg, fg, bold, align = BG_WHITE, TEXT_DEF, False, 'left'

            draw.rectangle([x, y, x + mw, y + mh], fill=bg)

            if text:
                f = _get_font(text, FONT_SIZE, bold)
                stroke_w = 0
                if bold and _has_khmer(text):
                    # Simulate bold for Khmer OS fonts using a stroke outline
                    stroke_w = max(1, int(0.4 * SCALE))
                try:
                    bb = draw.textbbox((0, 0), text, font=f, stroke_width=stroke_w)
                    tw, th = bb[2] - bb[0], bb[3] - bb[1]
                except Exception:
                    tw = len(text) * (FONT_SIZE - 2) + stroke_w * 2
                    th = FONT_SIZE + stroke_w * 2

                ty = y + (mh - th) // 2
                if align == 'center':
                    tx = x + (mw - tw) // 2
                elif align == 'right':
                    tx = x + mw - tw - PAD_X
                else:
                    tx = x + PAD_X

                if stroke_w > 0:
                    draw.text((tx, ty), text, font=f, fill=fg, stroke_width=stroke_w, stroke_fill=fg)
                else:
                    draw.text((tx, ty), text, font=f, fill=fg)

            draw.rectangle([x, y, x + mw, y + mh], outline=BORDER_COL, width=1 * SCALE)
            x += cw
        y += rh

    # ── Aspect ratio & max dimension safety check for Telegram ─────────────────
    # Telegram rejects photos with dimensions > 2500px or aspect ratio > 20 (Photo_invalid_dimensions).
    w, h = img.size
    max_dim = 2400
    if w > max_dim or h > max_dim:
        scale_factor = min(max_dim / float(w), max_dim / float(h))
        new_w = max(1, int(w * scale_factor))
        new_h = max(1, int(h * scale_factor))
        img = img.resize((new_w, new_h), Image.Resampling.LANCZOS)

    # Check aspect ratio safety
    w, h = img.size
    max_ratio = 18.0
    new_w, new_h = w, h
    if h > 0 and w / h > max_ratio:
        new_h = int(w / max_ratio)
    elif w > 0 and h / w > max_ratio:
        new_w = int(h / max_ratio)

    if (new_w, new_h) != (w, h):
        padded_img = Image.new('RGB', (new_w, new_h), BG_WHITE)
        padded_img.paste(img, (0, 0))
        img = padded_img

    buf = io.BytesIO()
    img.save(buf, format='PNG', optimize=True)
    buf.seek(0)
    return buf
