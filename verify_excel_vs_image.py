"""Verify what's actually in the Excel vs what we see in the image"""
from openpyxl import load_workbook
import glob
import os

print("=" * 100)
print("VERIFYING EXCEL FILE CONTENTS")
print("=" * 100)

# Find most recent SVAP001 Delivery file
pattern = 'bo/Report_SVAP001_Delivery*.xlsx'
files = glob.glob(pattern)

if not files:
    # Try cache or other directories
    pattern = '**/Report_*Delivery*05_08_2026*.xlsx'
    import pathlib
    files = list(pathlib.Path('.').rglob('Report_*Delivery*05_08_2026*.xlsx'))
    files = [str(f) for f in files]

if not files:
    print("❌ No Delivery Excel files found!")
    exit(1)

# Sort by modification time
files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
latest_file = files[0]

print(f"\n📄 Checking: {latest_file}")
print(f"   Modified: {os.path.getmtime(latest_file)}")

wb = load_workbook(latest_file, data_only=True)
ws = wb[wb.sheetnames[0]]

print(f"   Sheet: {ws.title}")

# Find Age column
age_col = None
for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
    for idx, cell in enumerate(row, 1):
        val = str(cell.value or '').strip()
        if 'Age' in val or 'យ៉ាលើកពលកម្ម' in val:
            age_col = idx
            print(f"   Found Age column at index: {idx}")
            break

if not age_col:
    print("❌ Age column not found!")
    wb.close()
    exit(1)

# Check EXACT age values with emojis
print(f"\n🔍 EXACT Age Cell Values:")
green_count = 0
yellow_count = 0
red_count = 0

for row_idx in range(5, min(ws.max_row + 1, 30)):
    cell = ws.cell(row_idx, age_col)
    val = str(cell.value or '').strip()
    
    if val and ('h' in val or 'm' in val):
        emoji = ''
        color = 'UNKNOWN'
        
        if '🟢' in val:
            emoji = '🟢'
            color = 'GREEN'
            green_count += 1
        elif '🟡' in val:
            emoji = '🟡'
            color = 'YELLOW ← ❌ BUG!'
            yellow_count += 1
        elif '🔴' in val:
            emoji = '🔴'
            color = 'RED'
            red_count += 1
        
        # Extract age hours
        import re
        match = re.search(r'(\d+)h', val)
        hours = int(match.group(1)) if match else 0
        
        print(f"   Row {row_idx:3d}: {emoji} {val:20s} → {color:15s} ({hours}h)")

print(f"\n📊 Summary:")
print(f"   🟢 Green: {green_count}")
print(f"   🟡 Yellow: {yellow_count} {'← ❌ PROBLEM!' if yellow_count > 0 else '← ✅ OK'}")
print(f"   🔴 Red: {red_count}")

wb.close()

print("\n" + "=" * 100)
if yellow_count > 0:
    print("❌ EXCEL FILE HAS YELLOW DOTS - Bot is generating yellow!")
    print("   This means the code is NOT being loaded correctly")
else:
    print("✅ EXCEL FILE IS CORRECT - No yellow dots")
    print("   The brown/orange color is from IMAGE RENDERING")
print("=" * 100)
