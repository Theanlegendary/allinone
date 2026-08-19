"""
/report command - generates Excel with all active bills (real statuses),
excluding done statuses (201, -99, 100, 410). Includes sort/filter.
"""
import os, re, tempfile
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

DONE_CODES = {'201', '-99', '99', '100', '410', '520'}

STATUS_INFO = {
    '110': ('Confirmed', 'Sending Store'),
    '120': ('Pickup in progress', 'Sending Store'),
    '200': ('Received', 'Sending Store'),
    '210': ('In Transit', 'Transit Vehicle'),
    '300': ('In Transit', 'Transit Vehicle'),
    '302': ('Sack completed / Accept handover', 'Sending/Receiving'),
    '306': ('Hub Transit / Forwarding', 'Hub / Receiving'),
    '309': ('Exploited-Forwarding', 'Receiving Store'),
    '310': ('Accept handover', 'Receiving Store'),
    '311': ('Assign order to sack', 'Sending Store'),
    '400': ('Assignment confirmation', 'Receiving Store'),
    '401': ('Delivering', 'Receiving Store'),
    '402': ('Delivering', 'Receiving Store'),
    '420': ('At Store for Pickup', 'Receiving Store'),
    '430': ('Redelivery', 'Receiving Store'),
    '460': ('Redelivery', 'Receiving Store'),
    '472': ('Stored / Hold', 'Receiving Store'),
    '480': ('Change Address', 'Receiving Store'),
    '500': ('Return Processing', 'Sending Store'),
    '540': ('Return Failed', 'Sending Store'),
}

def get_responsible_party(status_code, cur_po, recv_po):
    sc = str(status_code).strip()
    cur_po_upper = str(cur_po).strip().upper()
    recv_po_upper = str(recv_po).strip().upper()
    
    if sc in ('110', '120', '200', '311', '201'):
        return 'ហាងផ្ញើ / Sender Store'
    elif sc == '302':
        if cur_po_upper == recv_po_upper:
            return 'ហាងផ្ញើ / Sender Store'
        else:
            return 'ហាងទទួល / Receiver Store'
    elif sc in ('210', '300'):
        return 'អ្នកបើកបរដឹកជញ្ជូន / Transit Driver'
    elif sc == '306':
        if any(x in cur_po_upper for x in ('MEGA', 'HUB', 'DVC')):
            return 'មជ្ឈមណ្ឌលចែកចាយ / Transit Hub'
        else:
            return 'ហាងទទួល / Receiver Store'
    elif sc in ('310', '309', '400', '420', '472', '480'):
        return 'ហាងទទួល / Receiver Store'
    elif sc in ('401', '402', '430', '460'):
        return 'អ្នកដឹកជញ្ជូន / Delivery Rider'
    elif sc in ('500', '520', '540'):
        return 'ហាងផ្ញើ (ត្រឡប់វិញ) / Return Sender Store'
    return 'Unknown'



def build_report_excel(export_path, out_path, cfg):
    """Build the /report Excel from the export data."""
    xl = pd.ExcelFile(export_path)
    sheet = xl.sheet_names[0]
    for sname in xl.sheet_names:
        try:
            s = xl.parse(sname, nrows=3)
            if 'ORDER ID' in s.columns and 'CURRENT STATUS' in s.columns:
                sheet = sname
                break
        except Exception:
            continue
    df = xl.parse(sheet)

    if 'CURRENT STATUS' not in df.columns or 'ORDER ID' not in df.columns:
        raise ValueError("Export missing ORDER ID or CURRENT STATUS columns")

    df['STATUS_CODE'] = df['CURRENT STATUS'].astype(str).str.extract(r'^(-?\d+)')[0].fillna('')
    # Exclude done codes
    df = df[~df['STATUS_CODE'].isin(DONE_CODES)].copy()
    # Exclude test orders
    from generate_report import load_test_order_ids, normalize_id
    test_ids = load_test_order_ids(cfg)
    if test_ids:
        df = df[~df['ORDER ID'].astype(str).str.strip().isin(test_ids)].copy()

    if df.empty:
        raise ValueError("No active bills found after filtering.")

    # Build output columns
    rows_out = []
    for _, row in df.iterrows():
        sc = str(row.get('STATUS_CODE', '')).strip()
        oid = str(row.get('ORDER ID', '')).strip()
        if oid.endswith('.0'):
            oid = oid[:-2]
        po_handle = str(row.get('POST OFFICE HANDLE', '')).strip().upper()
        cur_po = str(row.get('CURRENT POST OFFICE', '')).strip().upper()
        recv_po = str(row.get('RECEIVE POST OFFICE', '')).strip().upper()
        created = row.get('CREATED DATE', '')
        info = STATUS_INFO.get(sc, (f'Status {sc}', 'Unknown'))
        resp = get_responsible_party(sc, cur_po, recv_po)
        rows_out.append({
            'BRANCH': po_handle,
            'CURRENT PO': cur_po,
            'ORDER ID': oid,
            'STATUS': sc,
            'STATUS NAME': info[0],
            'RESPONSIBLE': resp,
            'CREATED DATE': created,
        })

    df_out = pd.DataFrame(rows_out)
    df_out = df_out.sort_values(['BRANCH', 'STATUS', 'ORDER ID']).reset_index(drop=True)
    _write_report_excel(df_out, out_path)
    return len(df_out)


def _write_report_excel(df, out_path):
    """Write DataFrame to styled Excel with auto-filter table."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Active Bills"

    cols = list(df.columns)
    # Header
    hdr_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    hdr_font = Font(name='Segoe UI', color='FFFFFF', bold=True, size=10)
    thin = Side(style='thin', color='BFBFBF')
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    HEADER_KHMER_MAP_CMD = {
        'BRANCH': 'ប៉ុស្តិ៍ទទួលខុសត្រូវ',
        'CURRENT PO': 'ប៉ុស្តិ៍បច្ចុប្បន្ន',
        'ORDER ID': 'លេខបុង / លេខកូដបញ្ជាទិញ',
        'STATUS': 'កូដស្ថានភាព',
        'STATUS NAME': 'ឈ្មោះស្ថានភាព',
        'RESPONSIBLE': 'ការទទួលខុសត្រូវ',
        'CREATED DATE': 'កាលបរិច្ឆេទបង្កើត'
    }

    for ci, col in enumerate(cols, 1):
        translated_col = HEADER_KHMER_MAP_CMD.get(col, col)
        cell = ws.cell(1, ci, translated_col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bdr

    # Data rows
    data_font = Font(name='Segoe UI', size=10)
    for ri, row in df.iterrows():
        for ci, col in enumerate(cols, 1):
            val = row[col]
            if pd.isna(val):
                val = ''
            cell = ws.cell(ri + 2, ci, val)
            cell.font = data_font
            cell.border = bdr
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Auto-filter table
    last_row = len(df) + 1
    last_col = get_column_letter(len(cols))
    table_ref = f"A1:{last_col}{last_row}"
    tab = Table(displayName="ActiveBills", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Column widths
    widths = {
        'BRANCH': 22,
        'CURRENT PO': 20,
        'ORDER ID': 26,
        'STATUS': 12,
        'STATUS NAME': 28,
        'RESPONSIBLE': 26,
        'CREATED DATE': 20
    }
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 14)

    wb.save(out_path)
