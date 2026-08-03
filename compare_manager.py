import os
import json
import re
import glob
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COMPARE_DIR = "compare"
os.makedirs(COMPARE_DIR, exist_ok=True)

SNAPSHOT_FILE = os.path.join(COMPARE_DIR, "compare_snapshots.json")
EXCLUDE_STATUSES = {"99", "100", "201", "410", "520"}
RESOLVED_STATUSES = {"99", "100", "201", "410", "420", "520"}
EXCLUDE_KEYWORDS = ["TRAINER", "GLOBAL", "EXTERNAL", "TEST"]

CATEGORIES = ["Pickup", "Delivery", "Transit", "Not Assign"]
SHIFTS = ["9AM", "2PM", "5PM"]

PREFIX_ZONE_MAP = {
    "PNP": "Phnom Penh",
    "BAN": "Banteay Meanchey",
    "BAT": "Battambang",
    "KMP": "Kandal",
    "TAK": "Takeo",
    "KPC": "Kampong Cham",
    "KPT": "Kampot",
    "SRP": "Siem Reap",
    "KPS": "Preah Sihanouk",
    "KPE": "Kampong Speu",
    "KTH": "Kampong Thom",
    "SVR": "Svay Rieng",
    "PVG": "Prey Veng",
    "PUR": "Pursat",
    "KRT": "Kratie",
    "STU": "Stung Treng",
    "RAT": "Ratanakiri",
    "MON": "Mondulkiri",
    "PRE": "Preah Vihear",
    "ODD": "Oddar Meanchey",
    "PVI": "Pailin",
    "TKG": "Tbong Khmum",
    "KEP": "Kep",
    "KOH": "Koh Kong"
}

def clean_cache_3_times_daily():
    """Cleans up old cache and report files in compare/ directory 3 times daily."""
    try:
        if os.path.exists(SNAPSHOT_FILE):
            with open(SNAPSHOT_FILE, "r", encoding="utf-8") as f:
                snapshots = json.load(f)
            
            cutoff_date = datetime.now() - timedelta(days=3)
            cleaned = {}
            for d_str, val in snapshots.items():
                try:
                    dt = datetime.strptime(d_str, "%d/%m/%Y")
                    if dt >= cutoff_date:
                        cleaned[d_str] = val
                except Exception:
                    cleaned[d_str] = val

            with open(SNAPSHOT_FILE, "w", encoding="utf-8") as f:
                json.dump(cleaned, f, ensure_ascii=False, indent=2)

        for fpath in glob.glob(os.path.join(COMPARE_DIR, "*.xlsx")):
            try:
                mtime = datetime.fromtimestamp(os.path.getmtime(fpath))
                if datetime.now() - mtime > timedelta(hours=24):
                    os.remove(fpath)
            except Exception:
                pass
    except Exception:
        pass

def load_post_office_lookup_map():
    po_map = {}
    if os.path.exists("post_office_lookup.csv"):
        try:
            df_po = pd.read_csv("post_office_lookup.csv", encoding="utf-8-sig")
            df_po.columns = [str(c).strip().lower() for c in df_po.columns]
            if "current_post_office" in df_po.columns and "post_office_handle" in df_po.columns:
                for _, r in df_po.iterrows():
                    cur = str(r["current_post_office"]).strip().upper()
                    hnd = str(r["post_office_handle"]).strip().upper()
                    if cur and hnd and cur != "NAN" and hnd != "NAN":
                        po_map[cur] = hnd
        except Exception:
            pass
    return po_map

PO_LOOKUP_MAP = load_post_office_lookup_map()

def resolve_post_office_handle(raw_po):
    po = str(raw_po).strip().upper()
    if not po or po == "NAN":
        return None
    if any(kw in po for kw in EXCLUDE_KEYWORDS):
        return None

    global PO_LOOKUP_MAP
    if not PO_LOOKUP_MAP:
        PO_LOOKUP_MAP = load_post_office_lookup_map()

    if po in PO_LOOKUP_MAP:
        return PO_LOOKUP_MAP[po]

    if po.startswith("A") or po.startswith("S"):
        po_s = po[1:]
        if po_s in PO_LOOKUP_MAP:
            return PO_LOOKUP_MAP[po_s]

    prefix = po[:3] + "P001" if len(po) >= 3 else po
    if prefix in PO_LOOKUP_MAP.values():
        return prefix

    return po

def resolve_branch_zone(branch_code, df_detail=None):
    b_code = str(branch_code).strip().upper()

    if df_detail is not None:
        df = df_detail.copy()
        df.columns = [str(c).strip().upper() for c in df.columns]
        h_col = "POST OFFICE HANDLE" if "POST OFFICE HANDLE" in df.columns else ("CURRENT POST OFFICE" if "CURRENT POST OFFICE" in df.columns else None)
        z_col = "ZONE" if "ZONE" in df.columns else ("RECEIVE PROVINCE" if "RECEIVE PROVINCE" in df.columns else None)
        if h_col and z_col:
            m = df[df[h_col].astype(str).str.strip().str.upper() == b_code]
            if not m.empty:
                zv = str(m.iloc[0][z_col]).strip()
                if zv and zv.lower() != "nan":
                    return zv

    if os.path.exists("post_office_lookup.csv"):
        try:
            df_po = pd.read_csv("post_office_lookup.csv", encoding="utf-8-sig")
            df_po.columns = [str(c).strip().lower() for c in df_po.columns]
            if "post_office_handle" in df_po.columns and "zone" in df_po.columns:
                m = df_po[df_po["post_office_handle"].astype(str).str.strip().str.upper() == b_code]
                if not m.empty:
                    zv = str(m.iloc[0]["zone"]).strip()
                    if zv and zv.lower() != "nan":
                        return zv
        except Exception:
            pass

    prefix3 = b_code[:3]
    if prefix3 in PREFIX_ZONE_MAP:
        return PREFIX_ZONE_MAP[prefix3]

    return "Other Zone"

def load_test_bills():
    test_bills = set()
    if os.path.exists("test_bills.txt"):
        with open("test_bills.txt", "r", encoding="utf-8") as f:
            for line in f:
                b = line.strip()
                if b and not b.startswith("#"):
                    test_bills.add(b)
    return test_bills

def get_all_known_branches(df_detail=None):
    branches = set()
    if os.path.exists("post_office_lookup.csv"):
        try:
            df_po = pd.read_csv("post_office_lookup.csv", encoding="utf-8-sig")
            if "post_office_handle" in df_po.columns:
                for h in df_po["post_office_handle"].dropna().unique():
                    h_str = str(h).strip().upper()
                    if h_str and h_str != "NAN" and not any(kw in h_str for kw in EXCLUDE_KEYWORDS):
                        branches.add(h_str)
        except Exception:
            pass

    if df_detail is not None:
        df = df_detail.copy()
        df.columns = [str(c).strip().upper() for c in df.columns]
        handle_col = "POST OFFICE HANDLE" if "POST OFFICE HANDLE" in df.columns else ("CURRENT POST OFFICE" if "CURRENT POST OFFICE" in df.columns else None)
        if handle_col:
            for h in df[handle_col].dropna().unique():
                h_str = resolve_post_office_handle(h)
                if h_str:
                    branches.add(h_str)

    return sorted(list(branches))

def determine_shift(now=None):
    if now is None:
        now = datetime.now()
    hour = now.hour
    if hour < 12:
        return "9AM"
    elif hour < 16:
        return "2PM"
    else:
        return "5PM"

def load_snapshots():
    clean_cache_3_times_daily()
    if os.path.exists(SNAPSHOT_FILE):
        try:
            with open(SNAPSHOT_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_snapshots(data):
    with open(SNAPSHOT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def classify_category(row_dict):
    cat_raw = str(row_dict.get("REPORT TYPE", "") or row_dict.get("TYPE", "") or row_dict.get("_REPORT_CLASS", "") or "").upper()
    if "PICKUP" in cat_raw:
        return "Pickup"
    elif "DELIVERY" in cat_raw:
        return "Delivery"
    elif "TRANSIT" in cat_raw or "MEGA" in cat_raw:
        return "Transit"
    elif "BRANCH" in cat_raw or "NOT ASSIGN" in cat_raw:
        return "Not Assign"

    sc = str(row_dict.get("STATUS_CODE", "") or row_dict.get("STATUS", "") or "").lstrip("S").strip()
    if sc in ("200", "210", "302", "310"):
        return "Pickup"
    elif sc in ("311", "401", "402", "470", "472", "480", "500"):
        return "Delivery"
    elif sc in ("306", "309"):
        return "Transit"
    else:
        return "Not Assign"

def parse_holding_age_hours(scan_time_str):
    if not scan_time_str or str(scan_time_str).lower() == "nan":
        return 0.0
    try:
        match = re.search(r'(\d+)\s*h', str(scan_time_str), re.IGNORECASE)
        if match:
            return float(match.group(1))
        parsed_dt = pd.to_datetime(scan_time_str, dayfirst=True, format='mixed', errors='coerce')
        if pd.notna(parsed_dt):
            return (datetime.now() - parsed_dt).total_seconds() / 3600.0
    except Exception:
        pass
    return 0.0

def is_urgent_bill(row_dict):
    sc = str(row_dict.get("STATUS_CODE", "") or row_dict.get("STATUS", "") or "").lstrip("S").strip()
    if sc and " " in sc:
        sc = sc.split()[0].strip()

    if sc in ("420", "472") or sc in EXCLUDE_STATUSES:
        return False

    if "_IS_OVERDUE" in row_dict and pd.notna(row_dict["_IS_OVERDUE"]):
        return bool(row_dict["_IS_OVERDUE"])

    age_val = str(row_dict.get("AGE", "") or "")
    if "🔴" in age_val:
        return True

    cat = classify_category(row_dict)
    threshold = 48 if cat == "Transit" else 24

    scan_time = ""
    for tc in ["CURRENT TIME", "STATUS 306 AT STORE / AGENT (LAST TIME)", "STATUS 306 AT STORE / AGENT FROM HUB (FIRST TIME)", "SCAN TIME", "CREATED DATE"]:
        if tc in row_dict and pd.notna(row_dict[tc]):
            val_str = str(row_dict[tc]).strip()
            if val_str and val_str.lower() != "nan":
                scan_time = val_str
                break

    age_hours = parse_holding_age_hours(scan_time)
    return age_hours >= threshold

def extract_all_bills_map(df_detail):
    df = df_detail.copy()
    df.columns = [str(c).strip().upper() for c in df.columns]

    test_bills = load_test_bills()
    handle_col = "POST OFFICE HANDLE" if "POST OFFICE HANDLE" in df.columns else ("CURRENT POST OFFICE" if "CURRENT POST OFFICE" in df.columns else None)
    oid_col = "ORDER ID" if "ORDER ID" in df.columns else ("BILL ID" if "BILL ID" in df.columns else None)
    sc_col = "STATUS_CODE" if "STATUS_CODE" in df.columns else ("STATUS" if "STATUS" in df.columns else None)

    bills_map = {}
    if not handle_col or not oid_col or not sc_col:
        return bills_map

    for _, r in df.iterrows():
        oid = str(r.get(oid_col, "") or "").strip()
        if not oid or oid in test_bills:
            continue

        raw_h = str(r.get(handle_col, "") or "").strip()
        handle = resolve_post_office_handle(raw_h)
        if not handle:
            continue

        sc = str(r.get(sc_col, "") or "").lstrip("S").strip()
        if sc and " " in sc:
            sc = sc.split()[0].strip()

        scan_time = ""
        for tc in ["CURRENT TIME", "STATUS 306 AT STORE / AGENT (LAST TIME)", "STATUS 306 AT STORE / AGENT FROM HUB (FIRST TIME)", "SCAN TIME", "CREATED DATE"]:
            if tc in r and pd.notna(r[tc]):
                val_str = str(r[tc]).strip()
                if val_str and val_str.lower() != "nan":
                    scan_time = val_str
                    break

        cat = classify_category(r)
        is_urg = is_urgent_bill(r)

        bills_map[oid] = {
            "branch": handle,
            "category": cat,
            "status": sc,
            "scan_time": scan_time,
            "is_urgent": is_urg
        }

    return bills_map

def record_shift_snapshot(date_str, shift_name, df_detail):
    snapshots = load_snapshots()
    if date_str not in snapshots:
        snapshots[date_str] = {}

    bills_map = extract_all_bills_map(df_detail)
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    snapshots[date_str][shift_name] = {
        "captured_at": now_str,
        "bills": bills_map
    }

    baseline_map = snapshots[date_str].get("9AM", {}).get("bills", {})
    if shift_name != "9AM" and baseline_map:
        shift_resolved = {}
        for oid, base_info in baseline_map.items():
            if not base_info.get("is_urgent", False):
                continue

            curr_info = bills_map.get(oid)
            if not curr_info:
                shift_resolved[oid] = {
                    "branch": base_info["branch"],
                    "category": base_info["category"],
                    "status_9am": base_info["status"],
                    "new_status": "410",
                    "resolved_at": now_str
                }
            else:
                curr_sc = curr_info["status"]
                is_res = (curr_sc in RESOLVED_STATUSES or not curr_info.get("is_urgent", False))
                if is_res:
                    shift_resolved[oid] = {
                        "branch": base_info["branch"],
                        "category": base_info["category"],
                        "status_9am": base_info["status"],
                        "new_status": curr_sc,
                        "resolved_at": curr_info["scan_time"] if curr_info["scan_time"] else now_str
                    }

        snapshots[date_str][f"{shift_name}_resolved_urgent"] = shift_resolved

    save_snapshots(snapshots)
    return snapshots

def build_comparison_summary(date_str, df_detail=None):
    if df_detail is not None:
        shift_name = determine_shift()
        record_shift_snapshot(date_str, shift_name, df_detail)

    snapshots = load_snapshots()
    day_data = snapshots.get(date_str, {})

    s_9am = day_data.get("9AM", {}).get("bills", {})
    s_2pm = day_data.get("2PM", {}).get("bills", {})
    s_5pm = day_data.get("5PM", {}).get("bills", {})

    r_2pm = day_data.get("2PM_resolved_urgent", {})
    r_5pm = day_data.get("5PM_resolved_urgent", {})

    all_known_branches = get_all_known_branches(df_detail)

    # branch -> category -> {9AM, 2PM, 5PM, URG_9AM, URG_2PM, URG_5PM}
    branch_counts = {br: {c: {"9AM": 0, "2PM": 0, "5PM": 0} for c in CATEGORIES} for br in all_known_branches}
    urgent_counts = {br: {"URG_9AM": 0, "URG_2PM": 0, "URG_5PM": 0} for br in all_known_branches}

    # Count 9AM
    for oid, bdata in s_9am.items():
        br = bdata.get("branch", "UNKNOWN")
        cat = bdata.get("category", "Not Assign")
        if cat not in CATEGORIES: cat = "Not Assign"
        if br in branch_counts:
            branch_counts[br][cat]["9AM"] += 1
            if bdata.get("is_urgent", False):
                urgent_counts[br]["URG_9AM"] += 1

    # Count 2PM
    for oid, bdata in s_2pm.items():
        br = bdata.get("branch", "UNKNOWN")
        cat = bdata.get("category", "Not Assign")
        if cat not in CATEGORIES: cat = "Not Assign"
        if br in branch_counts:
            branch_counts[br][cat]["2PM"] += 1
            if bdata.get("is_urgent", False):
                urgent_counts[br]["URG_2PM"] += 1

    # Count 5PM
    for oid, bdata in s_5pm.items():
        br = bdata.get("branch", "UNKNOWN")
        cat = bdata.get("category", "Not Assign")
        if cat not in CATEGORIES: cat = "Not Assign"
        if br in branch_matrix:
            branch_counts[br][cat]["5PM"] += 1
            if bdata.get("is_urgent", False):
                urgent_counts[br]["URG_5PM"] += 1

    itemized_transitions = []
    for oid, base_info in s_9am.items():
        if not base_info.get("is_urgent", False):
            continue
        br = base_info.get("branch", "UNKNOWN")
        cat = base_info.get("category", "Not Assign")

        r_info_2pm = r_2pm.get(oid)
        r_info_5pm = r_5pm.get(oid)

        if r_info_2pm or r_info_5pm:
            final_info = r_info_5pm if r_info_5pm else r_info_2pm
            resolved_shift = "5PM" if r_info_5pm else "2PM"
            itemized_transitions.append({
                "BILL ID": oid,
                "ZONE": resolve_branch_zone(br, df_detail),
                "POST OFFICE HANDLE": br,
                "CATEGORY": cat,
                "STATUS 9AM": base_info.get("status", ""),
                "NEW STATUS": final_info.get("new_status", ""),
                "RESOLVED SHIFT": resolved_shift,
                "RESOLUTION TIME": final_info.get("resolved_at", "")
            })

    has_2pm = bool(s_2pm)
    has_5pm = bool(s_5pm)

    rows = []
    sorted_branch_tuples = []
    for br in branch_counts.keys():
        z = resolve_branch_zone(br, df_detail)
        sorted_branch_tuples.append((z, br))

    sorted_branch_tuples.sort(key=lambda x: (x[0], x[1]))

    tot_cols = {c: {"9AM": 0, "2PM": 0, "5PM": 0} for c in CATEGORIES}
    tot_urg = {"9AM": 0, "2PM": 0, "5PM": 0}

    for z, br in sorted_branch_tuples:
        u9 = urgent_counts[br]["URG_9AM"]
        u2 = urgent_counts[br]["URG_2PM"] if has_2pm else 0
        u5 = urgent_counts[br]["URG_5PM"] if has_5pm else 0

        t9 = sum(branch_counts[br][c]["9AM"] for c in CATEGORIES)
        t2 = sum(branch_counts[br][c]["2PM"] for c in CATEGORIES) if has_2pm else 0
        t5 = sum(branch_counts[br][c]["5PM"] for c in CATEGORIES) if has_5pm else 0

        if t9 == 0 and t2 == 0 and t5 == 0 and u9 == 0 and u2 == 0 and u5 == 0:
            continue

        if has_5pm:
            urg_change = u5 - u9
            clear_pct = ((u9 - u5) / u9 * 100.0) if u9 > 0 else 100.0
        elif has_2pm:
            urg_change = u2 - u9
            clear_pct = ((u9 - u2) / u9 * 100.0) if u9 > 0 else 100.0
        else:
            urg_change = 0
            clear_pct = 0.0

        urg_change_str = f"{urg_change}" if urg_change == 0 else (f"+{urg_change}" if urg_change > 0 else f"{urg_change}")

        row = {
            "ZONE": z,
            "POST OFFICE HANDLE": br,
            "URGENT_9AM": u9,
            "URGENT_2PM": u2,
            "URGENT_5PM": u5,
            "URGENT_CHANGE": urg_change_str,
            "CLEARANCE_PCT": clear_pct
        }

        for c in CATEGORIES:
            row[f"{c}_9AM"] = branch_counts[br][c]["9AM"]
            row[f"{c}_2PM"] = branch_counts[br][c]["2PM"] if has_2pm else 0
            row[f"{c}_5PM"] = branch_counts[br][c]["5PM"] if has_5pm else 0
            tot_cols[c]["9AM"] += branch_counts[br][c]["9AM"]
            tot_cols[c]["2PM"] += (branch_counts[br][c]["2PM"] if has_2pm else 0)
            tot_cols[c]["5PM"] += (branch_counts[br][c]["5PM"] if has_5pm else 0)

        rows.append(row)

        tot_urg["9AM"] += u9
        tot_urg["2PM"] += u2
        tot_urg["5PM"] += u5

    if has_5pm:
        grand_urg_change = tot_urg["5PM"] - tot_urg["9AM"]
        grand_clear_pct = ((tot_urg["9AM"] - tot_urg["5PM"]) / tot_urg["9AM"] * 100.0) if tot_urg["9AM"] > 0 else 100.0
    elif has_2pm:
        grand_urg_change = tot_urg["2PM"] - tot_urg["9AM"]
        grand_clear_pct = ((tot_urg["9AM"] - tot_urg["2PM"]) / tot_urg["9AM"] * 100.0) if tot_urg["9AM"] > 0 else 100.0
    else:
        grand_urg_change = 0
        grand_clear_pct = 0.0

    grand_urg_change_str = f"{grand_urg_change}" if grand_urg_change == 0 else (f"+{grand_urg_change}" if grand_urg_change > 0 else f"{grand_urg_change}")
    grand_clear_pct = ((tot_urg["9AM"] - tot_urg["5PM"]) / tot_urg["9AM"] * 100.0) if tot_urg["9AM"] > 0 else 100.0

    totals = {
        "ZONE": "ALL ZONES",
        "POST OFFICE HANDLE": "TOTAL",
        "URGENT_9AM": tot_urg["9AM"],
        "URGENT_2PM": tot_urg["2PM"],
        "URGENT_5PM": tot_urg["5PM"],
        "URGENT_CHANGE": grand_urg_change_str,
        "CLEARANCE_PCT": grand_clear_pct
    }

    for c in CATEGORIES:
        totals[f"{c}_9AM"] = tot_cols[c]["9AM"]
        totals[f"{c}_2PM"] = tot_cols[c]["2PM"]
        totals[f"{c}_5PM"] = tot_cols[c]["5PM"]

    return rows, totals, itemized_transitions

def build_compare_excel(date_str, rows, totals, itemized_transitions, out_filepath=None):
    if out_filepath is None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        out_filepath = os.path.join(COMPARE_DIR, f"Urgent_Clearance_Compare_{stamp}.xlsx")

    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Urgent Shift Clearance Matrix"

    ws_det = wb.create_sheet(title="Itemized Resolved Urgent Bills")

    font_family = "Segoe UI"
    f_title = Font(name=font_family, size=13, bold=True, color="FFFFFF")
    f_cat = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    f_hdr = Font(name=font_family, size=9, bold=True, color="FFFFFF")
    f_data = Font(name=font_family, size=9)
    f_total = Font(name=font_family, size=10, bold=True, color="0F172A")
    f_good = Font(name=font_family, size=9, bold=True, color="065F46")

    fill_title = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_cat1 = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_cat2 = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
    fill_cat3 = PatternFill(start_color="9A3412", end_color="9A3412", fill_type="solid")
    fill_cat4 = PatternFill(start_color="581C87", end_color="581C87", fill_type="solid")
    fill_hdr = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    fill_alt = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    fill_s1 = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid") # Blue 9AM
    font_s1 = Font(name=font_family, size=9, bold=True, color="1E40AF")
    hdr_s1  = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

    fill_s2 = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid") # Orange 2PM
    font_s2 = Font(name=font_family, size=9, bold=True, color="9A3412")
    hdr_s2  = PatternFill(start_color="C2410C", end_color="C2410C", fill_type="solid")

    fill_s3 = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Green 5PM
    font_s3 = Font(name=font_family, size=9, bold=True, color="065F46")
    hdr_s3  = PatternFill(start_color="047857", end_color="047857", fill_type="solid")

    cat_fills = [fill_cat1, fill_cat2, fill_cat3, fill_cat4]
    thin = Side(border_style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws_sum.merge_cells("A1:S1")
    t_cell = ws_sum.cell(1, 1, f"EXPRESS URGENT SHIFT CLEARANCE MATRIX (P1..3, D1..3, T1..3, N1..3, URGENT) — {date_str}")
    t_cell.font = f_title
    t_cell.fill = fill_title
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 30

    ws_sum.row_dimensions[2].height = 22
    ws_sum.merge_cells("A2:B2")
    z_cell = ws_sum.cell(2, 1, "ZONE & POST OFFICE HANDLE")
    z_cell.fill = fill_title
    z_cell.font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    z_cell.alignment = Alignment(horizontal="center", vertical="center")

    col_idx = 3
    cat_codes = [("PICKUP", "P"), ("DELIVERY", "D"), ("TRANSIT", "T"), ("NOT ASSIGN", "N")]
    for idx, (cat_name, code) in enumerate(cat_codes):
        ws_sum.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx+2)
        c_cell = ws_sum.cell(2, col_idx, f"{cat_name} ({code}1..3)")
        c_cell.font = f_cat
        c_cell.fill = cat_fills[idx % len(cat_fills)]
        c_cell.alignment = Alignment(horizontal="center", vertical="center")
        col_idx += 3

    ws_sum.merge_cells("O2:Q2")
    tot_cell = ws_sum.cell(2, 15, "URGENT COLUMN (9AM | 2PM | 5PM)")
    tot_cell.font = f_cat
    tot_cell.fill = fill_hdr
    tot_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_sum.cell(2, 18, "URGENT Δ").fill = fill_hdr
    ws_sum.cell(2, 18).font = f_cat
    ws_sum.cell(2, 18).alignment = Alignment(horizontal="center", vertical="center")

    ws_sum.cell(2, 19, "CLEARANCE").fill = fill_hdr
    ws_sum.cell(2, 19).font = f_cat
    ws_sum.cell(2, 19).alignment = Alignment(horizontal="center", vertical="center")

    sub_headers = [
        "ZONE", "POST OFFICE HANDLE",
        "P1 (9AM)", "P2 (2PM)", "P3 (5PM)",
        "D1 (9AM)", "D2 (2PM)", "D3 (5PM)",
        "T1 (9AM)", "T2 (2PM)", "T3 (5PM)",
        "N1 (9AM)", "N2 (2PM)", "N3 (5PM)",
        "URGENT (9AM)", "URGENT (2PM)", "URGENT (5PM)",
        "URGENT CHANGE", "CLEAR %"
    ]
    ws_sum.row_dimensions[3].height = 20

    for c_i, sh_text in enumerate(sub_headers, 1):
        cell = ws_sum.cell(3, c_i, sh_text)
        cell.font = Font(name=font_family, size=9, bold=True, color="FFFFFF")

        if c_i in (3, 6, 9, 12, 15):
            cell.fill = hdr_s1 # Blue 9AM
        elif c_i in (4, 7, 10, 13, 16):
            cell.fill = hdr_s2 # Orange 2PM
        elif c_i in (5, 8, 11, 14, 17):
            cell.fill = hdr_s3 # Green 5PM
        else:
            cell.fill = fill_hdr

        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    r_idx = 4
    for r in rows:
        row_fill = fill_alt if r_idx % 2 == 0 else None
        ws_sum.row_dimensions[r_idx].height = 19

        vals = [r["ZONE"], r["POST OFFICE HANDLE"]]
        for cat in CATEGORIES:
            vals.extend([r[f"{cat}_9AM"], r[f"{cat}_2PM"], r[f"{cat}_5PM"]])
        vals.extend([r["URGENT_9AM"], r["URGENT_2PM"], r["URGENT_5PM"], r["URGENT_CHANGE"], f"{r['CLEARANCE_PCT']:.1f}%"])

        for c_i, val in enumerate(vals, 1):
            cell = ws_sum.cell(r_idx, c_i, val)
            cell.border = border

            if c_i in (3, 6, 9, 12, 15):
                cell.fill = fill_s1
                cell.font = font_s1
            elif c_i in (4, 7, 10, 13, 16):
                cell.fill = fill_s2
                cell.font = font_s2
            elif c_i in (5, 8, 11, 14, 17):
                cell.fill = fill_s3
                cell.font = font_s3
            elif c_i in (18, 19):
                cell.font = f_good
                if row_fill: cell.fill = row_fill
            else:
                cell.font = f_data
                if row_fill: cell.fill = row_fill

            cell.alignment = Alignment(horizontal="center" if c_i > 2 else "left", vertical="center")
        r_idx += 1

    ws_sum.row_dimensions[r_idx].height = 22
    tot_vals = [
        totals["ZONE"], totals["POST OFFICE HANDLE"]
    ]
    for cat in CATEGORIES:
        tot_vals.extend([totals[f"{cat}_9AM"], totals[f"{cat}_2PM"], totals[f"{cat}_5PM"]])
    tot_vals.extend([totals["URGENT_9AM"], totals["URGENT_2PM"], totals["URGENT_5PM"], totals["URGENT_CHANGE"], f"{totals['CLEARANCE_PCT']:.1f}%"])

    for c_i, val in enumerate(tot_vals, 1):
        cell = ws_sum.cell(r_idx, c_i, val)
        cell.font = f_total
        cell.fill = fill_total
        cell.border = border
        cell.alignment = Alignment(horizontal="center" if c_i > 2 else "left", vertical="center")

    for col in ws_sum.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_sum.column_dimensions[col_letter].width = max(max_len + 4, 13)

    det_headers = ["BILL ID", "ZONE", "POST OFFICE HANDLE", "CATEGORY", "STATUS 9AM", "NEW STATUS", "RESOLVED SHIFT", "RESOLUTION TIME"]
    ws_det.merge_cells("A1:H1")
    d_cell = ws_det.cell(1, 1, f"ITEMIZED RESOLVED URGENT BILLS ({date_str})")
    d_cell.font = f_title
    d_cell.fill = fill_title
    d_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_det.row_dimensions[1].height = 30

    ws_det.row_dimensions[3].height = 24
    for c_idx, h_text in enumerate(det_headers, 1):
        cell = ws_det.cell(3, c_idx, h_text)
        cell.font = Font(name=font_family, size=9, bold=True, color="FFFFFF")
        cell.fill = fill_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    dr_idx = 4
    for item in itemized_transitions:
        row_fill = fill_alt if dr_idx % 2 == 0 else None
        ws_det.row_dimensions[dr_idx].height = 19
        vals = [item["BILL ID"], item["ZONE"], item["POST OFFICE HANDLE"], item["CATEGORY"], item["STATUS 9AM"], item["NEW STATUS"], item["RESOLVED SHIFT"], item["RESOLUTION TIME"]]
        for c_idx, val in enumerate(vals, 1):
            cell = ws_det.cell(dr_idx, c_idx, val)
            cell.font = f_good if c_idx == 6 else f_data
            cell.border = border
            if row_fill: cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center" if c_idx not in (1, 2, 3) else "left", vertical="center")
        dr_idx += 1

    for col in ws_det.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_det.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(out_filepath)
    return out_filepath
