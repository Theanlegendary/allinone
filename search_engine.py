"""
search_engine.py — Advanced Order & Phone Search Engine
Indexes all order data, senders, receivers, phone numbers, and branches into SQLite database (`search_database.db`).
Usage:
  python search_engine.py index             # Index/update database from all data files
  python search_engine.py <query>           # Search by phone, name, order ID, branch
  python search_engine.py <query> --branch KANP001  # Search within a specific branch
"""

import os
import sys
import glob
import re
import json
import sqlite3
from datetime import datetime
import pandas as pd

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ('utf-8', 'utf_8'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "search_database.db")

def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS orders (
            order_id TEXT PRIMARY KEY,
            created_date TEXT,
            sender_raw TEXT,
            sender_name TEXT,
            sender_phone TEXT,
            receiver_raw TEXT,
            receiver_name TEXT,
            receiver_phone TEXT,
            receive_po TEXT,
            delivery_po TEXT,
            current_po TEXT,
            post_office_handle TEXT,
            zone TEXT,
            status_code TEXT,
            current_status TEXT,
            current_time TEXT,
            cod REAL,
            total_fee REAL,
            vip TEXT,
            source_file TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_sender_phone ON orders(sender_phone)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_receiver_phone ON orders(receiver_phone)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_po_handle ON orders(post_office_handle)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_current_po ON orders(current_po)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_delivery_po ON orders(delivery_po)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_receive_po ON orders(receive_po)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_status ON orders(status_code)")
    conn.commit()
    conn.close()

def clean_phone(phone_str):
    if not phone_str or pd.isna(phone_str):
        return ""
    digits = re.sub(r"\D", "", str(phone_str))
    if digits.startswith("855") and len(digits) > 8:
        digits = "0" + digits[3:]
    elif len(digits) in (8, 9) and not digits.startswith("0"):
        digits = "0" + digits
    return digits

def parse_contact(raw_str):
    raw = str(raw_str or "").strip()
    phone, name = "", raw
    if " - " in raw:
        parts = raw.split(" - ", 1)
        phone = parts[0].strip()
        name = parts[1].strip()
    elif " " in raw and re.match(r"^\+?\d[\d\s\-]+$", raw.split()[0]):
        parts = raw.split(" ", 1)
        phone = parts[0].strip()
        name = parts[1].strip()
    return raw, name, clean_phone(phone)

def build_index():
    init_db()
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    base_dir = os.path.dirname(os.path.abspath(__file__))
    excel_files = glob.glob(os.path.join(base_dir, "*.xlsx")) + glob.glob(os.path.join(base_dir, "cache", "*.xlsx"))

    # Load VIP database for VIP tagging
    vip_phones = set()
    vip_names = set()
    vip_file = os.path.join(base_dir, "VIP_Phone_Numbers_FORMATTED_20260805_1006.xlsx")
    if os.path.exists(vip_file):
        try:
            df_vip = pd.read_excel(vip_file, sheet_name="Found (Phone Numbers)")
            for _, r in df_vip.iterrows():
                p = clean_phone(r.get("Phone Number"))
                n = str(r.get("VIP Name (Search)", "")).strip().lower()
                if p: vip_phones.add(p)
                if n: vip_names.add(n)
        except Exception:
            pass

    import openpyxl

    total_indexed = 0

    for fpath in excel_files:
        fname = os.path.basename(fpath)
        if "VIP_Phone_Numbers" in fname:
            continue
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True)
            for sname in wb.sheetnames:
                sheet = wb[sname]
                headers = []
                for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    if idx == 0:
                        headers = [str(c).strip().upper() if c is not None else "" for c in row]
                        continue
                    if not any(row):
                        continue

                    row_dict = {}
                    for h, v in zip(headers, row):
                        if h:
                            row_dict[h] = v

                    # Find ORDER ID
                    oid_val = next((row_dict[h] for h in headers if any(k in h for k in ("ORDER ID", "WAYBILL", "BILL NO")) and row_dict[h] is not None), None)
                    if not oid_val:
                        continue

                    oid = str(oid_val).strip()
                    if not oid or oid.lower() in ("nan", "none", "order id"):
                        continue
                    if oid.endswith(".0"):
                        oid = oid[:-2]

                    c_date = str(row_dict.get("CREATED DATE", row_dict.get("CREATED_DATE", "")))
                    s_raw, s_name, s_phone = parse_contact(row_dict.get("SENDER", row_dict.get("CUS NAME", "")))
                    r_raw, r_name, r_phone = parse_contact(row_dict.get("RECEIVER", ""))

                    rec_po = str(row_dict.get("RECEIVE POST OFFICE", row_dict.get("RECEIVE_POST_OFFICE", ""))).strip()
                    del_po = str(row_dict.get("DELIVERY POST OFFICE", row_dict.get("DELIVERY_POST_OFFICE", ""))).strip()
                    cur_po = str(row_dict.get("CURRENT POST OFFICE", row_dict.get("CURRENT_POST_OFFICE", ""))).strip()
                    po_handle = str(row_dict.get("POST OFFICE HANDLE", row_dict.get("POST_OFFICE_HANDLE", cur_po or del_po))).strip()
                    zone = str(row_dict.get("ZONE", "")).strip()

                    status_raw = str(row_dict.get("CURRENT STATUS", row_dict.get("CURRENT_STATUS", ""))).strip()
                    status_code = str(row_dict.get("STATUS_CODE", "")).strip()
                    if not status_code and status_raw:
                        match = re.match(r"^(\d{3})", status_raw)
                        if match:
                            status_code = match.group(1)

                    c_time = str(row_dict.get("CURRENT TIME", row_dict.get("CURRENT_TIME", "")))
                    cod_val = 0.0
                    try:
                        raw_cod = str(row_dict.get("COD (USD)", row_dict.get("COD", 0.0)) or 0.0).replace(",", "").strip()
                        cod_val = float(raw_cod) if raw_cod else 0.0
                    except Exception:
                        pass

                    fee_val = 0.0
                    try:
                        raw_fee = str(row_dict.get("TOTAL FEE (USD)", row_dict.get("TOTAL_FEE", 0.0)) or 0.0).replace(",", "").strip()
                        fee_val = float(raw_fee) if raw_fee else 0.0
                    except Exception:
                        pass

                    # Determine VIP status (Sender or Receiver match)
                    is_vip = ""
                    if s_phone in vip_phones or r_phone in vip_phones:
                        is_vip = "VIP"
                    elif s_name and any(vn in s_name.lower() for vn in vip_names if len(vn) >= 3):
                        is_vip = "VIP"
                    elif r_name and any(vn in r_name.lower() for vn in vip_names if len(vn) >= 3):
                        is_vip = "VIP"

                    cursor.execute("""
                        INSERT INTO orders (
                            order_id, created_date, sender_raw, sender_name, sender_phone,
                            receiver_raw, receiver_name, receiver_phone, receive_po, delivery_po,
                            current_po, post_office_handle, zone, status_code, current_status,
                            current_time, cod, total_fee, vip, source_file
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ON CONFLICT(order_id) DO UPDATE SET
                            current_status=excluded.current_status,
                            status_code=excluded.status_code,
                            current_time=excluded.current_time,
                            current_po=excluded.current_po,
                            post_office_handle=excluded.post_office_handle,
                            vip=excluded.vip,
                            source_file=excluded.source_file,
                            updated_at=CURRENT_TIMESTAMP
                    """, (
                        oid, c_date, s_raw, s_name, s_phone,
                        r_raw, r_name, r_phone, rec_po, del_po,
                        cur_po, po_handle, zone, status_code, status_raw,
                        c_time, cod_val, fee_val, is_vip, fname
                    ))
                    total_indexed += 1
            wb.close()
        except Exception as e:
            print(f"Warning indexing {fname}: {e}")

    conn.commit()
    conn.close()
    print(f"✅ Indexing Complete! {total_indexed} order records processed into SQLite database.")

def search_orders(query="", branch=None, role="all", status_code=None, limit=50):
    init_db()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    q_clean = query.strip()
    phone_clean = clean_phone(q_clean)

    sql = "SELECT * FROM orders WHERE 1=1"
    params = []

    role_str = (role or "all").lower()

    if status_code:
        sql += " AND status_code = ?"
        params.append(str(status_code).strip())
    elif q_clean and len(q_clean) == 3 and q_clean.isdigit():
        sql += " AND (status_code = ? OR order_id LIKE ?)"
        params.extend([q_clean, f"%{q_clean}%"])
    elif q_clean:
        if phone_clean and len(phone_clean) >= 6:
            if role_str in ("sender", "store"):
                sql += " AND (sender_phone LIKE ? OR order_id LIKE ?)"
                params.extend([f"%{phone_clean}%", f"%{q_clean}%"])
            elif role_str in ("receiver", "customer", "cus"):
                sql += " AND (receiver_phone LIKE ? OR order_id LIKE ?)"
                params.extend([f"%{phone_clean}%", f"%{q_clean}%"])
            else:
                sql += " AND (sender_phone LIKE ? OR receiver_phone LIKE ? OR order_id LIKE ?)"
                params.extend([f"%{phone_clean}%", f"%{phone_clean}%", f"%{q_clean}%"])
        else:
            p_like = f"%{q_clean}%"
            if role_str in ("sender", "store"):
                sql += " AND (order_id LIKE ? OR sender_name LIKE ? OR sender_phone LIKE ? OR post_office_handle LIKE ? OR current_po LIKE ?)"
                params.extend([p_like, p_like, p_like, p_like, p_like])
            elif role_str in ("receiver", "customer", "cus"):
                sql += " AND (order_id LIKE ? OR receiver_name LIKE ? OR receiver_phone LIKE ? OR post_office_handle LIKE ? OR current_po LIKE ?)"
                params.extend([p_like, p_like, p_like, p_like, p_like])
            else:
                sql += " AND (order_id LIKE ? OR sender_name LIKE ? OR receiver_name LIKE ? OR sender_phone LIKE ? OR receiver_phone LIKE ? OR post_office_handle LIKE ? OR current_po LIKE ?)"
                params.extend([p_like, p_like, p_like, p_like, p_like, p_like, p_like])

    if branch:
        sql += " AND (post_office_handle LIKE ? OR current_po LIKE ? OR delivery_po LIKE ? OR receive_po LIKE ?)"
        b_like = f"%{branch.strip()}%"
        params.extend([b_like, b_like, b_like, b_like])

    sql += " ORDER BY updated_at DESC, created_date DESC LIMIT ?"
    params.append(limit)

    cursor.execute(sql, params)
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]

def main():
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python search_engine.py index")
        print("  python search_engine.py <phone/name/order_id/branch>")
        print("  python search_engine.py <query> --branch KANP001")
        sys.exit(0)

    arg = sys.argv[1].strip()
    if arg == "index":
        print("Indexing all order & phone data into SQLite database...")
        build_index()
        return

    branch_filter = None
    if "--branch" in sys.argv:
        idx = sys.argv.index("--branch")
        if idx + 1 < len(sys.argv):
            branch_filter = sys.argv[idx + 1]

    role_filter = "all"
    if "--sender" in sys.argv or "--store" in sys.argv:
        role_filter = "sender"
    elif "--receiver" in sys.argv or "--customer" in sys.argv:
        role_filter = "receiver"

    status_filter = None
    if "--status" in sys.argv:
        idx = sys.argv.index("--status")
        if idx + 1 < len(sys.argv):
            status_filter = sys.argv[idx + 1]

    results = search_orders(arg, branch=branch_filter, role=role_filter, status_code=status_filter)
    role_desc = f" ({role_filter.upper()} ONLY)" if role_filter != "all" else ""
    st_desc = f" (Status: {status_filter})" if status_filter else ""
    print(f"\n🔍 SEARCH RESULTS FOR '{arg}'{role_desc}{st_desc}" + (f" (Branch: {branch_filter})" if branch_filter else "") + f" — Found {len(results)} matches:\n")
    print("=" * 110)
    for r in results:
        vip_tag = " [VIP 🌟]" if r["vip"] == "VIP" else ""
        print(f"📦 ORDER ID: {r['order_id']}{vip_tag}")
        print(f"   👤 SENDER (STORE)   : {r['sender_name']} (Phone: {r['sender_phone'] or 'N/A'})")
        print(f"   📥 RECEIVER (CUS)   : {r['receiver_name']} (Phone: {r['receiver_phone'] or 'N/A'})")
        print(f"   📍 BRANCHES         : Receive={r['receive_po']} | Current={r['current_po']} | Delivery={r['delivery_po']}")
        print(f"   📊 STATUS           : {r['current_status']} (COD: ${r['cod']:.2f})")
        print("-" * 110)

if __name__ == "__main__":
    main()
