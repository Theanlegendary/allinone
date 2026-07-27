"""mega_detail.py - Executive CEO-Level Detail Excel for MEGA / DVC Hubs.

Features:
  1. Executive KPI Summary Dashboard Cards (Total Orders, Urgent SLA, Normal SLA, Total COD Value).
  2. 2 Dedicated Sheet Tabs:
       Sheet 1: ⚠ Urgent (>=1 Day Hold at MEGA Hub)
       Sheet 2: ✓ Normal (<1 Day Hold at MEGA Hub)
  3. Executive Styling: Dark Slate headers, zebra striping, formatted currency ($#,##0.00), bold order IDs, gridlines.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

STATUS_KM = {
    "110": "មិនទាន់បញ្ជូន",
    "120": "កំពុងប្រមូល",
    "200": "បានទទួល",
    "210": "កំពុងដឹក",
    "300": "កំពុងដឹក",
    "302": "បែងចែក",
    "306": "ដល់មណ្ឌល",
    "309": "ស្កែន",
    "310": "បានទទួល",
    "311": "បន្តដឹក",
    "400": "រង់ចាំ",
    "401": "ដឹកជូន",
    "402": "ដឹកជូន",
    "420": "រង់នៅហាង",
    "430": "ដឹកឡើងវិញ",
    "460": "ដឹកឡើងវិញ",
    "472": "រក្សាទុក",
    "480": "កែអាសយដ្ឋាន",
    "500": "ត្រឡប់",
    "520": "បានត្រឡប់",
    "540": "ត្រឡប់បរាជ័យ",
}

# Column indices in source Excel (0-based)
CI_CREATED     = 1   # CREATED DATE
CI_ORDER       = 2   # ORDER ID
CI_SENDER      = 3   # SENDER
CI_RECEIVER    = 4   # RECEIVER
CI_RECV_PO     = 11  # RECEIVE POST OFFICE  (origin)
CI_DELIV_PO    = 13  # DELIVERY POST OFFICE (destination)
CI_CURRENT_PO  = 15  # CURRENT POST OFFICE
CI_TOTAL_FEE   = 19  # TOTAL FEE (USD)
CI_COD         = 20  # COD (USD)
CI_STATUS      = 23  # CURRENT STATUS
CI_ACTION_USER = 25  # ACTION USER (Staff ID & Name)
CI_HUB_CODE    = 35  # ACTION POST OFFICE at ORIGIN HUB (MEGA1 / DVCMEGA1)


def _sc(val):
    if val is None:
        return ""
    s = str(val).strip()
    if " - " in s:
        s = s.split(" - ", 1)[0]
    return s.split()[0].strip() if s else ""


def _parse_created(val, dt_cls):
    if val is None:
        return None
    if isinstance(val, dt_cls):
        return val.date()
    s = str(val).strip().split(" ")[0]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return dt_cls.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _fmt_money(val):
    """Return float or 0.0."""
    try:
        f = float(val)
        return round(f, 2) if f > 0 else 0.0
    except Exception:
        return 0.0


def build_mega_detail(source_path, out_path, cfg):
    """Build Executive CEO-Level detail Excel with 2 separate sheets.
    Returns (total_orders, urgent_count).
    """
    from datetime import datetime as _dt

    excl_test = cfg.get("pivot", {}).get("exclude_test", False)
    test_kw   = cfg.get("pivot", {}).get("test_keywords", ["test"])

    wb_src = openpyxl.load_workbook(source_path, data_only=True)
    ws_src = wb_src.active
    all_rows = list(ws_src.iter_rows(values_only=True))
    if not all_rows:
        return 0, 0
    data_rows = all_rows[1:]

    today = _dt.now().date()

    urgent_rows = []
    normal_rows = []

    hub_statuses = {"306", "309", "302", "311", "310"}

    for row in data_rows:
        if not row or len(row) <= CI_STATUS:
            continue
        if row[CI_ORDER] is None or str(row[CI_ORDER]).strip() == "":
            continue
        sc = _sc(row[CI_STATUS])
        if sc not in hub_statuses:
            continue
        if excl_test:
            blob = " ".join(str(row[c] or "") for c in (CI_SENDER, CI_RECEIVER)).lower()
            if any(k.lower() in blob for k in test_kw):
                continue

        # Filter matching Metfone Web Order Status Report (Select Branch = MEGA HUB)
        col35 = str(row[CI_HUB_CODE] if len(row) > CI_HUB_CODE and row[CI_HUB_CODE] else "").strip().upper()
        po    = str(row[CI_CURRENT_PO] or "").strip().upper()

        if col35 == "MEGA1" or po == "MEGA1":
            hub_label = "MEGA1"
        elif "DVC" in col35 or "DVC" in po or "MEGA" in col35 or "MEGA" in po or "HUB" in po:
            hub_label = "DVMEGA"
        else:
            continue



        # Date parsing: Use Action Date (Col 24) or Created Date
        action_val = row[24] if len(row) > 24 and row[24] else row[CI_CREATED]
        action_date = _parse_created(action_val, _dt)
        if not action_date:
            action_date = _parse_created(row[CI_CREATED], _dt)

        age = (today - action_date).days if action_date else 0
        # MEGA SLA Rule: Hold >= 1 day at MEGA hub is Urgent
        is_urgent = age >= 1

        action_user = str(row[CI_ACTION_USER] or "").strip() if len(row) > CI_ACTION_USER else ""

        fee_val = _fmt_money(row[CI_TOTAL_FEE] if len(row) > CI_TOTAL_FEE else None)
        cod_val = _fmt_money(row[CI_COD]       if len(row) > CI_COD       else None)

        record = {
            "hub":         hub_label,
            "status_km":   f"{sc} - {STATUS_KM.get(sc, sc)}" if sc in STATUS_KM else str(row[CI_STATUS] or sc),
            "action_user": action_user,
            "origin":      str(row[CI_RECV_PO]  or "").strip(),
            "dest":        str(row[CI_DELIV_PO] or "").strip(),
            "order_id":    str(row[CI_ORDER]    or "").strip(),
            "fee":         fee_val,
            "cod":         cod_val,
            "created":     str(row[CI_CREATED]  or "").strip(),
            "age":         age,
        }

        if is_urgent:
            urgent_rows.append(record)
        else:
            normal_rows.append(record)

    # Sort: Hub (MEGA1 first, DVMEGA second), then age desc, then order_id
    _hub_rank = {"MEGA1": 0, "DVMEGA": 1}
    urgent_rows.sort(key=lambda r: (_hub_rank.get(r["hub"], 9), -r["age"], r["order_id"]))
    normal_rows.sort(key=lambda r: (_hub_rank.get(r["hub"], 9), -r["age"], r["order_id"]))

    tot_count = len(urgent_rows) + len(normal_rows)
    urg_count = len(urgent_rows)
    norm_count = len(normal_rows)

    tot_cod = sum(r["cod"] for r in urgent_rows) + sum(r["cod"] for r in normal_rows)
    tot_fee = sum(r["fee"] for r in urgent_rows) + sum(r["fee"] for r in normal_rows)

    # ── Build Executive Workbook with 2 Sheets ──────────────────────────────
    wb = openpyxl.Workbook()

    # Sheet 1: Urgent
    ws_urg = wb.active
    ws_urg.title = "⚠ Urgent"

    # Sheet 2: Normal
    ws_norm = wb.create_sheet(title="✓ Normal")

    # Executive Color Palette & Styles
    thin_border   = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )
    thick_bottom = Border(bottom=Side(style="medium", color="0F172A"))

    ctr_align   = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align  = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    right_align = Alignment(horizontal="right",  vertical="center", wrap_text=False)

    # Fills
    title_fill  = PatternFill("solid", fgColor="0F172A") # Executive Deep Slate Navy
    kpi_bg_tot  = PatternFill("solid", fgColor="F1F5F9") # Soft Slate Grey Card
    kpi_bg_urg  = PatternFill("solid", fgColor="FCE8E6") # Soft Red Card
    kpi_bg_norm = PatternFill("solid", fgColor="E6F4EA") # Soft Green Card
    kpi_bg_cod  = PatternFill("solid", fgColor="EFF6FF") # Soft Blue Card

    hdr_fill_dark = PatternFill("solid", fgColor="1E293B") # Dark Slate Table Header
    urg_sec_fill  = PatternFill("solid", fgColor="991B1B") # Dark Red Banner
    norm_sec_fill = PatternFill("solid", fgColor="166534") # Dark Green Banner

    urg_row_even  = PatternFill("solid", fgColor="FFF5F5") # Soft Red Zebra Tint
    urg_row_odd   = PatternFill("solid", fgColor="FFFFFF")
    norm_row_even = PatternFill("solid", fgColor="F8FAFC") # Soft Grey Zebra Tint
    norm_row_odd  = PatternFill("solid", fgColor="FFFFFF")

    # Fonts
    title_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    kpi_lbl_font = Font(name="Segoe UI", size=8, bold=True, color="475569")
    kpi_val_tot  = Font(name="Segoe UI", size=13, bold=True, color="0F172A")
    kpi_val_urg  = Font(name="Segoe UI", size=13, bold=True, color="991B1B")
    kpi_val_norm = Font(name="Segoe UI", size=13, bold=True, color="166534")
    kpi_val_cod  = Font(name="Segoe UI", size=13, bold=True, color="1E40AF")

    hdr_font  = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    sec_font  = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=9, color="1E293B")
    id_font   = Font(name="Segoe UI", size=9, bold=True, color="1E3A8A")
    red_font  = Font(name="Segoe UI", size=9, bold=True, color="991B1B")
    gray_font = Font(name="Segoe UI", size=9, color="64748B")
    fee_font  = Font(name="Segoe UI", size=9, color="15803D")

    HDR = [
        "មណ្ឌល",         # 1 — Hub (MEGA1 / DVMEGA)
        "ស្ថានភាព",      # 2 — Status KM (e.g. 306 - Accept handover)
        "បុគ្គលិក",       # 3 — Action User
        "ប.ស.ចេញ",       # 4 — Origin PO
        "ប.ស.ទៅ",        # 5 — Dest PO
        "លេខបញ្ជា",      # 6 — Order ID
        "ថ្លៃ (USD)",    # 7 — Fee
        "COD (USD)",     # 8 — COD
        "កាលបរិច្ឆេទ",  # 9 — Created Date
        "ថ្ងៃ",          # 10 — Age
    ]
    COL_WIDTHS = [14, 28, 34, 12, 12, 18, 12, 12, 20, 8]
    NCOLS = len(HDR)

    def _build_executive_sheet(ws, rows_data, sec_title, sec_fill, is_urg):
        ws.views.sheetView[0].showGridLines = True

        # ── 1. Top Executive Banner (Row 1) ──────────────────────────────────
        stamp_str = _dt.now().strftime("%d/%m/%Y %H:%M")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
        c_title = ws.cell(1, 1, f"📊 EXECUTIVE MEGA HUB OPERATIONAL REPORT  |  GENERATED: {stamp_str}")
        c_title.fill = title_fill
        c_title.font = title_font
        c_title.alignment = left_align
        ws.row_dimensions[1].height = 26

        # ── 2. Executive KPI Summary Cards (Rows 3-4) ────────────────────────
        # Card 1: Total Orders (Cols 1-2)
        ws.merge_cells("A3:B3"); ws.merge_cells("A4:B4")
        c = ws.cell(3, 1, "TOTAL PARCELS AT HUB"); c.font = kpi_lbl_font; c.fill = kpi_bg_tot; c.alignment = ctr_align
        c = ws.cell(4, 1, tot_count); c.font = kpi_val_tot; c.fill = kpi_bg_tot; c.alignment = ctr_align

        # Card 2: Urgent Orders (Cols 3-4)
        ws.merge_cells("C3:D3"); ws.merge_cells("C4:D4")
        c = ws.cell(3, 3, "URGENT SLA (≥1 DAY HOLD)"); c.font = kpi_lbl_font; c.fill = kpi_bg_urg; c.alignment = ctr_align
        c = ws.cell(4, 3, urg_count); c.font = kpi_val_urg; c.fill = kpi_bg_urg; c.alignment = ctr_align

        # Card 3: Normal Orders (Cols 5-6)
        ws.merge_cells("E3:F3"); ws.merge_cells("E4:F4")
        c = ws.cell(3, 5, "NORMAL (<1 DAY HOLD)"); c.font = kpi_lbl_font; c.fill = kpi_bg_norm; c.alignment = ctr_align
        c = ws.cell(4, 5, norm_count); c.font = kpi_val_norm; c.fill = kpi_bg_norm; c.alignment = ctr_align

        # Card 4: Total COD Value (Cols 7-10)
        ws.merge_cells("G3:J3"); ws.merge_cells("G4:J4")
        c = ws.cell(3, 7, "TOTAL COD VALUE (USD)"); c.font = kpi_lbl_font; c.fill = kpi_bg_cod; c.alignment = ctr_align
        c = ws.cell(4, 7, round(tot_cod, 2)); c.font = kpi_val_cod; c.fill = kpi_bg_cod; c.alignment = ctr_align
        c.number_format = "$#,##0.00"

        for r_kpi in (3, 4):
            ws.row_dimensions[r_kpi].height = 18
            for c_kpi in range(1, NCOLS + 1):
                ws.cell(r_kpi, c_kpi).border = thin_border

        # ── 3. Section Banner Header (Row 6) ──────────────────────────────────
        ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=NCOLS)
        c_sec = ws.cell(6, 1, f"  {sec_title} — {len(rows_data)} RECORDS")
        c_sec.fill = sec_fill
        c_sec.font = sec_font
        c_sec.alignment = left_align
        ws.row_dimensions[6].height = 22

        # ── 4. Table Column Headers (Row 7) ──────────────────────────────────
        ws.row_dimensions[7].height = 22
        for ci, h in enumerate(HDR, 1):
            cell = ws.cell(7, ci, h)
            cell.fill = hdr_fill_dark
            cell.font = hdr_font
            cell.alignment = ctr_align
            cell.border = thin_border

        # ── 5. Data Rows (Row 8+) ─────────────────────────────────────────────
        curr = 8
        row_even = urg_row_even if is_urg else norm_row_even
        row_odd  = urg_row_odd  if is_urg else norm_row_odd

        for idx, rec in enumerate(rows_data):
            row_bg = row_even if idx % 2 == 0 else row_odd
            ws.row_dimensions[curr].height = 19

            vals = [
                rec["hub"],
                rec["status_km"],
                rec["action_user"],
                rec["origin"],
                rec["dest"],
                rec["order_id"],
                rec["fee"] if rec["fee"] > 0 else "",
                rec["cod"] if rec["cod"] > 0 else "",
                rec["created"],
                rec["age"],
            ]

            for ci, val in enumerate(vals, 1):
                cell = ws.cell(curr, ci, val)
                cell.fill = row_bg
                cell.border = thin_border

                if ci == 1:         # Hub
                    cell.font = data_font; cell.alignment = ctr_align
                elif ci == 2:       # Status
                    cell.font = data_font; cell.alignment = left_align
                elif ci == 3:       # Action User
                    cell.font = gray_font; cell.alignment = left_align
                elif ci in (4, 5):  # Origin / Dest PO
                    cell.font = data_font; cell.alignment = ctr_align
                elif ci == 6:       # Order ID
                    cell.font = id_font; cell.alignment = ctr_align
                elif ci in (7, 8):  # Fee / COD
                    cell.font = fee_font; cell.alignment = right_align
                    if val != "":
                        cell.number_format = "$#,##0.00"
                elif ci == 9:       # Created Date
                    cell.font = gray_font; cell.alignment = ctr_align
                elif ci == 10:      # Age
                    cell.font = red_font if is_urg else data_font
                    cell.alignment = ctr_align

            curr += 1

        # Auto-fit Column Widths
        for ci, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    _build_executive_sheet(ws_urg, urgent_rows, "⚠ URGENT ORDERS (HOLD TIME ≥ 1 DAY AT MEGA HUB)", urg_sec_fill, True)
    _build_executive_sheet(ws_norm, normal_rows, "✓ NORMAL ORDERS (HOLD TIME < 1 DAY AT MEGA HUB)", norm_sec_fill, False)

    wb.save(out_path)
    return tot_count, urg_count
