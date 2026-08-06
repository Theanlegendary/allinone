"""Check the most recent SVAP001 Excel file for dot colors"""
from openpyxl import load_workbook
import glob
import os

print("=" * 100)
print("CHECKING LATEST SVAP001 EXCEL FOR DOT COLORS")
print("=" * 100)

# Find most recent SVAP001 file
pattern = 'bo/Report_SVAP001*.xlsx'
files = glob.glob(pattern)

if not files:
    print("❌ No SVAP001 Excel files found!")
    exit(1)

# Sort by modification time
files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
latest_file = files[0]

print(f"\n📄 Latest file: {latest_file}")
print(f"   Modified: {os.path.getmtime(latest_file)}")

# Load and check
wb = load_workbook(latest_file, data_only=True)
ws = wb[wb.sheetnames[0]]

print(f"\n📋 Sheet: {ws.title}")

# Find Age column
age_col = None
for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
    for idx, cell in enumerate(row, 1):
        val = str(cell.value or '').strip()
        if 'Age' in val and 'ខ្មែរពលកម្ម' in val:  # Khmer text
            age_col = idx
            print(f"   Found Age column at index: {idx}")
            break

if not age_col:
    print("❌ Age column not found!")
    wb.close()
    exit(1)

# Check age values
print(f"\n🔍 Age values analysis:")
green_count = 0
yellow_count = 0
red_count = 0
samples = {'green': [], 'yellow': [], 'red': []}

for row_idx in range(5, min(ws.max_row + 1, 100)):
    cell = ws.cell(row_idx, age_col)
    val = str(cell.value or '').strip()
    
    if '🟢' in val:
        green_count += 1
        if len(samples['green']) < 3:
            samples['green'].append(f"Row {row_idx}: {val}")
    elif '🟡' in val:
        yellow_count += 1
        if len(samples['yellow']) < 3:
            samples['yellow'].append(f"Row {row_idx}: {val}")
    elif '🔴' in val:
        red_count += 1
        if len(samples['red']) < 3:
            samples['red'].append(f"Row {row_idx}: {val}")

print(f"\n   Distribution:")
print(f"      🟢 Green (0-10h): {green_count}")
print(f"      🟡 Yellow (SHOULD BE 0): {yellow_count} {'← ❌ PROBLEM!' if yellow_count > 0 else '← ✅ CORRECT!'}")
print(f"      🔴 Red (>10h): {red_count}")

if yellow_count > 0:
    print(f"\n   ⚠️  YELLOW DOTS FOUND IN EXCEL!")
    print(f"   Sample yellow values:")
    for sample in samples['yellow']:
        print(f"      {sample}")

if samples['red']:
    print(f"\n   Sample red values:")
    for sample in samples['red']:
        print(f"      {sample}")

wb.close()

print("\n" + "=" * 100)
print("CONCLUSION:")
if yellow_count > 0:
    print("   ❌ The EXCEL FILE itself contains yellow dots!")
    print("   ❌ The bot is generating yellow dots in the Excel")
    print("   ❌ Code changes were NOT loaded by the bot")
else:
    print("   ✅ The EXCEL FILE is correct (no yellow)")
    print("   ⚠️  The IMAGE renderer might be the problem")
print("=" * 100)
